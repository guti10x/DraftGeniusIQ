# Dependencias
from PyQt6.QtWidgets import QApplication, QDialog, QGridLayout, QLabel, QLineEdit, QSpinBox, QPushButton, QFileDialog, QWidget, QTextEdit, QProgressBar, QVBoxLayout, QTextEdit, QMainWindow, QStackedWidget, QHBoxLayout,QComboBox
from PyQt6.QtGui import QColor, QTextCharFormat
from PyQt6.QtCore import QMetaObject, Qt, Q_ARG
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementNotInteractableException, NoSuchElementException
from bs4 import BeautifulSoup
import requests
import openpyxl
import time
import os
import threading
import sys
from datetime import datetime
import json
import pandas as pd
import numpy as np
from unidecode import unidecode
import Levenshtein
from difflib import get_close_matches
import re
import category_encoders as ce
import seaborn as sns
import matplotlib
matplotlib.use('TkAgg')  #  'TkAgg' / 'Agg' 
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from sklearn.model_selection import train_test_split, cross_val_score, train_test_split, cross_val_score, KFold
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split, cross_validate
from sklearn.preprocessing import MinMaxScaler 
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from xgboost import XGBRegressor
import joblib
import copy


headers = {
    "X-RapidAPI-Key": "11822210cdmsha855c4a12c471b5p18100fjsn3972b17b3be8",
    "X-RapidAPI-Host": "sofascore.p.rapidapi.com"
}

#Credenciales ususario
usuario=""
contrasena=""

def select_folder(self):
    # Obtener el directorio del script de Python
    script_directory = os.path.dirname(__file__) if __file__ else os.getcwd()

    folder_path = QFileDialog.getExistingDirectory(self, "Seleccionar Carpeta", script_directory)
    if folder_path:
        # Actualizar las variables de clase con la carpeta y la ruta seleccionadas
        selected_folder = folder_path
        selected_path = folder_path

        # Actualizar el QLineEdit con la ruta seleccionada
        self.text_input.setText(selected_path)

def select_folder2(self):
    # Obtener el directorio del script de Python
    script_directory = os.path.dirname(__file__) if __file__ else os.getcwd()

    folder_path = QFileDialog.getExistingDirectory(self, "Seleccionar Carpeta", script_directory)
    if folder_path:
        # Actualizar las variables de clase con la carpeta y la ruta seleccionadas
        selected_folder = folder_path
        selected_path = folder_path

        # Actualizar el QLineEdit con la ruta seleccionada
        self.text_input2.setText(selected_path)

def select_file(self):
    # Obtener el directorio del script de Python
    script_directory = os.path.dirname(__file__)

    # Abrir el cuadro de diálogo para seleccionar un archivo
    file_path, _ = QFileDialog.getOpenFileName(self, "Seleccionar Archivo", script_directory)

    if file_path:
        # Actualizar las variables de clase con el archivo y la ruta seleccionadas
        self.selected_file = os.path.basename(file_path)
        self.selected_path = file_path

        # Actualizar el QLineEdit con la ruta seleccionada
        self.text_file_input.setText(self.selected_path)

def select_file2(self):
    # Obtener el directorio del script de Python
    script_directory = os.path.dirname(__file__)

    # Abrir el cuadro de diálogo para seleccionar un archivo
    file_path, _ = QFileDialog.getOpenFileName(self, "Seleccionar Archivo", script_directory)

    if file_path:
        # Actualizar las variables de clase con el archivo y la ruta seleccionadas
        self.selected_file = os.path.basename(file_path)
        self.selected_path = file_path

        # Actualizar el QLineEdit con la ruta seleccionada
        self.text_file2_input.setText(self.selected_path)

def realizar_login(driver):

    driver.get("https://mister.mundodeportivo.com/new-onboarding/#market")
    driver.implicitly_wait(15)

    # Consentir
    consent_button = driver.find_element(By.XPATH, '//*[@id="didomi-notice-agree-button"]')
    consent_button.click()

    # Click en "Siguiente"
    siguiente_button = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/button')
    for _ in range(4):  # Haz clic 4 veces
        siguiente_button.click()

    # Click en "Sign in with gmail"
    gmail_button = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/button[3]')
    gmail_button.click()

    # Ingresar usuario
    input_gmail = driver.find_element(By.XPATH, '//*[@id="email"]')
    input_gmail.clear()
    input_gmail.send_keys(usuario)

    # Ingresar contraseña
    input_psw = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/form/div[2]/input')
    input_psw.clear()
    input_psw.send_keys(contrasena)

    # Click en "Sign in with gmail"
    login_button = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/form/div[3]/button')
    login_button.click()

class VentanaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("DraftGeniousIQ")

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QGridLayout(self.central_widget)

        self.stacked_widget = QStackedWidget()

        self.ventana1 = squadWindow()
        self.ventana2 = marketWindow()
        self.ventana3 = scrapear_datos()
        self.ventana4 = dataset_creator()
        self.ventana5 = trainWindow()
        self.ventana6 = predictWindow()
        self.ventana7 = login()

        self.stacked_widget.addWidget(self.ventana1)
        self.stacked_widget.addWidget(self.ventana2)
        self.stacked_widget.addWidget(self.ventana3)
        self.stacked_widget.addWidget(self.ventana4)
        self.stacked_widget.addWidget(self.ventana5)
        self.stacked_widget.addWidget(self.ventana6)
        self.stacked_widget.addWidget(self.ventana7)

        self.btn_ventana1 = QPushButton("Mi plantilla")
        self.btn_ventana1.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(0))

        self.btn_ventana2 = QPushButton("Mercado")
        self.btn_ventana2.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))

        self.btn_ventana3 = QPushButton("Obtener datos de futbolistas")
        self.btn_ventana3.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(2))

        self.btn_ventana4 = QPushButton("Crear dataset")
        self.btn_ventana4.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(3))

        self.btn_ventana5 = QPushButton("Entrenar modelo")
        self.btn_ventana5.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(4))

        self.btn_ventana6 = QPushButton("Predecir")
        self.btn_ventana6.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(5))

        self.btn_ventana7 = QPushButton("Mi perfil")
        self.btn_ventana7.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(6))

        self.layout.addWidget(self.btn_ventana1, 0, 0)
        self.layout.addWidget(self.btn_ventana2, 0, 1)
        self.layout.addWidget(self.btn_ventana3, 0, 2)
        self.layout.addWidget(self.btn_ventana4, 0, 3)
        self.layout.addWidget(self.btn_ventana5, 0, 4)
        self.layout.addWidget(self.btn_ventana6, 0, 5)
        self.layout.addWidget(self.btn_ventana7, 0, 6)


        self.layout.addWidget(self.stacked_widget, 1, 0, 1, 7)


class squadWindow(QWidget):
    def __init__(self):
        super().__init__()

        #Varaible para guardar la plantilla scrapeada
        self.nombres_jugadores=[]

        # Crear un diseño principal usando QVBoxLayout
        layout = QVBoxLayout()

        # Crear un diseño de cuadrícula dentro del QVBoxLayout
        grid_layout = QGridLayout(self)

        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("MI PLANTILLA")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        grid_layout.addWidget(label_text, 0, 0)

        # LABEL SUBTÍTULO
        label_subtext = QLabel("Obten el listado de todos los jugaodres en mi plantilla de Mister Fantasy MD")
        grid_layout.addWidget(label_subtext, 1, 0, 1, 2)

        # BOTÓN PARA INICIAR LA OBTENCIÓN DE MI PLANTILLA ###########################################################
        # Crear un botón
        self.scrape_button = QPushButton("Obtener mi plantilla")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.scrape_button.clicked.connect(self.iniciar_scrapear_thread)

        # Alineación y estilos
        grid_layout.addWidget(self.scrape_button, 2, 0)
        self.scrape_button.setMaximumWidth(150)

        # VENTANA OUTPUT SCRAPER ####################################################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        grid_layout.addWidget(self.output_textedit, 3, 0, 10, 2)  # row, column, rowSpan, columnSpan

        # SELECCIONAR RUTA DONDE GUARDAR EL EXCEL OUTPUT DEL SCRAPER #################################################
        # LABEL DE TEXTO
        label_text = QLabel("Guardar mi plantilla:")
        grid_layout.addWidget(label_text, 13, 0)

        # INPUT DE TEXTO
        self.text_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input, 13, 1)

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 14, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        # BOTÓN PARA GUARDAR MI PLANTILLA ###########################################################################
        # Crear un botón
        self.save_button = QPushButton("Guardar plantilla")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.save_button.clicked.connect(self.guardar_excell)

        # Alineación
        grid_layout.addWidget(self.save_button, 15, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        self.save_button.setMinimumWidth(100)
        self.save_button.setMaximumWidth(150)
    

        # Agregar el diseño de cuadrícula al diseño principal
        layout.addLayout(grid_layout)

        # Agregar el diseño principal al widget
        self.setLayout(layout)

    def guardar_excell(self):
        if self.text_input.text()=="": 
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText("\nRuta de la carpeta donde guardar la plantilla no ha sido inicializada.\n")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return
        
        self.output_textedit.append(f"________________________________________________________________________________________")
        output_textedit = self.output_textedit
        color_azul = QColor(0, 0, 255)  # Valores RGB para azul
        formato_azul = QTextCharFormat()
        formato_azul.setForeground(color_azul)
        output_textedit.mergeCurrentCharFormat(formato_azul)
        output_textedit.insertPlainText("\nGuardando plantilla...\n")
        formato_negro = QTextCharFormat()
        formato_negro.setForeground(QColor(0, 0, 0))
        output_textedit.mergeCurrentCharFormat(formato_negro)

        if len(self.nombres_jugadores) > 0:
            # Obtener la fecha actual
            fecha_actual = datetime.now()

            # Formatear la fecha como una cadena (opcional)
            fecha_actual_str = fecha_actual.strftime("%Y-%m-%d--%H-%M-S")

            ruta_output = self.text_input.text()
            excel_file_path= ruta_output +"/mi_plantilla"+fecha_actual_str+".xlsx"
            
            # Crear un nuevo libro de Excel
            workbook = openpyxl.Workbook()

            # Seleccionar la hoja activa (por defecto, es la primera hoja)
            sheet = workbook.active

            # Iterar sobre la lista y almacenar cada elemento en una nueva fila
            for index, nombre in enumerate(self.nombres_jugadores, start=1):
                sheet.cell(row=index, column=1, value=nombre)

            # Guardar el libro de Excel
            workbook.save(excel_file_path)
            self.output_textedit.append(f"Plantilla guardada en {excel_file_path}")
        else:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText("\n¡La plantilla no se puede guardar porque no esta inicializada")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)

    def iniciar_scrapear_thread(self):
        # Crear un hilo y ejecutar la función en segundo plano
        thread = threading.Thread(target=self.scrapear_funcion)
        thread.start()

    def click_mas(self):
        # Pinchar en el botón del menu "Más"
        masMenu = self.driver.find_element(By.XPATH, '//*[@id="content"]/header/div[2]/ul/li[3]/a')

        try:
            masMenu.click()
        except (ElementNotInteractableException, NoSuchElementException):
            # Maneja la excepción y espera antes de intentar nuevamente
            self.output_textedit.append("Anuncio detectado, reiniciando driver...")
            self.driver.refresh()
            time.sleep(3) 
            masMenu.click()

    def scrapear_funcion(self):
        self.output_textedit.append(f"________________________________________________________________________________________")
        output_textedit = self.output_textedit
        color_azul = QColor(0, 0, 255)  # Valores RGB para azul
        formato_azul = QTextCharFormat()
        formato_azul.setForeground(color_azul)
        output_textedit.mergeCurrentCharFormat(formato_azul)
        output_textedit.insertPlainText("\nObteniendo plantilla...\n")
        formato_negro = QTextCharFormat()
        formato_negro.setForeground(QColor(0, 0, 0))
        output_textedit.mergeCurrentCharFormat(formato_negro)
       
        if usuario!="":
            try:
                self.driver = webdriver.Chrome()
                realizar_login(self.driver)
                time.sleep(5)

                # Espera a que se cargue la página
                self.driver.implicitly_wait(10)

                #Hacer click en el btn Jugadores con la función click_mas() para manejar errores generados por anuncios intrusiovos
                self.click_mas()

                # Encontrar el elemento div con la clase "team__squad"
                team_squad_div = self.driver.find_element(By.CLASS_NAME, 'team__squad')

                # Encontrar todos los elementos con la clase "name" dentro del div
                names_elements = team_squad_div.find_elements(By.CLASS_NAME, 'name')

                # Iterar sobre los elementos encontrados e imprimir el texto
                for name_element in names_elements:
                    name_element.click() 

                    nombre = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[2]")
                    apellido = self.driver.find_element(By.XPATH, " /html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[3]")

                    jugador = nombre.text + " " + apellido.text

                    self.output_textedit.append(jugador)
                    self.nombres_jugadores.append(jugador)
                    time.sleep(1)

                    self.driver.back()

                self.driver.quit()

            except: 
                output_textedit = self.output_textedit
                color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
                formato_rojo = QTextCharFormat()
                formato_rojo.setForeground(color_rojo)
                output_textedit.mergeCurrentCharFormat(formato_rojo)
                output_textedit.insertPlainText('Un anuncio bloquea al scraper el acceso a la información, volviendo a intentarlo...\n')
                formato_negro = QTextCharFormat()
                formato_negro.setForeground(QColor(0, 0, 0))
                output_textedit.mergeCurrentCharFormat(formato_negro)
                # Crear un nuevo hilo y volver a intentarlo 
                thread = threading.Thread(target=self.scrapear_funcion)
                thread.start()
        else:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('No has iniciado sesion en la aplicación. Logueate con tus credenciales de Mister Fantasy MD en la ventana Perfil para acceder a tu plantilla.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)

class marketWindow(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()
        
        # Crear un diseño principal usando QVBoxLayout
        layout = QVBoxLayout()

        # Crear un diseño de cuadrícula dentro del QVBoxLayout
        grid_layout = QGridLayout(self)

        #Varaible para guardar la plantilla scrapeada
        self.nombres_jugadores=[]
        
        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("MERCADO")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        grid_layout.addWidget(label_text, 0, 0)

        # LABEL SUBTÍTULO
        label_subtext = QLabel("Obten el listado de todos los jugaodres en venta en el mercado de Mister Fantasy MD")
        grid_layout.addWidget(label_subtext, 1, 0, 1, 2)

        # BOTÓN PARA INICIAR LA OBTENCIÓN DE MI PLANTILLA ###########################################################
        # Crear un botón
        self.scrape_button = QPushButton("Obtener jugaodres en venta")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.scrape_button.clicked.connect(self.iniciar_scrapear_thread)

        # Alineación y estilos
        grid_layout.addWidget(self.scrape_button, 2, 0)
        self.scrape_button.setMaximumWidth(190)

        # VENTANA OUTPUT SCRAPER #####################################################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        grid_layout.addWidget(self.output_textedit, 3, 0, 11, 2)  # row, column, rowSpan, columnSpan

        # SELECCIONAR RUTA DONDE GUARDAR EL EXCEL OUTPUT DEL SCRAPER ##################################################
        # LABEL DE TEXTO
        label_text = QLabel("Guardar jugadores en venta:")
        grid_layout.addWidget(label_text, 15, 0)

        # INPUT DE TEXTO
        self.text_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input, 15, 1)

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 16, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        # BOTÓN PARA GUARDAR MI PLANTILLA ###########################################################################
        # Crear un botón
        self.save_button = QPushButton("Guardar jugadores")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.save_button.clicked.connect(self.guardar_excell)

        # Alineación
        grid_layout.addWidget(self.save_button, 17, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        self.save_button.setMinimumWidth(100)
        self.save_button.setMaximumWidth(150)

        # Agregar el diseño de cuadrícula al diseño principal
        layout.addLayout(grid_layout)

        # Agregar el diseño principal al widget
        self.setLayout(layout)

    def guardar_excell(self):
        if self.text_input.text()=="": 
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText("\nRuta de la carpeta donde guardar los jugadores del mercado no ha sido inicializada.\n")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return
        
        self.output_textedit.append(f"________________________________________________________________________________________")
        output_textedit = self.output_textedit
        color_azul = QColor(0, 0, 255)  # Valores RGB para azul
        formato_azul = QTextCharFormat()
        formato_azul.setForeground(color_azul)
        output_textedit.mergeCurrentCharFormat(formato_azul)
        output_textedit.insertPlainText("\nGuardando jugaodres del mercado...\n")
        formato_negro = QTextCharFormat()
        formato_negro.setForeground(QColor(0, 0, 0))
        output_textedit.mergeCurrentCharFormat(formato_negro)

        if len(self.nombres_jugadores) > 0:
            # Obtener la fecha actual
            fecha_actual = datetime.now()

            # Formatear la fecha como una cadena (opcional)
            fecha_actual_str = fecha_actual.strftime("%Y-%m-%d--%H-%M-S")

            ruta_output = self.text_input.text()
            excel_file_path= ruta_output +"/mercado"+fecha_actual_str+".xlsx"
            
            # Crear un nuevo libro de Excel
            workbook = openpyxl.Workbook()

            # Seleccionar la hoja activa (por defecto, es la primera hoja)
            sheet = workbook.active

            # Iterar sobre la lista y almacenar cada elemento en una nueva fila
            for index, nombre in enumerate(self.nombres_jugadores, start=1):
                sheet.cell(row=index, column=1, value=nombre)

            # Guardar el libro de Excel
            workbook.save(excel_file_path)
            self.output_textedit.append(f"Plantilla guardada en {excel_file_path}")
        else:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText(f"\n¡Los jugaodres del mercado no se puede guardar porque no esta inicializada")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)

    def click_mas(self):
        # Pinchar en el botón del menu "Más"
        masMenu = self.driver.find_element(By.XPATH, '//*[@id="content"]/header/div[2]/ul/li[2]/a')

        try:
            masMenu.click()
        except (ElementNotInteractableException, NoSuchElementException):
            # Maneja la excepción y espera antes de intentar nuevamente
            self.output_textedit.append("Anuncio detectado, reiniciando driver...")
            self.driver.refresh()
            time.sleep(3) 
            masMenu.click()

    def iniciar_scrapear_thread(self):
        # Crear un hilo y ejecutar la función en segundo plano
        thread = threading.Thread(target=self.scrapear_funcion)
        thread.start()

    def scrapear_funcion(self):
        self.output_textedit.append(f"________________________________________________________________________________________")
        output_textedit = self.output_textedit
        color_azul = QColor(0, 0, 255)  # Valores RGB para azul
        formato_azul = QTextCharFormat()
        formato_azul.setForeground(color_azul)
        output_textedit.mergeCurrentCharFormat(formato_azul)
        output_textedit.insertPlainText("\nObteniendo jugadores en el mercado...\n")
        formato_negro = QTextCharFormat()
        formato_negro.setForeground(QColor(0, 0, 0))
        output_textedit.mergeCurrentCharFormat(formato_negro)

        if usuario!="":
            try:
                self.driver = webdriver.Chrome()
                realizar_login(self.driver)
                time.sleep(5)

                # Espera a que se cargue la página
                self.driver.implicitly_wait(10)

                #Hacer click en el btn Jugadores con la función click_mas() para manejar errores generados por anuncios intrusiovos
                self.click_mas()

                # Encuentra el elemento <ul> con el id "list-on-sale"
                ul_element = self.driver.find_element(By.ID, "list-on-sale")

                # Encuentra los elementos <div> con la clase "name" dentro del elemento <ul>
                div_elements = ul_element.find_elements(By.CSS_SELECTOR, "div.name")

                # Itera sobre los elementos <div> encontrados e imprime el nombre del jugador
                for div_element in div_elements:
                    # Obtener el contenido del elemento <div>
                    time.sleep(3)
                    try:
                        div_element.click()
                    except:
                        self.driver.execute_script("window.scrollBy(0, arguments[0]);", 300)  
                        time.sleep(0.5)
                        div_element.click()
                    time.sleep(2)

                    nombre = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[2]")
                    apellido = self.driver.find_element(By.XPATH, " /html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[3]")

                    jugador = nombre.text + " " + apellido.text

                    output_textedit.insertPlainText(f"{jugador}\n")
                    self.nombres_jugadores.append(jugador)  


                    self.driver.back()

            except:
                output_textedit = self.output_textedit
                color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
                formato_rojo = QTextCharFormat()
                formato_rojo.setForeground(color_rojo)
                output_textedit.mergeCurrentCharFormat(formato_rojo)
                output_textedit.insertPlainText('Un anuncio bloquea al scraper el acceso a la información, volviendo a intentarlo...\n')
                formato_negro = QTextCharFormat()
                formato_negro.setForeground(QColor(0, 0, 0))
                output_textedit.mergeCurrentCharFormat(formato_negro)
                # Crear un nuevo hilo y volver a intentarlo 
                thread = threading.Thread(target=self.scrapear_funcion)
                thread.start()
  
        else:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('No has iniciado sesion en la aplicación. Logueate con tus credenciales de Mister Fantasy MD en la ventana Perfil para acceder al mercado de jugaodes.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)

        self.driver.quit()  


class dataset_creator(QWidget):
  def __init__(self):
        super().__init__()

        main_layout = QVBoxLayout(self)

        self.stacked_widget = QStackedWidget()

        self.ventana1 = dataset_entrenamiento()
        self.ventana2 = dataset_predecir()

        self.stacked_widget.addWidget(self.ventana1)
        self.stacked_widget.addWidget(self.ventana2)

        button_layout = QHBoxLayout()  

        self.btn_ventana1 = QPushButton("Generar dataset para entrenar modelos")
        self.btn_ventana1.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(0))

        self.btn_ventana2 = QPushButton("Generar dataset para predecir valores")
        self.btn_ventana2.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))

        button_layout.addWidget(self.btn_ventana1)
        button_layout.addWidget(self.btn_ventana2)

        main_layout.addLayout(button_layout)  
        main_layout.addWidget(self.stacked_widget)

class dataset_entrenamiento(QWidget):
    
    def __init__(self):
        super().__init__()
        
        self.jugadoresS_noencontrados = ["Marc-André ter Stegen", "Adria Miquel Bosch Sanchis", "Sergio Ruiz Alonso", "Abderrahman Rebbach", "Kaiky", "Alejandro Pozo", "Lázaro", "Luis Javier Suárez", "Abdessamad Ezzalzouli", "Iván Cuéllar", "Djené", "Maximiliano Gómez", "Mamadou Mbaye", "Fali", "Anthony Lozano", "José María Giménez", "Sandro Ramírez", "Reinildo Isnard Mandava", "Chimy Ávila", "Pablo Ibáñez Lumbreras", "Portu", "Juan Carlos", "José Manuel Arnáiz", "Federico Valverde", "Alfonso Espino", "Ismaila Ciss", "Josep Chavarría", "José Pozo", "Imanol García de Albéniz", "Peru Nolaskoain Esnal", "Malcom Ares"] 
        self.jugadoresMD_noencontrados = ["Ter Stegen", "Miki Bosch", "Sergio Ruiz", "Abde Rebbach", "Kaiky Fernandes", "Álex Pozo", "Lázaro Vinicius", "Luis Suárez", "Abde Ezzalzouli", "Pichu Cuéllar", "Dakonam Djené", "Maxi Gómez", "Momo Mbaye", "Fali Giménez", "Choco Lozano", "José Giménez", "Sandro", "Reinildo Mandava", "Ezequiel Ávila", "Pablo Ibáñez", "Cristian Portu", "Juan Carlos Martín", "José Arnaiz", "Fede Valverde", "Pacha Espino", "Pathé Ciss", "Pep Chavarría", "José Ángel Pozo", "Imanol García", "Peru Nolaskoain", "Malcom Adu Ares"]

        # Crear un diseño de cuadrícula dentro del QVBoxLayout
        grid_layout = QGridLayout(self)
       
        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("Crear dataset para una jornada de LaLiga")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        grid_layout.addWidget(label_text, 0, 0)

        # LABEL SUBTÍTULO
        label_subtext = QLabel("Una vez scrapeado los partidos de la jornada de Sofaescore y las estadisticas de los jugaodres del Mister Fantasy,crea un dataset para entrenar un modelo con todos")
        label_subtext2 = QLabel("los datos fusionados y preprocesados.")
        grid_layout.addWidget(label_subtext, 1, 0, 1, 2)
        grid_layout.addWidget(label_subtext2, 2, 0, 1, 2)
           
        ### SELECCIONAR JORNADA INPUT ####################################################
        # INPUT NÚMERO JORNADA 
        label_number = QLabel("Jornada del dataset:")
        grid_layout.addWidget(label_number, 3, 0)
        # Estilos 
        self.number_input = QSpinBox(self)
        self.number_input.setMinimum(1)  # Establecer el valor mínimo (jornada 1)
        self.number_input.setMaximum(38)  # Establecer el valor máximo (Jornada 38)
        self.number_input.setSingleStep(1)  # Establecer el paso
        self.number_input.setMaximumSize(44, 20)
        self.number_input.setMinimumSize(44, 20)
        # Aliniación
        grid_layout.addWidget(self.number_input, 4, 0)

        ### SELECCIONAR RUTA DATASET DE ENTRADA SOFAESCORE #########################################################################
        # LABEL DE TEXTO
        label_text = QLabel("Selecionar carpeta donde se almacenaron todos los partidos scrapeados de la jornada de la web de Sofascore: ")
        grid_layout.addWidget(label_text, 5, 0)

        # INPUT DE TEXTO
        self.text_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input, 6, 0)

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 6, 1, alignment=Qt.AlignmentFlag.AlignLeft)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        ### SELECCIONAR RUTA DATASET DE ENTRADA MISTER FANTASY #####################################################################
        # LABEL DE TEXTO
        label_text = QLabel("Selecionar archivo resultante del scrapeo de la jornada de la web de Mister Fantasy Mundo Deportivo: ")
        grid_layout.addWidget(label_text, 7, 0)

        # INPUT DE TEXTO
        self.text_file_input= QLineEdit(self)  
        # Alineación
        grid_layout.addWidget(self.text_file_input, 8, 0)

        # BOTÓN PARA SELECCIONAR ARCHIVO
        select_file_button = QPushButton("Seleccionar archivo")
        select_file_button.clicked.connect(lambda: select_file(self))

        # Alineación
        grid_layout.addWidget(select_file_button, 8, 1, alignment=Qt.AlignmentFlag.AlignLeft)

        # Estilos
        select_file_button.setMinimumWidth(140)

        ### SELECCIONAR RUTA DONDE GUARDAR DATASET RESULTANTE #####################################################################
        # LABEL DE TEXTO
        label_text = QLabel("Selecionar ruta donde guardar el dataset generado de la jornada: ")
        grid_layout.addWidget(label_text, 9, 0)

        # INPUT DE TEXTO
        self.text_input2 = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input2, 10, 0)

        # BOTÓN PARA SELECCIONAR ARCHIVO
        select_folder_button = QPushButton("Seleccionar carpeta")
        select_folder_button.clicked.connect(lambda: select_folder2(self))

        # Alineación
        grid_layout.addWidget(select_folder_button, 10, 1, alignment=Qt.AlignmentFlag.AlignLeft)

        # Estilos
        select_folder_button.setMinimumWidth(140)

        ### BOTÓN PARA EJECUTAR FUNCIÓN PARA FUSIONAR EXCELLS ######################################################################
        # Crear un botón
        self.generate_button = QPushButton("Generar dataset")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.generate_button.clicked.connect(self.iniciar_thread_function)

        # Alineación
        grid_layout.addWidget(self.generate_button, 14, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        # Estilos
        self.generate_button.setMinimumWidth(150)

        # VENTANA OUTPUT SCRAPER #####################################################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        grid_layout.addWidget(self.output_textedit,15, 0, 2, 2)  # row, column, rowSpan, columnSpan
    
    def iniciar_thread_function(self):  
        # Crear un hilo y ejecutar la función en segundo plano
        thread = threading.Thread(target=self.json_a_excel)
        thread.start()

    def json_a_excel(self):

        def guardar_en_excell(output_archivo):

            # Obtener las listas de las filas
            fila_excel1 = df1.iloc[index_df1, :].tolist()
            fila_excel2 = df2.iloc[index_df2, :].tolist()

            # Concatenar las listas
            fila_concatenada = fila_excel2 + fila_excel1

            # Crear un DataFrame de pandas con una sola fila y múltiples columnas
            df_nueva_fila = pd.DataFrame([fila_concatenada])

            # Leer el archivo Excel existente
            df_existente = pd.read_excel(output_archivo, header=None)

            # Concatenar el DataFrame existente con la nueva fila
            df_final = pd.concat([df_existente, df_nueva_fila], ignore_index=True)

            # Escribir el DataFrame final en el archivo Excel
            df_final.to_excel(output_archivo, index=False, header=False)

        def procesar_cadena(cadena):
            match = re.match(r'(\d+)\s*\((\d+)\)', str(cadena))
            if match:
                numerator = int(match.group(1))
                denominator = int(match.group(2))
                if denominator != 0:
                    result = numerator / denominator
                    return float(result)  # Asegurarse de que el resultado sea float
                else:
                    return 0.0
            else:
                return 0.0
        
        #### PARTE 0: LEER INPUTS + COMPROBAR QUE TODAS LOS INPUTS (rutas de archivos y carpetas) HAN SIDO INICIALIZADAS
            
        # Número de la jornada
        num_jornada = self.number_input.text()

        # Ruta a la carpeta que contiene los archivos json de Sofaescore
        carpeta_SF = self.text_input.text()

        # Ruta del excell del fichero Mister Fantasy
        file_mf = self.text_file_input.text()

        # Ruta de la carpeta donde guardar el excell generado
        ruta_output = self.text_input2.text()

        if not num_jornada or not carpeta_SF or not file_mf or not ruta_output:
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            self.output_textedit.mergeCurrentCharFormat(formato_rojo)

            ## Comprobar si las variables han sido inicializadas
            if not num_jornada:
                self.output_textedit.insertPlainText("El número de la jornada no ha sido inicializado.\n")

            if not carpeta_SF:
                self.output_textedit.insertPlainText("La ruta de la carpeta donde encontrar todos los ficheros de estadisticas de todos los jugaores de LaLiga de Sofaescore no se ha inicializada.\n")

            if not file_mf:
                self.output_textedit.insertPlainText("La ruta del fichero donde encontrar todos las estadisticas de todos los jugaores de LaLiga de Mister Fantasy no se ha inicializada.\n")

            if not ruta_output:
                self.output_textedit.insertPlainText("La ruta de la carpeta donde guardar el datatset generado no ha sido inicializada.\n")

            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            self.output_textedit.mergeCurrentCharFormat(formato_negro)
            return


        # Parte 1: fusionar todos los jsons de todos los partidos scrapeados de la jornada ##############################################
        try:
            # Rutas globales
            carpeta_json = self.text_input.text()
            carpeta_xlsx = self.text_input.text()
            nombre_archivo_excel = 'todos_los_partidos_de_la_jornada.xlsx'

            # Lista para almacenar los DataFrames de cada archivo JSON
            dfs = []

            # Iterar sobre cada archivo en la carpeta
            for archivo_json in os.listdir(carpeta_json):
                if archivo_json.endswith(".json"):
                    with open(os.path.join(carpeta_json, archivo_json), "r") as file:
                        data = json.load(file)

                    # Crear un DataFrame vacío para cada archivo JSON
                    df = pd.DataFrame()

                    # Iterar sobre los elementos del JSON y agregarlos al DataFrame
                    for jugador, estadisticas in data.items():
                        df = pd.concat([df, pd.DataFrame([[jugador, estadisticas["puntuacion"]] + [stat[key] for stat in estadisticas["estadisticas"] for key in stat.keys()]], columns=["Nombre", "Puntuación"] + [key for stat in estadisticas["estadisticas"] for key in stat.keys()])], ignore_index=True)

                    # Agregar el DataFrame a la lista
                    dfs.append(df)

            # Concatenar todos los DataFrames en uno solo
            df_final = pd.concat(dfs, ignore_index=True)

            # Guardar el DataFrame en un archivo Excel
            ruta_excel = os.path.join(carpeta_xlsx, nombre_archivo_excel)
            df_final.to_excel(ruta_excel, index=False)

            # Parte 2: Fusionar MD con SC por nombre ########################################################################################
            # Rutas de los archivos Excel
            excel1_path = ruta_excel
            excel2_path = self.text_file_input.text()
            output = self.text_input2.text()
            numero_jornada = str(self.number_input.value())
            # Obtener la fecha actual
            fecha_actual = datetime.now()
            # Formatear la fecha como una cadena (opcional)
            fecha_actual_str = fecha_actual.strftime("%Y-%m-%d--%H-%M-%S")

            output_archivo=output+"/dataset_entrenamiento_jornada"+numero_jornada+"__"+fecha_actual_str+".xlsx"
            
            # Leer los datos de los archivos Excel
            df1 = pd.read_excel(excel1_path, header=None)
            df2 = pd.read_excel(excel2_path, header=None)
            
            # Obtener todas las celdas de la fila 1 (que ahora es la segunda fila después de desactivar el encabezado)
            fila_excel1 = df1.iloc[0, :].dropna().tolist()
            fila_excel2 = df2.iloc[0, :].dropna().tolist()

            # Concatenar las listas
            fila_concatenada =  fila_excel2 + fila_excel1

            # Crear un DataFrame de pandas con una sola fila y múltiples columnas
            df = pd.DataFrame([fila_concatenada])


            # Escribir el DataFrame en un archivo Excel
            df.to_excel(output_archivo, index=False, header=False)
        except:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('\nCarpeta de partidos scrpaeados de Sofascore o fichero de jugaodres scrapeados de Mister Fantasy erroneo. Asegurate de introducir la ruta correcta a la carpeta o archico correcto para poder generar el dataset.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return

        # PARTE 4: BUSCAR COINCIENDECIAS ENTRE AMBOS DATASETS  ######################################################################################

        # Inicializar el conjunto de valores encontrados
        valores_encontrados = set()
        df_fusionado = pd.DataFrame()
        encabezado=0
        contador_coincidencias=0
        contador_manual=0
        contador_global=0
        
        self.output_textedit.insertPlainText("_____________________________________________________________________________________________________\n")   
        self.output_textedit.insertPlainText("Buscando coincidencia entre jugadores...\n")
        
        # Iterar sobre las filas de df1 y comparar con las filas de df2
        for index_df1, row_df1 in df1.iterrows():
            value_to_compare1o =row_df1.iloc[0] 
            value_to_compare1 =value_to_compare1o.lower()
            value_to_compare1 =unidecode(value_to_compare1)
            value_to_compare1 =unidecode(value_to_compare1)
            value_to_compare1=value_to_compare1.replace(" ", "")
            
            coincidencia_encontrada = False
            contador_global+=1
            # Iterar sobre las filas de df2
            for index_df2, row_df2 in df2.iterrows():
                #print("-----",value_to_compare1,"-----",value_to_compare2,"-----")
                value_to_compare2o =row_df2.iloc[0]
                value_to_compare2 =value_to_compare2o.lower()
                value_to_compare2 =unidecode(value_to_compare2)
                value_to_compare2 = unidecode(value_to_compare2)
                value_to_compare2=value_to_compare2.replace(" ", "")
                
                # Calcular la distancia de Levenshtein
                distancia = Levenshtein.distance(value_to_compare1, value_to_compare2)
                # Establecer un umbral para considerar coincidencias
                umbral = 2  

                if distancia <= umbral:
                    self.output_textedit.insertPlainText(f"Coincidencia encontrada: excell1-fila-{index_df1} <-> excell2-fila.{index_df2} , {value_to_compare1} == {value_to_compare2}\n")
                    valores_encontrados.add(value_to_compare1) 

                    guardar_en_excell(output_archivo)

                    contador_coincidencias +=1
                    coincidencia_encontrada = True
                    time.sleep(0.02)

            # Imprimir si no se encontró ninguna coincidencia
            if not coincidencia_encontrada:
                if value_to_compare1!="nombre":
                    self.output_textedit.insertPlainText("------------------------------------------------------------------------------------------------\n")
                    self.output_textedit.insertPlainText(f"Coincidencia NO encontrada: excell1-fila-{index_df1} en {value_to_compare1}\n")
                    self.output_textedit.insertPlainText("------------------------------------------------------------------------------------------------\n")
        
        #Estadisticas de la fusion de los datasets
        self.output_textedit.insertPlainText("_____________________________________________________________________________________________________\n")       
        self.output_textedit.insertPlainText("Buscando jugaodres manualmente que no hicieron match...\n")
        for jugadorS, jugadorMD in zip(self.jugadoresS_noencontrados, self.jugadoresMD_noencontrados):
            for index_df1, row_df1 in df1.iterrows():
                value_to_compare1o = row_df1.iloc[0]
                for index_df2, row_df2 in df2.iterrows():
                    value_to_compare2o = row_df2.iloc[0]

                    if value_to_compare1o == jugadorS and value_to_compare2o == jugadorMD:
                        self.output_textedit.insertPlainText(f"Coincidencia encontrada: {jugadorS}\n")
                        guardar_en_excell(output_archivo)
                        contador_manual+=1
            
        #Resultados de la fusión de datasets
        self.output_textedit.insertPlainText("\n_____________________________________________________________________________________________________\n")
        self.output_textedit.insertPlainText(f"Total coincidencias: {contador_coincidencias} / {contador_global-1}\n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText(f"Añadidos manualmente: {contador_manual}\n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText(f"Jugadores no disponibles en MisterFantasy: {((contador_global-1)-(contador_coincidencias+contador_manual))}\n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText(f"Precisión: {(((contador_coincidencias+contador_manual)/(contador_global-1))*100)} %\n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText("Dataset fusionado correctamente.\n")

        time.sleep(0.5)
        self.output_textedit.insertPlainText("Procesando atributos...\n")

        # PARTE 4: PROCESAR DATOS DE LAS COLUMNAS ######################################################################################################################

        # Cargar el archivo Excel en un DataFrame
        df = pd.read_excel(output_archivo)

        #### Atributos Mister Fantasy
        ## NOMBRE #####
        # Verificar y convertir a str la columna 
        df.iloc[:, 0] = df.iloc[:, 0].apply(lambda x: str(x) if not isinstance(x, str) else x)

        ## VALOR #####
        # Verificar y convertir a float la columna 
        # Verificar y convertir a float la columna
        df.iloc[:, 1] = df.iloc[:, 1].apply(lambda x: float(str(x).replace('.', '').replace(',', '.')) if isinstance(x, (int, float, str)) else x)


        ## POSICIÓN #####
        # Verificar y convertir a str la columna 
        df.iloc[:, 2] = df.iloc[:, 2].apply(lambda x: str(x) if not isinstance(x, str) else x)

        ## EQUIPO #####
        # Verificar y convertir a str la columna 
        df.iloc[:, 3] = df.iloc[:, 3].apply(lambda x: str(x) if not isinstance(x, str) else x)

        ## PUNTUACIÓN FANTASY
        # Verificar y convertir a float la columna 
        df.iloc[:, 4] = pd.to_numeric(df.iloc[:, 4], errors='coerce')
        # Llenar los valores NaN con 0
        df.iloc[:, 4] = df.iloc[:, 4].fillna(0)

        ## AS
        # Verificar y convertir a float la columna 
        df.iloc[:, 5] = df.iloc[:, 5].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## MARCA
        # Verificar y convertir a float la columna 
        df.iloc[:, 6] = df.iloc[:, 6].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## MD 
        # Verificar y convertir a float la columna 
        df.iloc[:, 7] = df.iloc[:, 7].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## ULTIMO RIVAL #####
        # Verificar y convertir a str la columna 
        df.iloc[:, 8] = df.iloc[:, 8].apply(lambda x: str(x) if not isinstance(x, str) else x)

        ## RESULTADO DEL PARTIDO #####
        # Verificar y convertir "Win" a 1 y cualquier otro valor a 0 en toda la columna 9
        df.iloc[:, 9] = df.iloc[:, 9].map({"Win": 2, "Loss": 0, "Draw": 1}).fillna(0).astype(int)

        ## PROXIMO RIVAL #####
        # Verificar y convertir a str la columna 
        df.iloc[:, 10] = df.iloc[:, 10].apply(lambda x: str(x) if not isinstance(x, str) else x)

        ## PROXIMO PARTIDO ES LOCAL #####
        # Convertir valores booleanos a 1 y 0 en toda la columna 11
        df.iloc[:, 11] = df.iloc[:, 11].astype(int)

        ## MEDIA EN CASA #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 12] = df.iloc[:, 12].apply(lambda x: float(str(x).replace(',', '.')) if isinstance(x, str) else x)

        ## MEDIA FUERA #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 13] = df.iloc[:, 13].apply(lambda x: float(str(x).replace(',', '.')) if isinstance(x, str) else x)

        # EDAD #####
        # Verificar y convertir a entero la columna 19
        df.iloc[:, 19] = pd.to_numeric(df.iloc[:, 19], errors='coerce').fillna(0).astype(int)

        # ALTURA #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 15] = df.iloc[:, 15].apply(lambda x: float(str(x).replace('m', '').replace(',', '.')) if isinstance(x, (str, float)) else x)

        # PESO #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 16] = df.iloc[:, 16].apply(lambda x: float(str(x).replace('kg', '').replace(',', '.')) if isinstance(x, (str, float)) else x)

        #### Atributos Sofaescore
        ## PUNTUACIÓN SF#####
        # Verificar y convertir a float la columna 
        df.iloc[:, 18] = df.iloc[:, 18].apply(lambda x: float(str(x).replace(',', '.')) if not isinstance(x, float) else x)

        # MINUTES PLAYED #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 19] = df.iloc[:, 19].apply(lambda x: float(str(x).replace("'", '')) if not isinstance(x, (float, int)) else x)

        ## EXPECTED ASSISTS
        # Verificar y convertir a float la columna 
        df.iloc[:, 20] = df.iloc[:, 20].apply(lambda x: float(str(x).replace(',', '.')) if not isinstance(x, float) else x)

        ## (XA) SAVES
        # Verificar y convertir a float la columna 
        df.iloc[:, 21] = df.iloc[:, 21].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## GOALS PREVENTED
        # Verificar y convertir a float la columna 
        df.iloc[:, 22] = df.iloc[:, 22].apply(lambda x: float(str(x).replace(',', '.')) if not isinstance(x, float) else x)

        ## PUNCHES 
        # Verificar y convertir a float la columna 
        df.iloc[:, 23] = df.iloc[:, 23].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## RUNS OUT (SUCC.) 
        # Aplica la función a toda la columna 
        df.iloc[:, 24] = df.iloc[:, 24].apply(procesar_cadena)

        ## HIGH CLAIMS 
        # Verificar y convertir a float la columna 
        df.iloc[:, 25] = df.iloc[:, 25].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## TOUCHES #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 26] = df.iloc[:, 26].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## ACC. PASSES #####
        # Aplica la lógica de la expresión regular a toda la columna 
        df.iloc[:, 27] = df.iloc[:, 27].apply(lambda x: re.search(r'\((\d+)%\)', str(x)).group(1) if re.search(r'\((\d+)%\)', str(x)) else None)

        ## KEY PASSES ####
        # Verificar y convertir a float la columna 
        df.iloc[:, 28] = df.iloc[:, 28].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## CROSSES (ACC.) #####
        df.iloc[:, 29] = df.iloc[:, 29].apply(procesar_cadena)

        ## LONG BALLS (ACC.) #####
        df.iloc[:, 30] = df.iloc[:, 30].apply(procesar_cadena)

        ## CLEARANCES #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 31] = df.iloc[:, 31].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## BLOCKED SHOTS #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 32] = df.iloc[:, 32].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## INTERCEPTIONS #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 33] = df.iloc[:, 33].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## TACKLES #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 34] = df.iloc[:, 34].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## DRIBBLED PAST #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 35] = df.iloc[:, 35].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## GROUND DUELS (WON) ##### 
        # Apply the function to transform the column 
        df.iloc[:, 36] = df.iloc[:, 36].apply(procesar_cadena)

        ## AERIAL DUELS (WON) #####
        # Apply the function to transform the column 
        df.iloc[:, 37] = df.iloc[:, 37].apply(procesar_cadena)

        ## FOULS #####
        # Verificar y convertir a float la columna
        df.iloc[:, 38] = df.iloc[:, 38].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## WAS FOULED #####
        # Verificar y convertir a float la columna
        df.iloc[:, 39] = df.iloc[:, 39].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## SHOTS ON TARGET #####
        # Verificar y convertir a float la columna
        df.iloc[:, 40] = df.iloc[:, 40].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## SHOTS OFF TARGET #####
        # Verificar y convertir a float la columna
        df.iloc[:, 41] = df.iloc[:, 41].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## SHOTS BLOCKED #####
        # Verificar y convertir a float la columna
        df.iloc[:, 42] = df.iloc[:, 42].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## DRIBBLE ATTEMPTS (SUCC.) #####
        df.iloc[:, 43] = df.iloc[:, 43].apply(procesar_cadena)

        ## GOALS #####
        # Verificar y convertir a float la columna
        df.iloc[:, 44] = df.iloc[:, 44].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## ASSISTS #####
        # Verificar y convertir a float la columna
        df.iloc[:, 45] = df.iloc[:, 45].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## POSSESSION LOST #####
        # Verificar y convertir a float la columna
        df.iloc[:, 46] = df.iloc[:, 46].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## EXPECTED GOALS (XG) #####
        # Verificar y convertir a float la columna 
        df.iloc[:, 47] = df.iloc[:, 47].apply(lambda x: float(str(x).replace(',', '.')) if not isinstance(x, float) else x)

        ## PENALTY MISS #####
        # Verificar y convertir a float la columna
        df.iloc[:, 48] = df.iloc[:, 48].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## BIG CHANCES CREATED #####
        # Verificar y convertir a float la columna
        df.iloc[:, 49] = df.iloc[:, 49].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## PENALTY WON #####
        # Verificar y convertir a float la columna
        df.iloc[:, 50] = df.iloc[:, 50].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## BIG CHANCES MISSED #####
        # Verificar y convertir a float la columna
        df.iloc[:, 51] = df.iloc[:, 51].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## SAVES FROM INSIDE BOX #####
        # Verificar y convertir a float la columna
        df.iloc[:, 52] = df.iloc[:, 52].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## PENALTY COMMITTED #####
        # Verificar y convertir a float la columna
        df.iloc[:, 53] = df.iloc[:, 53].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## OFFSIDES #####
        # Verificar y convertir a float la columna
        df.iloc[:, 54] = df.iloc[:, 54].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## HIT WOODWORK #####
        # Verificar y convertir a float la columna
        df.iloc[:, 55] = df.iloc[:, 55].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## AUSENCIA #####
        # Verificar y convertir a str la columna 
        df.iloc[:, 56] = df.iloc[:, 56].apply(lambda x: str(x) if not isinstance(x, str) else x)

        ## ERROR LED TO SHOT #####
        # Verificar y convertir a float la columna
        df.iloc[:, 57] = df.iloc[:, 57].apply(lambda x: float(x) if not isinstance(x, float) else x)

        ## ERROR LED TO GOAL   #####
        # Verificar y convertir a float la columna
        df.iloc[:, 58] = df.iloc[:, 58].apply(lambda x: float(x) if not isinstance(x, float) else x)

        self.output_textedit.insertPlainText("Dataset procesado correctamente.\n")
        time.sleep(0.5)
        self.output_textedit.insertPlainText("Dataset guardao correctamente.\n")
        # Guardar el DataFrame modificado en un nuevo archivo Excel
        df.to_excel(output_archivo, index=False)

class dataset_predecir(QWidget):
    
    def __init__(self):
        super().__init__()

        self.lista_jugadores_datos_scrapeados=[]

        self.jugadoresS_noencontrados = ["Marc-André ter Stegen", "Adria Miquel Bosch Sanchis", "Sergio Ruiz Alonso", "Abderrahman Rebbach", "Kaiky", "Alejandro Pozo", "Lázaro", "Luis Javier Suárez", "Abdessamad Ezzalzouli", "Iván Cuéllar", "Djené", "Maximiliano Gómez", "Mamadou Mbaye", "Fali", "Anthony Lozano", "José María Giménez", "Sandro Ramírez", "Reinildo Isnard Mandava", "Chimy Ávila", "Pablo Ibáñez Lumbreras", "Portu", "Juan Carlos", "José Manuel Arnáiz", "Federico Valverde", "Alfonso Espino", "Ismaila Ciss", "Josep Chavarría", "José Pozo", "Imanol García de Albéniz", "Peru Nolaskoain Esnal", "Malcom Ares"] 
        self.jugadoresMD_noencontrados = ["Ter Stegen", "Miki Bosch", "Sergio Ruiz", "Abde Rebbach", "Kaiky Fernandes", "Álex Pozo", "Lázaro Vinicius", "Luis Suárez", "Abde Ezzalzouli", "Pichu Cuéllar", "Dakonam Djené", "Maxi Gómez", "Momo Mbaye", "Fali Giménez", "Choco Lozano", "José Giménez", "Sandro", "Reinildo Mandava", "Ezequiel Ávila", "Pablo Ibáñez", "Cristian Portu", "Juan Carlos Martín", "José Arnaiz", "Fede Valverde", "Pacha Espino", "Pathé Ciss", "Pep Chavarría", "José Ángel Pozo", "Imanol García", "Peru Nolaskoain", "Malcom Adu Ares"]
        self.teams_data = {
        "Real Madrid": "https://cdn.gomister.com/file/cdn-common/teams/15.png?version=20231117",
        "Real Sociedad": "https://cdn.gomister.com/file/cdn-common/teams/16.png?version=20231117",
        "Atlético de Madrid": "https://cdn.gomister.com/file/cdn-common/teams/2.png?version=20231117",
        "Girona": "https://cdn.gomister.com/file/cdn-common/teams/222.png?version=20231117",
        "Osasuna": "https://cdn.gomister.com/file/cdn-common/teams/50.png?version=20231117",
        "Athletic Club": "https://cdn.gomister.com/file/cdn-common/teams/1.png?version=20231117",
        "Valencia": "https://cdn.gomister.com/file/cdn-common/teams/19.png?version=20231117",
        "Granada": "https://cdn.gomister.com/file/cdn-common/teams/10.png?version=20231117",
        "Getafe": "https://cdn.gomister.com/file/cdn-common/teams/9.png?version=20231117",
        "Villarreal": "https://cdn.gomister.com/file/cdn-common/teams/20.png?version=20231117",
        "Las Palmas": "https://cdn.gomister.com/file/cdn-common/teams/11.png?version=20231117",
        "Mallorca": "https://cdn.gomister.com/file/cdn-common/teams/408.png?version=20231117",
        "Rayo Vallecano": "https://cdn.gomister.com/file/cdn-common/teams/14.png?version=20231117",
        "Barcelona": "https://cdn.gomister.com/file/cdn-common/teams/3.png?version=20231117",
        "Celta de Vigo": "https://cdn.gomister.com/file/cdn-common/teams/5.png?version=20231117",
        "Cádiz": "https://cdn.gomister.com/file/cdn-common/teams/499.png?version=20231117",
        "Alavés": "https://cdn.gomister.com/file/cdn-common/teams/48.png?version=20231117",
        "Almería": "https://cdn.gomister.com/file/cdn-common/teams/21.png?version=20231117",
        "Sevilla": "https://cdn.gomister.com/file/cdn-common/teams/17.png?version=20231117",
        "Betis": "https://cdn.gomister.com/file/cdn-common/teams/4.png?version=20231117",
        }

        # Crear un diseño de cuadrícula dentro del QVBoxLayout
        grid_layout = QGridLayout(self)
       
        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("Crear dataset sobre el que realizar prediciones")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        grid_layout.addWidget(label_text, 0, 0)

        # LABEL SUBTÍTULO
        label_subtext = QLabel("Crea un dataset con los jugaodores en mi plantilla o del mercado para realizar prediciones sobre su puntuación / valor de mercado en la prócima jornada")
        grid_layout.addWidget(label_subtext, 1, 0, 1, 2)

        ### SELECCIONAR JORNADA INPUT ####################################################
        # INPUT NÚMERO JORNADA 
        label_number = QLabel("Jornada del dataset:")
        grid_layout.addWidget(label_number, 2, 0)
        # Estilos 
        self.number_input = QSpinBox(self)
        self.number_input.setMinimum(1)  # Establecer el valor mínimo (jornada 1)
        self.number_input.setMaximum(38)  # Establecer el valor máximo (Jornada 38)
        self.number_input.setSingleStep(1)  # Establecer el paso
        self.number_input.setMaximumSize(44, 20)
        self.number_input.setMinimumSize(44, 20)
        # Aliniación
        grid_layout.addWidget(self.number_input, 3, 0)


        ### SELECCIONAR DATASET DE JUGAODRES DEL MERCADO O MI PLANTILLA #########################################################################
        # LABEL DE TEXTO
        label_text = QLabel("Selecionar fichero con los jugaodres en el mercado o de mi plantilla : ")
        grid_layout.addWidget(label_text, 4, 0)

        # INPUT DE TEXTO
        self.text_file_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_file_input, 5, 0)

        # BOTÓN PARA SELECCIONAR ARCHIVO
        select_file_button = QPushButton("Seleccionar Archivo")
        select_file_button.clicked.connect(lambda: select_file(self))
        # Alineación
        grid_layout.addWidget(select_file_button, 5, 1, alignment=Qt.AlignmentFlag.AlignLeft)
        # Estilos
        select_file_button.setMinimumWidth(140)

        ### SELECCIONAR RUTA DATASET DE ENTRADA SOFAESCORE #########################################################################
        # LABEL DE TEXTO
        label_text = QLabel("Selecionar carpeta donde se almacenaron tododos los partidos scrapeados de la jornada de la web de Sofaescore: ")
        grid_layout.addWidget(label_text, 6, 0)

        # INPUT DE TEXTO
        self.text_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input, 7, 0)

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 7, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        ### SELECCIONAR RUTA DATASET DE ENTRADA MISTER FANTASY #####################################################################
        # LABEL DE TEXTO
        label_text = QLabel("Selecionar archivo resultante del scrapeo de la jornada de la web de Mister Fantasy Mundo Deportivo: ")
        grid_layout.addWidget(label_text, 8, 0)

        # INPUT DE TEXTO
        self.text_file2_input= QLineEdit(self)  
        # Alineación
        grid_layout.addWidget(self.text_file2_input, 9, 0)

        # BOTÓN PARA SELECCIONAR ARCHIVO
        select_file_button = QPushButton("Seleccionar archivo")
        select_file_button.clicked.connect(lambda: select_file2(self))

        # Alineación
        grid_layout.addWidget(select_file_button, 9, 1, alignment=Qt.AlignmentFlag.AlignRight)

        # Estilos
        select_file_button.setMinimumWidth(140)


        ##### SELECCIONAR RUTA DONDE GUARDAR EL EXCEL CON LOS ATRIBUTOS PROCESADOS PARA INTRODUCIR AL MODELO #################################################
        ### LABEL DE TEXTO
        label_text = QLabel("Guardar dataset de los jugaodres a predecir:")
        grid_layout.addWidget(label_text, 10, 0)

        ### INPUT DE TEXTO
        self.text_input2 = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input2, 11, 0)

        ### BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder2(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 11, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        ### BOTÓN PARA GUARDAR DATASET ############
        # Crear un botón
        self.save_button = QPushButton("Generar dataset")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.save_button.clicked.connect(self.iniciar_scrapear_thread)

        # Alineación
        grid_layout.addWidget(self.save_button, 12, 0, alignment=Qt.AlignmentFlag.AlignLeft)

        # Estilos
        self.save_button.setMaximumWidth(200)

        #####  VENTANA OUTPUT  ####################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        grid_layout.addWidget(self.output_textedit, 13, 0, 5, 0)  # row, column, rowSpan, columnSpan

    def iniciar_scrapear_thread(self):  
        # Crear un hilo y ejecutar la función en segundo plano
        if usuario!="":
            thread = threading.Thread(target=self.generar_excell_predecir)
            thread.start()
        else:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('No has iniciado sesion en la aplicación. Logueate con tus credenciales de Mister Fantasy MD en la ventana Perfil para acceder al mercado de jugaodes.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)

    def json_a_excel(self):

        def guardar_en_excell():

            output = self.text_input2.text()
            numero_jornada = str(self.number_input.value())
            output_archivo=output+"/dataset_predecir_jornada"+numero_jornada+".xlsx"

            # Obtener las listas de las filas
            fila_excel1 = df1.iloc[index_df1, :].tolist()
            fila_excel2 = df2.iloc[index_df2, :].tolist()

            # Concatenar las listasnombre_archivo
            fila_concatenada = fila_excel2 + fila_excel1

            # Crear un DataFrame de pandas con una sola fila y múltiples columnas
            df_nueva_fila = pd.DataFrame([fila_concatenada])

            # Leer el archivo Excel existente
            df_existente = pd.read_excel(output_archivo, header=None)

            # Concatenar el DataFrame existente con la nueva fila
            df_final = pd.concat([df_existente, df_nueva_fila], ignore_index=True)

            # Escribir el DataFrame final en el archivo Excel
            df_final.to_excel(output_archivo, index=False, header=False)

        # Parte 1: fusionar todos los jsons de todos los partidos scrapeados de la jornada ##############################################
        # Rutas globales
        carpeta_json = self.text_input.text()
        nombre_archivo_excel = 'todos_los_partidos_de_la_jornada.xlsx'

        # Lista para almacenar los DataFrames de cada archivo JSON
        dfs = []

        # Iterar sobre cada archivo en la carpeta
        for archivo_json in os.listdir(carpeta_json):
            if archivo_json.endswith(".json"):
                with open(os.path.join(carpeta_json, archivo_json), "r") as file:
                    data = json.load(file)

                # Crear un DataFrame vacío para cada archivo JSON
                df = pd.DataFrame()

                # Iterar sobre los elementos del JSON y agregarlos al DataFrame
                for jugador, estadisticas in data.items():
                    df = pd.concat([df, pd.DataFrame([[jugador, estadisticas["puntuacion"]] + [stat[key] for stat in estadisticas["estadisticas"] for key in stat.keys()]], columns=["Nombre", "Puntuación"] + [key for stat in estadisticas["estadisticas"] for key in stat.keys()])], ignore_index=True)

                # Agregar el DataFrame a la lista
                dfs.append(df)

        # Concatenar todos los DataFrames en uno solo
        df_final = pd.concat(dfs, ignore_index=True)

        # Guardar el DataFrame en un archivo Excel
        ruta_excel = os.path.join(carpeta_json, nombre_archivo_excel)
        df_final.to_excel(ruta_excel, index=False)


        # Parte 2: Fusionar MD con SC por nombre ########################################################################################
        # Rutas de los archivos Excel
        excel1_path = ruta_excel
        excel2_path = self.text_file2_input.text()
        output = self.text_input2.text()
        numero_jornada = str(self.number_input.value())
        output_archivo=output+"/dataset_predecir_jornada"+numero_jornada+".xlsx"
        
        # Leer los datos de los archivos Excel
        df1 = pd.read_excel(excel1_path, header=None)
        df2 = pd.read_excel(excel2_path, header=None)
        
        # Obtener todas las celdas de la fila 1 (que ahora es la segunda fila después de desactivar el encabezado)
        fila_excel1 = df1.iloc[0, :].dropna().tolist()
        fila_excel2 = df2.iloc[0, :].dropna().tolist()

        # Concatenar las listas
        fila_concatenada =  fila_excel2 + fila_excel1

        # Crear un DataFrame de pandas con una sola fila y múltiples columnas
        df = pd.DataFrame([fila_concatenada])

        # Escribir el DataFrame en un archivo Excel
        df.to_excel(output_archivo, index=False, header=False)

        # Inicializar el conjunto de valores encontrados
        valores_encontrados = set()
        df_fusionado = pd.DataFrame()
        encabezado=0
        contador_coincidencias=0
        contador_manual=0
        contador_global=0
        
        self.output_textedit.insertPlainText("_____________________________________________________________________________________________________\n")   
        self.output_textedit.insertPlainText("Buscando coincidencia entre jugadores...\n")
        
        # Iterar sobre las filas de df1 y comparar con las filas de df2
        for index_df1, row_df1 in df1.iterrows():
            value_to_compare1o =row_df1.iloc[0] 
            value_to_compare1 =value_to_compare1o.lower()
            value_to_compare1 =unidecode(value_to_compare1)
            value_to_compare1 =unidecode(value_to_compare1)
            value_to_compare1=value_to_compare1.replace(" ", "")
            
            coincidencia_encontrada = False
            contador_global+=1
            # Iterar sobre las filas de df2
            for index_df2, row_df2 in df2.iterrows():
                #print("-----",value_to_compare1,"-----",value_to_compare2,"-----")
                value_to_compare2o =row_df2.iloc[0]
                value_to_compare2 =value_to_compare2o.lower()
                value_to_compare2 =unidecode(value_to_compare2)
                value_to_compare2 = unidecode(value_to_compare2)
                value_to_compare2=value_to_compare2.replace(" ", "")
                
                # Calcular la distancia de Levenshtein
                distancia = Levenshtein.distance(value_to_compare1, value_to_compare2)
                # Establecer un umbral para considerar coincidencias
                umbral = 2  

                if distancia <= umbral:
                    self.output_textedit.insertPlainText(f"Coincidencia encontrada: excell1-fila-{index_df1} <-> excell2-fila.{index_df2} , {value_to_compare1} == {value_to_compare2}\n")
                    valores_encontrados.add(value_to_compare1) 

                    guardar_en_excell()

                    contador_coincidencias +=1
                    coincidencia_encontrada = True
                    time.sleep(0.02)

            # Imprimir si no se encontró ninguna coincidencia
            if not coincidencia_encontrada:
                if value_to_compare1!="nombre":
                    self.output_textedit.insertPlainText("------------------------------------------------------------------------------------------------\n")
                    self.output_textedit.insertPlainText(f"Coincidencia NO encontrada: excell1-fila-{index_df1} en {value_to_compare1}\n")
                    self.output_textedit.insertPlainText("------------------------------------------------------------------------------------------------\n")
        
        #Estadisticas de la fusion de los datasets
        self.output_textedit.insertPlainText("_____________________________________________________________________________________________________\n")       
        self.output_textedit.insertPlainText("Buscando jugaodres manualmente que no hicieron match...\n")
        for jugadorS, jugadorMD in zip(self.jugadoresS_noencontrados, self.jugadoresMD_noencontrados):
            for index_df1, row_df1 in df1.iterrows():
                value_to_compare1o = row_df1.iloc[0]
                for index_df2, row_df2 in df2.iterrows():
                    value_to_compare2o = row_df2.iloc[0]

                    if value_to_compare1o == jugadorS and value_to_compare2o == jugadorMD:
                        self.output_textedit.insertPlainText(f"Coincidencia encontrada: {jugadorS}\n")
                        guardar_en_excell()
                        contador_manual+=1
            
        #Resultados de la fusión de datasets
        self.output_textedit.insertPlainText("\n_____________________________________________________________________________________________________\n")
        self.output_textedit.insertPlainText(f"Total coincidencias: {contador_coincidencias} / {contador_global-1}\n")
        self.output_textedit.insertPlainText(f"Añadidos manualmente: {contador_manual}\n")
        self.output_textedit.insertPlainText(f"Jugadores no disponibles en MisterFantasy: {((contador_global-1)-(contador_coincidencias+contador_manual))}\n")
        self.output_textedit.insertPlainText(f"Precisión: {(((contador_coincidencias+contador_manual)/(contador_global-1))*100)} %\n")
        self.output_textedit.insertPlainText("Dataset generado correctamente\n")

    def generar_excell_predecir(self):

        def encontrar_coincidencias(valores_Historico, valores_Mercado,ruta_archivo):
            self.output_textedit.insertPlainText("\n" + "=" * 100 + "\n")
            self.output_textedit.insertPlainText(f"Buscando coincidencias en {ruta_archivo}. \n")
            posiciones = []

            for pos, nombre_completo in enumerate(valores_Historico):
                coincidencias_iniciales = [nombre_abreviado for nombre_abreviado in valores_Mercado if nombre_abreviado.startswith(nombre_completo[0])]

                coincidencias_texto = get_close_matches(nombre_completo, coincidencias_iniciales)

                if coincidencias_texto:
                    posiciones.append(pos)
                    self.output_textedit.insertPlainText(f"Coincidencia en la fila {pos} para '{nombre_completo}'\n")
    
            return posiciones

        def scraping_data_selected():
            nombre = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[2]")
            apellido = self.driver.find_element(By.XPATH, " /html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[3]")
            jugador = nombre.text + " " + apellido.text
            self.output_textedit.insertPlainText(f"Nombre: {jugador}\n")
                    
            label_deseado= "Media en casa"
            elemento = self.driver.find_element(By.XPATH, f"//div[@class='item']//div[@class='label' and text()='{label_deseado}']/following-sibling::div[@class='value']")
            media_puntos_local = elemento.text
            self.output_textedit.insertPlainText(f"Media de puntos como local {media_puntos_local}\n")

            time.sleep(1)

            label_deseado= "Media fuera"
            elemento = self.driver.find_element(By.XPATH, f"//div[@class='item']//div[@class='label' and text()='{label_deseado}']/following-sibling::div[@class='value']")
            media_puntos_visitante = elemento.text
            self.output_textedit.insertPlainText(f"Media de puntos como visitante {media_puntos_visitante}\n")

            print("check 1 ")
            try:
                team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[1]/a/img")
            except:
                try:
                    team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div/div[3]/div/div[1]/div[2]/img[1]")
                except:
                    team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div/div[3]/div/div[1]/div[2]/img[2]")
            time.sleep(5)
            print("check 2 ")   
            image_url = team_logo_element.get_attribute("src")
            # Dividir la URL utilizando el signo de igual como delimitador
            parts = image_url.split('=')
            # El valor de version está en la segunda parte después del =
            version = parts[1]

            time.sleep(5)
            print("check 3 ")
            for equipo, url in self.teams_data.items():
                # Dividir la URL en base al signo de interrogación
                partes = url.split('?')
                    
                # Verificar si hay una parte después del signo de interrogación y actualizar la versión
                if len(partes) > 1:
                    partes[1] = f"version={version}"
                            
                    # Volver a unir las partes para formar la URL actualizada
                    nueva_url = '?'.join(partes)
                            
                    # Actualizar la URL en el diccionario
                    self.teams_data[equipo] = nueva_url
                        
            time.sleep(5)
            print("check 4 ")

            # Obtener src del equipo
            team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[1]/a/img")
            image_url = team_logo_element.get_attribute("src")

            equipo = None
            proximo_rival=None
            local= 0
            result=0
            ultimo_rival=None
            # Lista que contendrá múltiples jugadores
            # Comparar la URL de la imagen con las URLs en teams_data
            for equipo_nombre, equipo_url in self.teams_data.items():
                if image_url == equipo_url:
                    equipo = equipo_nombre

                    #### OBTENER RESULTADO ÚLTIMO PARTIDO ####
                    try:
                        divpartido = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div[1]/div[3]/div")
                    except:
                        divpartido = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div/div[2]/div")
                            
                    # Encuentra el div del partido
                    item_elements = divpartido.find_elements(By.CLASS_NAME, 'item')
                        
                    # Encuentra las imágenes dentro del div partido
                    img_elements = item_elements[0].find_elements(By.CLASS_NAME, 'team-logo')

                    # Guarda las src de las imágenes en variables
                    if len(img_elements) >= 2:
                        src_img1 = img_elements[0].get_attribute('src')
                        src_img2 = img_elements[1].get_attribute('src')
                        if src_img1 == image_url:
                            local = 1
                            for equipo_nombre, equipo_url in self.teams_data.items():
                                if src_img2 == equipo_url:
                                    proximo_rival=equipo_nombre
                        else:
                            local= 0
                            for equipo_nombre, equipo_url in self.teams_data.items():
                                if src_img1 == equipo_url:
                                    proximo_rival=equipo_nombre
                    else:
                        print("No se encontro el próximo partido")

            time.sleep(1)
            self.output_textedit.insertPlainText(f"Próximo equipo que enfrentará: {proximo_rival}\n")
            time.sleep(0.5)
            self.output_textedit.insertPlainText(f"Juega como local el próximo pártido: {local}\n")

            try:
                # Encuentra el elemento utilizando el XPath proporcionado con By
                try:
                    localizador = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div[1]/div[4]")
                    # Desplázate hacia el elemento
                    localizador.location_once_scrolled_into_view
                except:
                    localizador = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div[1]/div[6]")
                    # Desplázate hacia el elemento
                    localizador.location_once_scrolled_into_view

                # Encuentra el primer div con la clase "box box-scores"
                box_scores_div = self.driver.find_element(By.CLASS_NAME, 'box-scores')
                # Dentro de ese div, encuentra el primer div con la clase "line btn btn-player-gw"
                first_line_div = box_scores_div.find_element(By.CLASS_NAME, 'line.btn.btn-player-gw')


                # Interactúa con el elemento (por ejemplo, haz clic en él)
                first_line_div.click()

                # Encontrar el div principal con la clase "player-match"
                player_match_div = self.driver.find_element(By.CLASS_NAME, "player-match")

                # Encontrar los subelementos dentro del div principal
                team_1 = player_match_div.find_element(By.CLASS_NAME, "left").find_element(By.CLASS_NAME, "team").text
                goals_team_1 = [int(goal.text) for goal in player_match_div.find_elements(By.CLASS_NAME, "goals")][0]  
                goals_team_2 = [int(goal.text) for goal in player_match_div.find_elements(By.CLASS_NAME, "goals")][1]  
                team_2 = player_match_div.find_element(By.CLASS_NAME, "right").find_element(By.CLASS_NAME, "team").text

                if team_1 == equipo:
                    ultimo_rival=team_2

                    if goals_team_1 > goals_team_2:
                        result = 2 #Win
                    elif goals_team_1 < goals_team_2:  
                        result = 0 #Loss
                    else:
                        result = 1 #Draw

                else:
                    ultimo_rival=team_1

                    if goals_team_1 > goals_team_2:
                        result = 0 #Loss
                    elif goals_team_1 < goals_team_2:  
                        result = 2 #Win
                    else:
                        result = 1 #Draw
                                
                time.sleep(1)

                ausencia=""
                    
                self.driver.back()
                        
            except:
                result=None
                ultimo_rival=""

                # Encontrar el primer elemento con la clase "line btn btn-player-gw" usando By.CLASS_NAME
                elemento_line = self.driver.find_element(By.CLASS_NAME, "line.btn.btn-player-gw")

                # Obtener el HTML del elemento
                html = elemento_line.get_attribute("outerHTML")

                # Verificar si el div con la clase "status" existe e imprimir su texto
                if 'Sancionado' in html:
                    ausencia="Suspension"
                elif 'Lesionado' in html:
                    ausencia="Injury"
                            
            self.output_textedit.insertPlainText(f"Último equipo enfrentado: {ultimo_rival}\n")
            time.sleep(0.5)
            self.output_textedit.insertPlainText(f"Resultado último partido: {result}\n")
            time.sleep(0.5)
            self.output_textedit.insertPlainText(f"Ausencia en el último partido: {ausencia}\n")
            

            # Crea un diccionario para el jugador actual
            jugador = {"Nombre": jugador,
                    "ultimo_rival": ultimo_rival,
                    "resultado_partido": result, 
                    "prximo_rival": proximo_rival, 
                    "prximo_partido_local": local,
                    "media_casa": media_puntos_local,
                    "media_fuera": media_puntos_visitante,
                    "ausencia": ausencia,
                    }
            self.lista_jugadores_datos_scrapeados.append(jugador)
                    
            time.sleep(4)
            self.driver.back()
            time.sleep(4)

        #### PARTE 0 : LEER INPUTS + COMPROBAR QUE TODAS LOS INPUTS (rutas de archivos y carpetas) HAN SIDO INICIALIZADAS ########################################################################################

        # Número de la jornada
        num_jornada = self.number_input.text()

        # Fichero con los futbolistas (mercado/planilla) a procesar
        archivo_excel_selected_players = self.text_file_input.text()

        # Ruta del excell del fichero Mister Fantasy
        archivo_excel_MF = self.text_file2_input.text()

        # Ruta a la carpeta que contiene los archivos json de Sofaescore
        carpeta_SF = self.text_input.text()

        # Ruta a la carpeta donde guardar dataset generado
        carpeta_save = self.text_input2.text()

        output = self.text_input2.text()
        numero_jornada = str(self.number_input.value())
        output_archivo=output+"/dataset_predecir_jornada"+numero_jornada+".xlsx"

        # Ruta de la carpeta donde guardar el excell generado
        ruta_output = self.text_input2.text()
        if not num_jornada or not archivo_excel_selected_players or not archivo_excel_MF or not carpeta_SF or not carpeta_save:
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            self.output_textedit.mergeCurrentCharFormat(formato_rojo)

            ## Comprobar si las variables han sido inicializadas
            if not num_jornada:
                self.output_textedit.insertPlainText("El número de la jornada no ha sido inicializado.\n")

            if not archivo_excel_selected_players:
                self.output_textedit.insertPlainText("El fichero de los jugaodres del mercado o de la plantilla no se ha inicializado.\n")
                
            if not archivo_excel_MF:
                self.output_textedit.insertPlainText("La ruta de la carpeta donde encontrar todos los ficheros de estadisticas de Sofaecore de todos los jugaores de LaLiga no se ha inicializada.\n")
                
            if not carpeta_SF:
                self.output_textedit.insertPlainText("La ruta de la carpeta donde encontrar todos los ficheros de estadisticas de Mister Fantasy de todos los jugaores de LaLiga no se ha inicializada.\n")
            
            if not carpeta_save:
                self.output_textedit.insertPlainText("La ruta de la carpeta donde guardar el datatset generado no ha sido inicializada.\n")
            
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            self.output_textedit.mergeCurrentCharFormat(formato_negro)
            return
        
        # Obtener el nombre del archivo sin la extensión
        nombre_archivo = os.path.splitext(os.path.basename(archivo_excel_selected_players))[0]

        # Convierte el nombre del archivo a minúsculas para facilitar la comparación
        nombre_archivo = nombre_archivo.lower()
        
        if not ('plantilla' in nombre_archivo or 'mercado' in nombre_archivo):
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('\nArchivo de jugadores del mercado o plantilla erroneo.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return
        
        #### PARTE 1 : GENERAR DATASET resultate de fusionar los daasets de Sofaescore y Mister Fantasy ########################################################################################
        try:
            self.json_a_excel()
        except:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('\nCarpeta de partidos scrpaeados de Sofascore o fichero de jugaodres scrapeados de Mister Fantasy erroneo. Asegurate de introducir la ruta correcta a la carpeta o archico correcto para poder generar el dataset.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return
            
        #### PARTE 2 : ABRIR FICHERO DE JUGADORES DE MERCADO / MI PLANTILLA #####################################################################################################################
        try:
            # Ruta al archivo Excel
            self.output_textedit.insertPlainText("\n" + "_" * 100 + "\n")
            time.sleep(1)
            self.output_textedit.insertPlainText(f"Abriendo fichero de jugadores selecioandos...\n")
            time.sleep(1)
            self.output_textedit.insertPlainText("\n" + "" * 100 + "\n")
            time.sleep(1)

            # Lee el archivo Excel con pandas y especifica que no hay encabezado
            df = pd.read_excel(archivo_excel_selected_players, header=None)

            # Obtén los valores de la primera columna (columna 0)
            valores_Mercado = df.iloc[:, 0].tolist()

            # Imprime o utiliza los valores según sea necesario
            self.output_textedit.insertPlainText(f"Total de jugadores encontrados: {len(valores_Mercado)}\n")

            for valor in valores_Mercado:
                self.output_textedit.insertPlainText(f"{valor}\n")
                time.sleep(0.5)
            time.sleep(1)
            self.output_textedit.insertPlainText(f"\n") 
        except:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('\nFichero de jugadores del mercado o mi plantilla erroneo o inaccesible. Asegurate de introducir la ruta correcta del archico correcto para poder generar el dataset.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return
        
        #### PARTE 3 : SCRAPING DE DATOS DE MISTER FATASY DE JUGAODRES #########################################################################################################################
        time.sleep(0.5)
        self.output_textedit.insertPlainText("\n" + "_" * 100 + "\n")
        time.sleep(0.5)
        self.output_textedit.insertPlainText(f"Accediendo a la web de Mister Fnatasy para scrapear datos de los jugadores que se almacenarán en el dataset...\n")
        
        while True:
            try:
                #Creamos driver con la url del fantasy y nos logueamos
                self.driver = webdriver.Chrome()
                time.sleep(10)
                realizar_login(self.driver)
                
                # Analizamso si el fichero para comprobar si contiene jugaodres del mercado o de mi plantilla y en consecuencia scrapear en el apartado de "mi plantilla" o "mercado"
                # Obtener el nombre del archivo sin la extensión
                nombre_archivo = os.path.splitext(os.path.basename(archivo_excel_selected_players))[0]

                # Convierte el nombre del archivo a minúsculas para facilitar la comparación
                nombre_archivo = nombre_archivo.lower()

                # Verifica si el nombre del archivo contiene la palabra "mercado"
                if 'mercado' in nombre_archivo:
                    time.sleep(2)
                    self.output_textedit.insertPlainText("Scrapeando datos de jugadores del mercado...\n")
                    # PARTE 3.1 : SCRAPEAR MERCADO
                    try:
                        # Pinchar en el botón Market
                        market = self.driver.find_element(By.XPATH, '//*[@id="content"]/header/div[2]/ul/li[2]/a')
                        time.sleep(2)
                        market.click()
                        time.sleep(3)
                    except (ElementNotInteractableException, NoSuchElementException):
                        # Maneja la excepción y espera antes de intentar nuevamente
                        self.output_textedit.insertPlainText("Anuncio detectado, reiniciando driver...")
                        self.driver.refresh()
                        time.sleep(6) 
                        market = self.driver.find_element(By.XPATH, '//*[@id="content"]/header/div[2]/ul/li[2]/a')
                        market.click()
                    
                    # Encuentra el elemento <ul> con el id "list-on-sale"
                    ul_element = self.driver.find_element(By.ID, "list-on-sale")

                    # Encuentra los elementos <div> con la clase "name" dentro del elemento <ul>
                    div_elements = ul_element.find_elements(By.CSS_SELECTOR, "div.name")

                    # Itera sobre los elementos <div> encontrados e imprime el nombre del jugador
                    for div_element in div_elements:
                        self.output_textedit.insertPlainText("\n" + "-" * 40 + "\n")
                        time.sleep(0.5)

                        #nombre_elemento 
                        name_element = div_element.text
                        time.sleep(1)
                        #Imprime el nombre del jugador
                        self.output_textedit.insertPlainText(f"{name_element}:\n")
                        
                        time.sleep(1)
                        try:
                            div_element.click()
                        except: 
                            self.driver.execute_script("window.scrollBy(0, arguments[0]);", 300)  
                            time.sleep(0.5)
                            div_element.click()
                            
                        time.sleep(2)
                        scraping_data_selected()
                        time.sleep(2)

                # Verifica si el nombre del archivo contiene la palabra "plantilla"
                if 'plantilla' in nombre_archivo:
                    # PARTE 3.2 : SCRAPEAR PLANTILLA
                    time.sleep(2)
                    self.output_textedit.insertPlainText("Scrapeando datos de jugadores de mi plantilla...\n")
                    time.sleep(2)
                    
                    try:
                        # Pinchar en el botón Market
                        squad = self.driver.find_element(By.XPATH, '//*[@id="content"]/header/div[2]/ul/li[3]/a')

                        time.sleep(2)
                        squad.click()
                        time.sleep(3)
                    except (ElementNotInteractableException, NoSuchElementException):
                        # Maneja la excepción y espera antes de intentar nuevamente
                        self.output_textedit.insertPlainText(f"Anuncio detectado, reiniciando driver...\n")
                        self.driver.refresh()
                        time.sleep(6) 
                        squad.click()

                    time.sleep(3)
                    try:
                        div = self.driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div[4]/ul')
                    except:
                        self.driver.refresh()
                        time.sleep(6)
                        div = self.driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div[4]/ul')

                    # Encuentra todos los elementos li dentro de la ul
                    li_elementos = div.find_elements(By.TAG_NAME, "li")
            
                    # Itera sobre los elementos li
                    for li_elemento in li_elementos:
                        self.output_textedit.insertPlainText("\n" + "-" * 40 + "\n")
                        time.sleep(0.5)

                        # Encuentra el elemento 'div' con la clase 'name'
                        nombre_elemento = li_elemento.find_element(By.CLASS_NAME, "name")
                        time.sleep(1)
                        # Extrae el texto del elemento 'div' con la clase 'name'
                        nombre_jugador = nombre_elemento.text.strip()
                        # Imprime el nombre del jugador
                        self.output_textedit.insertPlainText(f"{nombre_jugador}:\n")
                        time.sleep(1)

                        try:
                            nombre_elemento.click()
                        except (ElementNotInteractableException, NoSuchElementException):
                            # Maneja la excepción y espera antes de intentar nuevamente
                            self.output_textedit.insertPlainText(f"Anuncio detectado, reiniciando driver...\n")
                            self.driver.refresh()

                        time.sleep(2)
                        scraping_data_selected()

                # Imprimir la información de cada jugador en la lista
                for jugador in self.lista_jugadores_datos_scrapeados:
                    time.sleep(0.5)
                    print(jugador)
                self.driver.quit()

                # Salimos del bucle xq se ha conseguid completar el scraping sin anuncios bloqueantes
                break
            
            except:
                output_textedit = self.output_textedit
                color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
                formato_rojo = QTextCharFormat()
                formato_rojo.setForeground(color_rojo)
                output_textedit.mergeCurrentCharFormat(formato_rojo)
                output_textedit.insertPlainText('Un anuncio bloquea al scraper el acceso a la información, volviendo a intentarlo...\n')
                formato_negro = QTextCharFormat()
                formato_negro.setForeground(QColor(0, 0, 0))
                output_textedit.mergeCurrentCharFormat(formato_negro)
                self.driver.quit()
                
        #### PARTE 4 : Buscar jugaodres en los datasets de estadisticas que me interesan (jugaodres de mi plantilla / jugaodres en el mercado actual)
        # Lista global para almacenar todas las filas seleccionadas
        filas_jugadores = []

        output = self.text_input2.text()
        numero_jornada = str(self.number_input.value())
        output_archivo=output+"/dataset_predecir_jornada"+numero_jornada+".xlsx"

        # Lista todos los archivos en la carpeta con extensión .xlsx
        archivos_excel = [output_archivo]
        self.output_textedit.insertPlainText("\n" + "_" * 100 + "\n")
        self.output_textedit.insertPlainText("Buscando coincidencias...\n")
        
        # Itera sobre cada archivo Excel en la carpeta
        for archivo in archivos_excel:
            
            # Ruta al archivo Excel actual
            ruta_archivo = os.path.join(output_archivo, archivo)

            # Lee el archivo Excel con pandas y especifica que no hay encabezado
            df = pd.read_excel(ruta_archivo, header=None)

            # Obtén los valores de la primera columna (o cualquier columna que desees)
            valores_Historico = df.iloc[:, 0].tolist()

            # Llama a la función encontrar_coincidencias y actualiza los valores de los índices
            posiciones_coincidentes = encontrar_coincidencias(valores_Historico, valores_Mercado, ruta_archivo)
            
            # Accede a las filas correspondientes a los índices proporcionados
            filas_seleccionadas = df.iloc[posiciones_coincidentes]

            # Convierte las filas seleccionadas a una lista de diccionarios para un fácil acceso a los atributos
            datos_filas = filas_seleccionadas.to_dict(orient='records')

            # Agrega las filas seleccionadas a la lista global
            filas_jugadores.extend(datos_filas)
 
        # Leer y almacenar cabecera
        df = pd.read_excel(ruta_archivo, header=None, nrows=1)

        # Almacena los nombres de las columnas en una lista
        nombres_columnas = list(df.iloc[0])

        self.output_textedit.insertPlainText("\n" + "_" * 100 + "\n")
        self.output_textedit.insertPlainText("\nObteniendo cabezera...\n")
        # Imprime los nombres de las columnas
        self.output_textedit.insertPlainText(f"{nombres_columnas}\n")


        #### PARTE 5 : CREAR EXCELL FINAL PARA PASAR AL MODELO Y PREDECIR  ####################################################################################################
        # Obtener la fecha actual
        fecha_actual = datetime.now()
        # Formatear la fecha como una cadena (opcional)
        fecha_actual_str = fecha_actual.strftime("%Y-%m-%d--%H-%M-S")

        nombre_archivo= ruta_output +"/jugadores_a_predecir"+fecha_actual_str+".xlsx"

        # Verificar si el archivo existe y eliminarlo si es necesario
        if os.path.exists(nombre_archivo):
            os.remove(nombre_archivo)

        # Crear un nuevo libro de trabajo (Workbook)
        wb = openpyxl.Workbook()

        # Seleccionar la hoja activa
        sheet = wb.active

        # Agregar los nombres de las columnas a la primera fila de la hoja
        sheet.append(nombres_columnas)

        # Guardar el libro de trabajo en el archivo especificado
        wb.save(nombre_archivo)


        #### PARTE 6 : #### Acceder a las estadisticas de cada jugaodor del mercado/plantilla en todos los datasets(partidos de cada jornada de LaLiga) donde se han encontrado registros de este, para: 
        ## - Generar medias, porcentajes... de valores numéricos estadisticos alamcenados en todos los dataset (Ej: media de goles)
        ## - Acceder a la web de Mister Fantasy par inicializar ciertos atributos (Ej: proximo rival al que se enfrentará un jugaodor)

        self.output_textedit.insertPlainText("\n" + "_" * 100 + "\n")
        self.output_textedit.insertPlainText(f"Procesando datos de cada jugaodor selecionado...\n")

        # Diccionario para almacenar los índices de las filas repetidas
        indice_nombres = {}

        # Iterar filas para las repetidas
        for indice, fila_jugador in enumerate(filas_jugadores):
            # Obtiene el nombre del jugador (asumiendo que está en la primera posición)
            nombre_jugador = fila_jugador[0]
            
            # Verifica si el nombre ya ha sido visto
            if nombre_jugador in indice_nombres:
                indice_nombres[nombre_jugador].append(indice)
            else:
                indice_nombres[nombre_jugador] = [indice]


        # Gestionar filas repetidas 
        for nombre, indices in indice_nombres.items():
            #print("Vuelta Done")
            self.output_textedit.insertPlainText("\n" + "_" * 75+ "\n")
            self.output_textedit.insertPlainText(f"Jugador {nombre}: \n")

            # Acceder al nombre de cada índice repetido
            for indice in indices:
                nombre_repetido = filas_jugadores[indice][0]
                self.output_textedit.insertPlainText(f"   Nombre del jugador detectado en la fila {indice} como {nombre_repetido}\n")
            
            self.output_textedit.insertPlainText("\n" + "-" * 60+ "\n")
            # Buscar el nombre en el diccionario
            posicion = next((index for (index, jugador) in enumerate(self.lista_jugadores_datos_scrapeados) if jugador['Nombre'] == nombre), None)
           
            if posicion is not None:
                self.output_textedit.insertPlainText(f"El nombre '{nombre}' se encuentra en la posición {posicion} del diccionario.")

                # GENERAR DATAFRAME para cada jugador ###################################################################
                self.output_textedit.insertPlainText("Datos del jugador: \n")

                time.sleep(0.5) 

                ## NOMBRE #####
                nombre= filas_jugadores[indice][0]
                self.output_textedit.insertPlainText(f"Nombre: {nombre}\n")

                time.sleep(0.5) 

                ## VALOR #####
                valor= 0.0
                try:
                    for indice in indices:
                        valor_jugador1 = str(filas_jugadores[indice][1])
                        valor_duplicado = float(valor_jugador1.replace('.', ''))
                        valor = valor + valor_duplicado
                    valor= valor/len(indices)
                    self.output_textedit.insertPlainText(f"Valor:{valor}\n")
                except:
                    valor= None

                time.sleep(0.5) 

                ## POSICIÓN #####
                try:
                    pos_jugador=filas_jugadores[indice][2]
                    self.output_textedit.insertPlainText(f"Posición: {pos_jugador}\n")
                except:
                    pos_jugador= ""
                time.sleep(0.5)    
                ## EQUIPO #####
                try:
                    eq_jugador=filas_jugadores[indice][3]
                    self.output_textedit.insertPlainText(f"Equipo: {eq_jugador}\n")
                except:
                    pos_jugador= ""

                time.sleep(0.5)    

                ## PUNTUACIÓN FANTASY, AS, MARCA, MD #####
                fantasy = 0.0
                ass = 0.0
                marca = 0.0
                md = 0.0
                for indice in indices:
                    try:
                        mf = float(filas_jugadores[indice][4])
                    except:
                        mf=0.0
                    try:
                        a = float(filas_jugadores[indice][5])
                    except:
                        a=0.0
                    try:
                        mc = float(filas_jugadores[indice][6])
                    except:
                        mc=0.0
                    try:
                        m = float(filas_jugadores[indice][7])
                    except:
                        m=0.0

                    fantasy = fantasy + mf
                    ass = ass + a
                    marca = marca + mc
                    md = md + m

                time.sleep(0.5) 
                valor= valor/len(indices)
                self.output_textedit.insertPlainText(f"Puntuación Mister Fantasy:{fantasy}\n")

                time.sleep(0.5) 
                valor= valor/len(indices)
                self.output_textedit.insertPlainText(f"Puntuación AS: {ass}\n")

                time.sleep(0.5) 
                valor= valor/len(indices)
                self.output_textedit.insertPlainText(f"Puntuación Marca: {marca}\n")

                time.sleep(0.5) 
                valor= valor/len(indices)
                self.output_textedit.insertPlainText(f"Puntuación Mundo Deportivo: {md}\n")
                
                time.sleep(0.5) 


                ## ULTIMO RIVAL ##### 
                try:
                    ultimo_equipo_rival = self.lista_jugadores_datos_scrapeados[posicion]['ultimo_rival']
                    # Comprobar si es str y convertir si no lo es
                    if not isinstance(ultimo_equipo_rival, str):
                        ultimo_equipo_rival = str(ultimo_equipo_rival)
                except:
                    ultimo_equipo_rival = None
                self.output_textedit.insertPlainText(f"Último equipo rival: {ultimo_equipo_rival}\n")

                time.sleep(0.5) 

                ## RESULTADO DEL PARTIDO #####
                try:
                    resultado_ultimo_partido = self.lista_jugadores_datos_scrapeados[posicion]['resultado_partido']
                    # Comprobar si es int y convertir si no lo es
                    if not isinstance(resultado_ultimo_partido, int):
                        try:
                            resultado_ultimo_partido = int(resultado_ultimo_partido)
                            # Comprobar si está en el rango esperado (0, 1, o 2)
                            if resultado_ultimo_partido not in {0, 1, 2}:
                                resultado_ultimo_partido = None
                                print("El resultado no está en el rango esperado (0, 1, o 2). Se estableció como nulo.")
                        except (ValueError, TypeError):
                            print("No se puede convertir a entero (1).")
                except:
                    resultado_ultimo_partido = None

                self.output_textedit.insertPlainText(f"Resultado último partido: {resultado_ultimo_partido}\n")

                time.sleep(0.5) 

                ## PROXIMO RIVAL #####
                try:
                    proximo_equipo_rival = self.lista_jugadores_datos_scrapeados[posicion]['prximo_rival']
                    # Comprobar si es str y convertir si no lo es
                    if not isinstance(proximo_equipo_rival, str):
                        proximo_equipo_rival = str(proximo_equipo_rival)
                except:
                    proximo_equipo_rival = None
                self.output_textedit.insertPlainText(f"Próximo equipo rival: {proximo_equipo_rival}\n")

                time.sleep(0.5) 

                ## PROXIMO PARTIDO ES LOCAL #####
                try:
                    proximo_equipo_local = self.lista_jugadores_datos_scrapeados[posicion]['prximo_partido_local']
                    if not isinstance(proximo_equipo_local, int):
                        try:
                            proximo_equipo_local = int(proximo_equipo_local)
                            # Comprobar si está en el rango esperado (0 o 1)
                            if proximo_equipo_local not in {0, 1}:
                                proximo_equipo_local = None
                                print("El valor no está en el rango esperado (0 o 1). Se estableció como nulo.")
                        except (ValueError, TypeError):
                            print("No se puede convertir a entero (2).")
                except:
                    proximo_equipo_local = None
                self.output_textedit.insertPlainText(f"Próximo equipo como local: {proximo_equipo_local}\n")

                time.sleep(0.5) 

                ## MEDIA EN CASA #####
                try:
                    media_puntos_local = self.lista_jugadores_datos_scrapeados[posicion]['media_casa']
                    # Comprobar si es float y convertir si no lo es
                    if not isinstance(media_puntos_local, float):
                        try:
                            media_puntos_local = float(media_puntos_local.replace(',', '.'))  # Reemplazar ',' por '.' para manejar decimales
                        except (ValueError, TypeError):
                            print("No se puede convertir a punto flotante.")
                except:
                    media_puntos_local = None
                self.output_textedit.insertPlainText(f"Media de puntos como local: {media_puntos_local}\n")
                
                time.sleep(0.5) 

                ## MEDIA FUERA #####
                try:
                    media_puntos_fuera = self.lista_jugadores_datos_scrapeados[posicion]['media_fuera']
                    # Comprobar si es float y convertir si no lo es
                    if not isinstance(media_puntos_fuera, float):
                        try:
                            media_puntos_fuera = float(media_puntos_fuera.replace(',', '.'))  # Reemplazar ',' por '.' para manejar decimales
                        except (ValueError, TypeError):
                            print("No se puede convertir a punto flotante.")
                except:
                    media_puntos_fuera = None
                self.output_textedit.insertPlainText(f"Media de puntos como visitante: {media_puntos_fuera}\n")

                time.sleep(0.5) 

                # EDAD #####
                try:
                    edad=filas_jugadores[indice][14]
                    self.output_textedit.insertPlainText(f"Edad: {edad}\n")
                except:
                    edad=None

                time.sleep(0.5)  
                
                # ALTURA #####
                try:
                    try:
                        altura_jugador = filas_jugadores[indice][15]
                        altura_jugador = float(altura_jugador)
                        self.output_textedit.insertPlainText(f"Altura: {altura_jugador}\n")
                    except (ValueError, TypeError):
                        altura_jugador = filas_jugadores[indice][15].replace('m', '').replace(',', '.')
                        altura_jugador = float(altura_jugador)
                        self.output_textedit.insertPlainText(f"Altura: {altura_jugador}\n")
                except:
                    altura_jugador=None

                time.sleep(0.5) 

                # PESO #####
                try:
                    try:
                        peso_jugador = filas_jugadores[indice][16]
                        peso_jugador = float(peso_jugador)
                        self.output_textedit.insertPlainText(f"Peso: {peso_jugador}\n")
                    except (ValueError, TypeError):
                        peso_jugador = filas_jugadores[indice][16].replace('kg', '').replace(',', '.')
                        peso_jugador = float(peso_jugador)
                        self.output_textedit.insertPlainText(f"Peso: {peso_jugador}\n")
                except:
                    peso_jugador=None

                time.sleep(0.5) 

                ## PUNTUACIÓN SF#####
                puntuacion= 0.0
                try:
                    for indice in indices:
                            pts=float(filas_jugadores[indice][18])
                            puntuacion = puntuacion + pts   
                    puntuacion= puntuacion/len(indices)
                    self.output_textedit.insertPlainText(f"Puntuación SF: {puntuacion}\n")
                except:
                    puntuacion=None

                time.sleep(0.5) 

                # MINUTES PLAYED #####
                minutes = 0.0
                try:
                    for indice in indices:
                        try:
                            mn = float(filas_jugadores[indice][19])
                        except (ValueError, TypeError):
                            mn = float(str(filas_jugadores[indice][19]).replace("'", ''))

                        minutes += mn

                    minutes = minutes / len(indices)
                    self.output_textedit.insertPlainText(f"Minutos disputados: {minutes}\n")
                except:
                    minutes=None

                time.sleep(0.5) 
                
                ## EXPECTED ASSISTS 
                expassists = 0
                try:
                    for indice in indices:
                        expa = float(filas_jugadores[indice][20])
                        expassists = expassists + expa
                    expassists = expassists / len(indices)
                    self.output_textedit.insertPlainText(f"Provavilidad de que un pase sea asistencia: {expassists}\n")
                except:
                    expassists=None

                time.sleep(0.5) 

                ## ESTADÍSCAS PORTEROS #################################################################
                ## (XA), SAVES, GOALS PREVENTED, PUNCHES, RUNS OUT (SUCC.), HIGH CLAIMS #####
                saves = 0
                golasPrev = 0
                punches = 0
                runsOut=0
                hightClaims=0
                for indice in indices:   
                    try:
                        saves = saves + float(filas_jugadores[indice][21])
                    except:
                        saves=None
                    try:
                        golasPrev = golasPrev + float(filas_jugadores[indice][22])
                    except:
                        golasPrev=None
                    try:
                        punches = punches + float(filas_jugadores[indice][23])
                    except:
                        punches=None
                    
                    try:
                        match = re.match(r'(\d+)\s*\((\d+)\)', str(filas_jugadores[indice][24]))
                        if match:
                            # Extraer los números y convertirlos a enteros
                            numero_externo = int(match.group(1))
                            numero_interno = int(match.group(2))
                            runs = 0
                            if  numero_interno !=0:
                                # Realizar la división 
                                runs = numero_externo / numero_interno

                            runsOut = runsOut + runs
                    except:
                        runsOut=None
                    try:
                        hightClaims = hightClaims+ float(filas_jugadores[indice][25])
                    except:
                        hightClaims=None

                time.sleep(0.5)    
                saves = saves / len(indices)
                self.output_textedit.insertPlainText(f"Paradas (porteros): {saves}\n")

                time.sleep(0.5) 
                golasPrev = golasPrev / len(indices)
                self.output_textedit.insertPlainText(f"Goles evitados (porteros): {golasPrev}\n")

                time.sleep(0.5)       
                punches = punches / len(indices)
                self.output_textedit.insertPlainText(f"Despeje de puños (porteros): {punches}\n")

                time.sleep(0.5) 
                runsOut = runsOut / len(indices)
                self.output_textedit.insertPlainText(f"Salida con balón esxitosa (porteros): {punches}\n")

                time.sleep(0.5) 
                hightClaims = hightClaims / len(indices)
                self.output_textedit.insertPlainText(f"Captura de balon en altura exitosa (porteros): {hightClaims}\n")

                time.sleep(0.5) 

                ## TOUCHES #####
                touches=0
                try:
                    for indice in indices:        
                        touch = float(filas_jugadores[indice][26])
                        touches = touches+ touch
                        
                    touches = touches / len(indices)
                    self.output_textedit.insertPlainText(f"Toques de balón: {touches}\n")
                except:
                    touches=None

                time.sleep(0.5) 

                ## ACC. PASSES #####
                accpass = 0
                try:
                    for indice in indices:        
                        accp_match = re.search(r'\((\d+)%\)', str(filas_jugadores[indice][27]))
                        accp = accp_match.group(1) if accp_match else None
                        if accp:
                            accpass += float(accp)

                    accpass = accpass / len(indices) if len(indices) > 0 else 0
                    accpass = accpass /100
                    self.output_textedit.insertPlainText(f"Pases precisos: {accpass}\n")
                except:
                    accpass=None 

                time.sleep(0.5) 

                ## KEY PASSES #####
                keypass=0
                try:
                    for indice in indices:        
                        keypass = keypass+ float(filas_jugadores[indice][28])
                    keypass = keypass / len(indices)
                    self.output_textedit.insertPlainText(f"Pases claves: {keypass}\n")
                except:
                    keypass=None

                time.sleep(0.5) 

                ## CROSSES (ACC.) #####
                crosses=0
                try:
                    for indice in indices: 
                        match = re.match(r'(\d+)\s*\((\d+)\)', str(filas_jugadores[indice][29]))
                        if match:
                            # Extraer los números y convertirlos a enteros
                            numero_externo = int(match.group(1))
                            numero_interno = int(match.group(2))
                            crs = 0
                            if  numero_interno !=0:
                                # Realizar la división y devolver el resultado
                                crs = numero_externo / numero_interno
                            
                            crosses = crs + crosses
                            
                    crosses = crosses / len(indices)
                    self.output_textedit.insertPlainText(f"Centros precisos: {crosses}\n")
                except:
                    crosses=None

                time.sleep(0.5) 

                ## LONG BALLS (ACC.) #####
                longball = 0.0
                try:
                    for indice in indices: 
                        match = re.match(r'(\d+)\s*\((\d+)\)', str(filas_jugadores[indice][30]))
                        if match:
                            # Extraer los números y convertirlos a enteros
                            numero_externo = int(match.group(1))
                            numero_interno = int(match.group(2))
                            longb = 0
                            if  numero_interno !=0:
                                # Realizar la división y devolver el resultado
                                longb = numero_externo / numero_interno
                            
                            longball = longb + longball
                            
                    longball = longball / len(indices)
                    self.output_textedit.insertPlainText(f"Pase largo preciso: {longball}\n")
                except:
                    longball=None

                time.sleep(0.5) 


                ## CLEARANCES #####
                clearances= 0.0
                try:
                    for indice in indices:        
                        clearances = clearances + float(filas_jugadores[indice][31])
                    clearances = clearances / len(indices)
                    self.output_textedit.insertPlainText(f"Ocasiones evitadas de gol: {clearances}\n")
                except:
                    clearances=None

                time.sleep(0.5) 

                ## BLOCKED SHOTS #####
                blocketshoots= 0.0
                try:
                    for indice in indices:        
                        blocketshoots = blocketshoots + float(filas_jugadores[indice][32])
                    blocketshoots = blocketshoots / len(indices)
                    self.output_textedit.insertPlainText(f"Tiros bloqueados: {blocketshoots}\n")
                except:
                    blocketshoots=None

                time.sleep(0.5) 

                ## INTERCEPTIONS #####
                interceptions= 0.0
                try:
                    for indice in indices:        
                        interceptions = interceptions + float(filas_jugadores[indice][33])
                    interceptions = interceptions / len(indices)
                    self.output_textedit.insertPlainText(f"Pases nterceptados: {interceptions}\n")
                except:
                    interceptions=None

                time.sleep(0.5) 

                ## TACKLES #####
                tackles=0.0
                try:
                    for indice in indices:        
                        tackles = tackles + float(filas_jugadores[indice][34])
                    tackles = tackles / len(indices)
                    self.output_textedit.insertPlainText(f"Robos de balón sin falta: {tackles}\n")
                except:
                    tackles=None

                time.sleep(0.5) 

                ## DRIBBLED PAST #####
                dribbled=0.0
                try:
                    for indice in indices:        
                        dribbled = dribbled + float(filas_jugadores[indice][35])
                    dribbled = dribbled / len(indices)
                    self.output_textedit.insertPlainText(f"Veces regateado: {dribbled}\n")
                except:
                    dribbled=None

                time.sleep(0.5) 

                ## GROUND DUELS (WON) #####
                groundduels=0.0
                try:
                    for indice in indices: 
                        match = re.match(r'(\d+)\s*\((\d+)\)', str(filas_jugadores[indice][36]))
                        if match:
                            # Extraer los números y convertirlos a enteros
                            numero_externo = int(match.group(1))
                            numero_interno = int(match.group(2))
                            grnddu = 0
                            if  numero_interno !=0:
                                # Realizar la división y devolver el resultado
                                grnddu = numero_externo / numero_interno

                            groundduels = groundduels + grnddu

                    groundduels = groundduels / len(indices)
                    self.output_textedit.insertPlainText(f"Duelos en el suelo Ganados: {groundduels}\n")
                except:
                    groundduels=None

                time.sleep(0.5) 

                ## AERIAL DUELS (WON) #####
                airduels=0.0
                try:
                    for indice in indices: 
                        match = re.match(r'(\d+)\s*\((\d+)\)', str(filas_jugadores[indice][37]))
                        if match:
                            # Extraer los números y convertirlos a enteros
                            numero_externo = int(match.group(1))
                            numero_interno = int(match.group(2))
                            airdu = 0
                            if  numero_interno !=0:
                                # Realizar la división y devolver el resultado
                                airdu = numero_externo / numero_interno

                            airduels = airduels + airdu

                    airduels = airduels / len(indices)
                    self.output_textedit.insertPlainText(f"Duelos aereos ganados: {airduels}\n")
                except:
                    airduels=None

                time.sleep(0.5) 

                ## FOULS #####
                fouls=0.0
                try:
                    for indice in indices:        
                        dribbled = dribbled + float(filas_jugadores[indice][38])
                    dribbled = dribbled / len(indices)
                    self.output_textedit.insertPlainText(f"Faltas cometidas: {fouls}\n")
                except:
                    fouls=None

                time.sleep(0.5) 

                ## WAS FOULED #####
                fouled=0.0
                try:
                    for indice in indices:        
                        fouled = fouled + float(filas_jugadores[indice][39])
                    fouled = fouled / len(indices)
                    self.output_textedit.insertPlainText(f"Faltas recibidas: {fouled}\n")
                except:
                    fouled=None

                time.sleep(0.5)

                ## SHOTS ON TARGET #####
                tarhetshoot=0.0
                try:
                    for indice in indices:        
                        tarhetshoot = tarhetshoot + float(filas_jugadores[indice][40])
                    dribbled = dribbled / len(indices)
                    self.output_textedit.insertPlainText(f"Tiros a puerta: {tarhetshoot}\n")
                except:
                    tarhetshoot=None

                time.sleep(0.5) 

                ## SHOTS OFF TARGET #####
                offtargetshoot=0.0
                try:
                    for indice in indices:        
                        offtargetshoot = offtargetshoot + float(filas_jugadores[indice][41])
                    offtargetshoot = offtargetshoot / len(indices)
                    self.output_textedit.insertPlainText(f"Tiros que no fueron a puerta: {offtargetshoot}\n")
                except:
                    offtargetshoot=None

                time.sleep(0.5) 

                ## SHOTS BLOCKED #####
                blockedshoots=0.0
                try:
                    for indice in indices:        
                        blockedshoots = blockedshoots + float(filas_jugadores[indice][42])
                    blockedshoots = blockedshoots / len(indices)
                    self.output_textedit.insertPlainText(f"Disparos realizados bloqueados : {blockedshoots}\n")
                except:
                    blockedshoots=None

                time.sleep(0.5) 

                ## DRIBBLE ATTEMPTS (SUCC.) #####
                dribbleatempts=0.0
                try:
                    for indice in indices: 
                        match = re.match(r'(\d+)\s*\((\d+)\)', str(filas_jugadores[indice][43]))
                        if match:
                            # Extraer los números y convertirlos a enteros
                            numero_externo = int(match.group(1))
                            numero_interno = int(match.group(2))
                            dribs = 0
                            if  numero_interno !=0:
                                # Realizar la división y devolver el resultado
                                dribs = numero_externo / numero_interno

                            dribbleatempts = dribbleatempts + dribs

                    dribbleatempts = dribbleatempts / len(indices)
                    self.output_textedit.insertPlainText(f"Regates exitosos: {dribbleatempts}\n")
                except:
                    dribbleatempts=None

                time.sleep(0.5) 

                ## GOALS #####
                goals=0.0
                try:
                    for indice in indices:        
                        goals = goals + int(filas_jugadores[indice][44])
                    goals = goals / len(indices)
                    self.output_textedit.insertPlainText(f"Goles: {goals}\n")
                except:
                    goals=None

                time.sleep(0.5) 

                ## ASSISTS #####
                assists=0.0
                try:
                    for indice in indices:        
                        assists = assists + int(filas_jugadores[indice][45])
                    assists = assists / len(indices)
                    self.output_textedit.insertPlainText(f"Asistencias: {assists}\n")
                except:
                    assists=None

                time.sleep(0.5) 

                ## POSSESSION LOST #####
                posesion_perdida = 0.0
                try:
                    for indice in indices:        
                        posesion_perdida = posesion_perdida + int(filas_jugadores[indice][46])
                    posesion_perdida = posesion_perdida / len(indices)
                    self.output_textedit.insertPlainText(f"Posesión perdida: {posesion_perdida}\n")
                except:
                    posesion_perdida=None

                time.sleep(0.5)

                ## EXPECTED GOALS (XG) #####
                xgoals = 0.0
                try:
                    for indice in indices:        
                        xgoals = xgoals + float(filas_jugadores[indice][47])
                    xgoals = xgoals / len(indices)
                    self.output_textedit.insertPlainText(f"Provavilidad de que ls disparos se conviertan en goles: {xgoals}\n")
                except:
                    xgoals=None

                time.sleep(0.5)

                ####
                penalty_miss =0.0
                big_chances_created =0.0
                penalty_won =0.0
                big_chance_miss=0.0
                savesBox=0.0
                penaltyDone=0.0   
                oddside=0.0
                palos=0.0
                
                for indice in indices:     
                    try:   
                        penalty_miss = penalty_miss + float(filas_jugadores[indice][48])
                    except:
                        penalty_miss=None

                    try:
                        big_chances_created = big_chances_created + float(filas_jugadores[indice][49])
                    except:
                        big_chances_created=None
                    
                    try:
                        penalty_won = penalty_won + float(filas_jugadores[indice][50])
                    except:
                        penalty_won=None
                    
                    try:
                        big_chance_miss = big_chance_miss + float(filas_jugadores[indice][51])
                    except:
                        big_chance_miss=None
                    
                    try:
                        savesBox = savesBox + float(filas_jugadores[indice][52])
                    except:
                        savesBox=None
                    
                    try:
                        penaltyDone = penaltyDone + float(filas_jugadores[indice][53])
                    except:
                        penaltyDone=None
                    
                    try:
                        oddside = oddside + float(filas_jugadores[indice][54])
                    except:
                        oddside=None

                    try:
                        palos = palos + float(filas_jugadores[indice][55])
                    except:
                        palos=None
                
                ## PENALTY MISS #####
                penalty_miss = penalty_miss / len(indices)
                self.output_textedit.insertPlainText(f"Penaltis fallados : {penalty_miss}\n")
                time.sleep(0.5) 

                ## BIG CHANCES CREATED #####
                big_chances_created = big_chances_created / len(indices)
                self.output_textedit.insertPlainText(f"Ocasiones de gol generadas: {big_chances_created}\n")       
                time.sleep(0.5) 

                ## PENALTY WON #####
                penalty_won = penalty_won / len(indices)
                self.output_textedit.insertPlainText(f"Penaltys conseguidos: {penalty_won}\n")
                time.sleep(0.5) 

                ## BIG CHANCES MISSED #####
                big_chance_miss = big_chance_miss / len(indices)
                self.output_textedit.insertPlainText(f"Ocasiones de gol desperdiciadas: {big_chance_miss}\n")
                time.sleep(0.5) 

                ## SAVES FROM INSIDE BOX #####
                savesBox = savesBox / len(indices)
                self.output_textedit.insertPlainText(f"Ocasione sd e gol evitadas desde dentro del area: {savesBox}\n")
                time.sleep(0.5) 

                ## PENALTY COMMITTED #####
                penaltyDone = penaltyDone / len(indices)
                self.output_textedit.insertPlainText(f"Penalties cometidos: {penaltyDone}\n")
                time.sleep(0.5) 

                ## OFFSIDES #####      
                oddside = oddside / len(indices)
                self.output_textedit.insertPlainText(f"Veces en fuera de juego: {oddside}\n")
                time.sleep(0.5) 

                ## HIT WOODWORK #####
                palos = palos / len(indices)
                self.output_textedit.insertPlainText(f"Disparos al paloooo: {palos}\n")

                time.sleep(0.5) 

                ## AUSENCIA #####    
                try:
                    ausencia = self.lista_jugadores_datos_scrapeados[posicion]['ausencia']
                    # Comprobar si es str y convertir si no lo es
                    if not isinstance(ausencia, str):
                        ausencia = str(ausencia)
                except:
                    ausencia = None
                self.output_textedit.insertPlainText(f"Causa por no estar convocado: {ausencia}\n")

                time.sleep(0.5)  
                print("check ausencia")

                ## ERROR LED TO SHOT #####
                shot_led_error=0
                try:
                    for indice in indices:
                        shot_led_error = shot_led_error + float(filas_jugadores[indice][57])
                    shot_led_error = shot_led_error / len(indices)
                    self.output_textedit.insertPlainText(f"Errores que derivan en un disparo del rival: {shot_led_error}\n")
                except:
                    shot_led_error=None

                time.sleep(0.5)  

                ## ERROR LED TO GOAL   ##### 
                goal_led_error=0
                try:
                    for indice in indices:
                        goal_led_error = goal_led_error + float(filas_jugadores[indice][58])
                    goal_led_error = goal_led_error / len(indices)
                    self.output_textedit.insertPlainText(f"Errores que derivan en un gol del rival: {goal_led_error}\n")
                except:
                    goal_led_error=None
                
                #########################################################################################################
                
                self.output_textedit.insertPlainText("\n")
                time.sleep(20)

                # Cargar excell
                wb = openpyxl.load_workbook(nombre_archivo)
                time.sleep(0.5)

                # Seleccionar la hoja activa
                sheet = wb.active
                        
                # Lista de variables a almacenar
                nueva_fila = [
                nombre, valor, pos_jugador, eq_jugador,
                fantasy, ass, marca, md, ultimo_equipo_rival, resultado_ultimo_partido, proximo_equipo_rival, proximo_equipo_local,
                media_puntos_local, media_puntos_fuera, edad, altura_jugador, peso_jugador, puntuacion, minutes, expassists, saves, golasPrev, punches, runsOut,
                hightClaims, touches, accpass, keypass, crosses, longball, tarhetshoot, offtargetshoot, blockedshoots, dribbleatempts, goals, assists, clearances,
                blocketshoots, interceptions, tackles, dribbled, groundduels, airduels, fouls, fouled, savesBox, penaltyDone, oddside, palos, ausencia, shot_led_error,
                goal_led_error, posesion_perdida, xgoals, penalty_miss, big_chances_created, penalty_won, big_chance_miss]

                    
                # Escribir la nueva fila en la hoja de cálculo
                sheet.append(nueva_fila)

                time.sleep(0.5)

                # Guardar el archivo Excel
                wb.save(nombre_archivo)
            else:
                print(f"El nombre '{nombre}' no se encontró en el diccionario.")
                self.output_textedit.insertPlainText(f"Coincidencia erronea\n")
            
        os.remove(output_archivo)
        self.output_textedit.insertPlainText(f"Se han generado exitosamente el dataset en {nombre_archivo} para realizar la sprediciones sobre los jugadores selecionados.\n")


class scrapear_datos(QWidget):
  def __init__(self):
        super().__init__()

        main_layout = QVBoxLayout(self)

        self.stacked_widget = QStackedWidget()

        self.ventana1 = PlayerScraperWindowMF("Players Scraper MD")
        self.ventana2 = PlayerScraperWindowSC("Players Scraper SF")

        self.stacked_widget.addWidget(self.ventana1)
        self.stacked_widget.addWidget(self.ventana2)

        button_layout = QHBoxLayout()  

        self.btn_ventana1 = QPushButton("Extraer datos de Mister Fantasy Mundo Deportivo")
        self.btn_ventana1.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(0))

        self.btn_ventana2 = QPushButton("Extraer datos de Sofaescore")
        self.btn_ventana2.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))

        button_layout.addWidget(self.btn_ventana1)
        button_layout.addWidget(self.btn_ventana2)

        main_layout.addLayout(button_layout)  
        main_layout.addWidget(self.stacked_widget)

class PlayerScraperWindowSC(QWidget):
    def __init__(self, window_title):
        super().__init__()
        self.setWindowTitle(window_title)

        # Variables de la clase
        self.selected_folder = ""
        self.selected_path = ""
        
        self.driver = None

        self.progress = 0
        self.ruta_jornada=""
        self.slugJson=""

        # Crear un diseño de cuadrícula
        layout = QGridLayout(self)

        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("Sofaescore Scraper")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        layout.addWidget(label_text, 0, 0,1, 2)

        # LABEL SUBTÍTULO
        label_subtext = QLabel("Obtener el listado de todos los jugaodres titulares, suplentes y no vonvocados y sus estadisticas de juego asociadas.")
        layout.addWidget(label_subtext, 1, 0, 1, 2)


        ### SELECCIONAR EQUIPO INPUT ####################################################
        # INPUT NOMBRRE EQUIPO 
        label_equipo = QLabel("Equipo a scrapear:")
        # Aliniación
        layout.addWidget(label_equipo, 2, 0)

        # Lista de equipos
        teams = ["Real Madrid", "Barcelona", "Atl. de Madrid", "Valencia", "Sevilla", "Villarreal",
                 "Real Sociedad", "Real Betis", "Athletic Club", "Celta Vigo", "Almeria", "Getafe",
                 "Mallorca", "Girona", "Granada", "", "Alavés", "Rayo Vallecano", "Osasuna", "Las Palmas"]

        # Crear un QComboBox y agregar los equipos
        self.team_combobox = QComboBox(self)
        self.team_combobox.addItems(teams)
        # Aliniación
        layout.addWidget(self.team_combobox, 2, 1)
        # Estilos 
        self.team_combobox.setFixedWidth(120)


        ### SELECCIONAR JORNADA INPUT ####################################################
        # INPUT NÚMERO JORNADA 
        label_number = QLabel("Jornada a scrapear:")
        layout.addWidget(label_number, 3, 0)
        # Estilos 
        self.number_input = QSpinBox(self)
        self.number_input.setMinimum(1)  # Establecer el valor mínimo (jornada 1)
        self.number_input.setMaximum(38)  # Establecer el valor máximo (Jornada 38)
        self.number_input.setSingleStep(1)  # Establecer el paso
        self.number_input.setMaximumSize(45, 20)
        self.number_input.setMinimumSize(45, 20)
        # Aliniación
        layout.addWidget(self.number_input, 3, 1)


        ###  SELECCIONAR RUTA DONDE GUARDAR EL EXCEL OUTPUT DEL SCRAPER  #################
        # LABEL TEXTO 
        label_text = QLabel("Ruta output scraper:")
        layout.addWidget(label_text, 4, 0)

        # INPUT TEXTO (QLineEdit en lugar de QSpinBox)
        self.text_input = QLineEdit(self)
        # Alineación
        layout.addWidget(self.text_input, 4, 1)
        # Estilos 
        self.text_input.setMinimumWidth(350)
        
        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación  
        layout.addWidget(select_folder_button, 5, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos 
        select_folder_button.setMinimumWidth(140)


        ###  BOTÓN PARA INICIAR SCRAPER  ################################################
        # Crear un botón llamado "Scrapear"
        scrape_button = QPushButton("Scrapear")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar barra progreso
        scrape_button.clicked.connect(self.iniciar_scrapear_thread)
        scrape_button.clicked.connect(self.start_progress)

        # Alineación 
        layout.addWidget(scrape_button, 6, 0)
        # Estilos
        self.number_input.setMaximumSize(38, 20)

        ###  BARRA DE PROGRESO  ################################################
        # Crear Barra de progreso
        self.progress_bar = QProgressBar(self)
        layout.addWidget(self.progress_bar, 6, 1)

        ###  VENTANA OUTPUT SCRAPER  ####################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        layout.addWidget(self.output_textedit, 7, 0, 11, 0)  # row, column, rowSpan, columnSpan

    def start_progress(self):
        # Establecer el rango de la barra de progreso según tus necesidades
        self.progress_bar.setRange(0, 22)

        ruta_output = self.text_input.text()
        if ruta_output!="":
            self.progress_bar.setValue(0) 
        
    def iniciar_scrapear_thread(self):  
        # Crear un hilo y ejecutar la función en segundo plano
        thread = threading.Thread(target=self.scrapear_funcion)
        thread.start()

    def invocar_actualizacion(self, nuevo_valor):
        QMetaObject.invokeMethod(self.progress_bar, "setValue", Qt.ConnectionType.QueuedConnection, Q_ARG(int, nuevo_valor))
    
    def performance_to_json(self, JsonJugador):
        nombre_archivo = f"performance_jugadores_partido_{self.slugJson}.json"
        ruta_completa_archivo = os.path.join(self.ruta_jornada, nombre_archivo)
        
        # Verificar si el archivo JSON existe
        if os.path.exists(ruta_completa_archivo):
            # Si el archivo existe, cargar su contenido
            with open(ruta_completa_archivo, 'r') as archivo:
                datos = json.load(archivo)
        else:
            # Si el archivo no existe, crear un diccionario vacío
            datos = {}
        # Actualizar el diccionario existente con la nueva entrada
        datos.update(JsonJugador)
        
        # Guardar los datos actualizados en el archivo JSON
        if not os.path.exists(ruta_completa_archivo):
            with open(ruta_completa_archivo, 'w') as archivo:
                archivo.write("{}")  # Crea un archivo vacío si no existe
        with open(ruta_completa_archivo, 'w') as archivo:
            json.dump(datos, archivo, indent=4)

    def obtener_informacion_jugador(self):

        # Obtiene el contenido HTML de la página
        pagina_html = self.driver.page_source
        #print(pagina_html)

        # Utiliza BeautifulSoup para analizar el HTML
        soup = BeautifulSoup(pagina_html, 'html.parser')

        try:        
            nombre= self.driver.find_element(By.XPATH,'//*[@id="__next"]/main/div[3]/div/div/div/div[1]/div/div[1]/a/div')
            self.output_textedit.append("_______________________________")
            self.output_textedit.append(nombre.text)
            
            try:
                puntuacion= self.driver.find_element(By.XPATH,'//*[@id="__next"]/main/div[3]/div/div/div/div[1]/div/div[2]/div/span')
                self.output_textedit.append(puntuacion.text)

                self.output_textedit.append("_______________________________")

                # DEVOLVER TODOS LOS PARÁMETROS DE RENDIMIENTO DEL JUGADOR: encontrar todos los div con la clase "sc-fqkvVR sc-dcJsrY litZes eFJwJL"
                entradas = []
                goal_elements = soup.find_all('div', class_='sc-fqkvVR sc-dcJsrY litZes eFJwJL')
                for element in goal_elements:
                    
                    # Encuentra el div con la clase "sc-jEACwC hFGVAX" y el span con la clase "sc-jEACwC jnyhQn" dentro de este div
                    ### ¡¡¡¡¡¡¡¡ Revisar clases de los dos elementos a continuación los cuales pueden cambiar con el tiempo !!!!!!!! 
                    div_goal = element.find('div', class_='sc-jEACwC dGzxxl')
                    span_goal = element.find('span', class_='sc-jEACwC jrhxaB')
                    ### ¡¡¡¡¡¡¡¡ Revisar clases de los dos elementos a continuación los cuales pueden cambiar con el tiempo !!!!!!!! 
                    if div_goal and span_goal:
                        self.output_textedit.append(f"{div_goal.text}: {span_goal.text}")

                        estadisticas = {}
                        clave = div_goal.text
                        valor = span_goal.text
                        estadisticas[clave] = valor

                        entrada = {
                            clave: valor
                        }

                        # Agrega la entrada JSON a la lista
                        entradas.append(entrada)
                        
                    nombreJson=nombre.text
                    puntuaciónJson=puntuacion.text
            
                    # Crear el diccionario para el jugador
                    JsonJugador = {
                        nombreJson: {
                            "puntuacion": puntuaciónJson,
                            "estadisticas": entradas
                        }
                    }
                    
                    self.performance_to_json(JsonJugador)

            except NoSuchElementException as e:
                self.output_textedit.append("Sin jugar")
                self.output_textedit.append("_______________________________")
                
                entradas = []
                estadisticas = {}
                clave = "Minutes played"
                valor = 0
                estadisticas[clave] = valor

                entrada = {
                    clave: valor
                }

                # Agrega la entrada JSON a la lista
                entradas.append(entrada)
                
                nombreJson=nombre.text
                puntuaciónJson=None
            
                # Crear el diccionario para el jugador
                JsonJugador = {
                        nombreJson: {
                            "puntuacion": puntuaciónJson,
                            "estadisticas": entradas
                        }
                }
                
                self.performance_to_json(JsonJugador)
                
                return

        except NoSuchElementException as e:
            try:
                manager= self.driver.find_element(By.XPATH,'//*[@id="__next"]/main/div[2]/div/div/div[1]/div[1]/div/div[2]/div[2]/div/div/span')
                self.output_textedit.append("Entrenador")
            except:
                self.output_textedit.append("No convocado")
        self.output_textedit.append("_______________________________")

    def scrapear_funcion(self):
        self.output_textedit.append("Conecting to API...")

        # GESTIÓN DEL INPUT DEL USUARIO
        selected_team = self.team_combobox.currentText()
        ruta_output = self.text_input.text()
        jornada=int(self.number_input.value())

        # Mostrar el valor en el QTextEdit
        self.output_textedit.append(f"Equipo seleccionado: {selected_team}")
        self.output_textedit.append(f"Jornada seleccionada: {jornada}")
        self.output_textedit.append(f"Ruta para la salida del scraper selecionada: {ruta_output}")

        if ruta_output=="":
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText("\n¡La ruta de la carpeta donde guardar los datos scrapeados no está inicializada!, Configúrala antes de empezar a scrapear")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return
        
        ############################################  eliminar excell a scrapear si ya existe !!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        ##############################################################################################################################################################
        # FASE 1: Obtener la url de la pagina web de Sofaescore de todas los partidos de LaLiga                                                                      #
        ##############################################################################################################################################################

        #######################################################################################################################################################
        #  PARTE 1 : Mediante una request a la API de Sofaescore obtenemos el id de todos los equipos de LaLiga para posteriores cnsultas                     #
        #######################################################################################################################################################

        url = "https://sofascore.p.rapidapi.com/teams/search"

        id_equipo = None
        valor_round = None
        slugJson = None

        self.output_textedit.append(f"________________________________________________________________________________________")
        self.output_textedit.append("Obteniendo id del equipo de LaLiga...")

        response = requests.get(url, headers=headers, params={"name": selected_team})
        data = response.json()
        if data['teams']:
            id_equipo = data['teams'][0]['id']
            self.output_textedit.append(f"GET ID {selected_team}: {id_equipo}")  

        #######################################################################################################################################################
        #  PARTE 2 : Mediante una request a la API de Sofaescore obtenemos informacion de todos los equipos de LaLiga                                         #
        #######################################################################################################################################################
        time.sleep(1)
        url = "https://sofascore.p.rapidapi.com/teams/get-last-matches"
        ultimo_partido_equipos = None
        liga="LaLiga"
        ulimo_partido=None

        self.output_textedit.append(f"________________________________________________________________________________________")  
        self.output_textedit.append(f"Obteniendo partido de la jornada {jornada} del {liga}...")

        querystring = {"teamId": str(id_equipo), "pageIndex": "0"}

        response = requests.get(url, headers=headers, params=querystring)

        if response.status_code == 200:
            data = response.json()
            partido=0
            while True: 
                valor_round = data['events'][partido]['roundInfo']['round']
                tournament = data['events'][partido]['tournament']['name']
                        
                if valor_round==jornada and tournament==liga:
                    nombre=data['events'][partido]['slug']
                    self.output_textedit.append(f"GET partido: {nombre}")
                    
                    ulimo_partido=data['events'][partido]

                    valor_round = data['events'][partido]['roundInfo']['round']
                    break
                partido+=1
        else:
             self.output_textedit.append(f"Error al obtener datos para el equipo {selected_team}. Código de estado: {response.status_code}")
                
        nombre_carpeta_jornada = "jornada_" + str(valor_round)+"_SF"
        self.ruta_jornada = os.path.join(ruta_output, nombre_carpeta_jornada)
        
        #######################################################################################################################################################       
        #  PARTE 3 : Mediante los datos obtenidos de la API cosntruimos la url de cada patido de todos los equipos de LaLiga                                  #
        #######################################################################################################################################################
        time.sleep(1)
        #Obtener componenets de la url   
        id = ulimo_partido.get('id')
        slug = ulimo_partido.get('slug')
        customId = ulimo_partido.get('customId')

        base_url = "https://www.sofascore.com/{}/{}#{}"

        self.output_textedit.append(f"________________________________________________________________________________________")  
        self.output_textedit.append("Generando urls del último partido de cada equipo...")

        #Fusionar elementos de la url
        url = base_url.format(id,customId,slug)
        self.output_textedit.append(url)

        ##############################################################################################################################################################
        # FASE 2: "Acceder y hacer scraping de todos las urls de todos los partidos de LaLiga                                                                        #
        ##############################################################################################################################################################

        #######################################################################################################################################################
        #  PARTE 1 : obtener el performance de los 22 jugadores titulares del partido                                                                         # 
        #    -abre la web                                                                                                                                     #
        #    -acepta las cookies                                                                                                                              #
        #    -hace click sobre de cada jugador para que emerja la tarjeta con los datos del performance del partido asociados a cada jugador y los extrae)    #
        #######################################################################################################################################################
        self.output_textedit.append(f"________________________________________________________________________________________")
        self.output_textedit.append("Starting scraper...")

        # Crea una instancia del controlador del navegador
        self.driver = webdriver.Chrome()
            
        # Maximizar la ventana del navegador
        self.driver.maximize_window()
            
        # Navega a la página web que deseas hacer scraping
        self.driver.get(url)

        # Espera a que se cargue la página
        self.driver.implicitly_wait(20)

        # Encuentra el botón de "Consentir" 
        button = self.driver.find_element(By.XPATH, '//button[@aria-label="Consentir"]')
        # Haz clic en el botón de "Consentir" 
        button.click()
        try:
            # Encuentra el botón de "Ask me later" 
            button = self.driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/button')
            # Haz clic en el botón de "Consentir" 
            button.click()
                
        except:
            pass
        time.sleep(45)
            
        # Reducir el nivel de zoom 
        #zoom_out_script = "document.body.style.zoom='60%';"
        #self.driver.execute_script(zoom_out_script)

        # Encuentra el nombre del partido" 
        local = self.driver.find_element(By.XPATH, '//*[@id="__next"]/main/div[2]/div[2]/div[1]/div[1]/div[1]/div[1]/div/div[1]/div/a/div/div/bdi')
        visitante = self.driver.find_element(By.XPATH, '//*[@id="__next"]/main/div[2]/div[2]/div[1]/div[1]/div[1]/div[1]/div/div[3]/div/a/div/div/bdi')
        # Concatenar los nombres para formar slugJson
        slugJsonConcatenado = f"{local.text}_{visitante.text}"
        slugJson = slugJsonConcatenado.replace(" ", "_")
            
        self.driver.quit()
            
        self.output_textedit.append(f"Scraping {slugJson}...")
        
        for i in range(22):

            # Crear carpeta para la jornada
            if not os.path.exists(self.ruta_jornada):
                os.makedirs(self.ruta_jornada)

            # Crea una instancia del controlador del navegador
            self.driver = webdriver.Chrome()
                
            # Maximizar la ventana del navegador
            self.driver.maximize_window()

            # Navega a la página web que deseas hacer scraping
            self.driver.get(url)

            # Espera a que se cargue la página
            self.driver.implicitly_wait(45)

            # Encuentra el botón de "Consentir" 
            button = self.driver.find_element(By.XPATH, '//button[@aria-label="Consentir"]')
            # Haz clic en el botón de "Consentir" 
            button.click()

            try:
                # Encuentra el botón de "Ask me later" 
                button = self.driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/button')
                # Haz clic en el botón de "Consentir" 
                button.click()
            except:
                pass
                
            # Reducir el nivel de zoom 
            #zoom_out_script = "document.body.style.zoom='60%';"
            #self.driver.execute_script(zoom_out_script)
            time.sleep(45)
                
            # Encuentra todos los elementos <a> con la clase 'sc-3937c22d-0 jrbLdB'
            divJugadores = self.driver.find_elements(By.XPATH, '//a[@class="sc-3937c22d-0 jrbLdB"]')
                
            numTitulares=len(divJugadores)
            self.output_textedit.append(f"{i+1}/{numTitulares}")
            divJugadores[i].click()
            time.sleep(45)
            self.obtener_informacion_jugador()
            self.progress += 1
            self.invocar_actualizacion(self.progress)

            self.driver.quit()
        
        #######################################################################################################################################################
        #  PARTE 2 :  obtener el performance de los jugadores que entraron de cambio al partido / suplentes / lesionados                                      #
        #######################################################################################################################################################

        self.progress_bar.setValue(0) 
        jugador=0

        while True:

            # Crea una instancia del controlador del navegador
            self.driver = webdriver.Chrome()
            # Maximizar la ventana del navegador
            self.driver.maximize_window()
            # Navega a la página web que deseas hacer scraping
            self.driver.get(url)
            # Espera a que se cargue la página
            self.driver.implicitly_wait(45)
            # Encuentra el botón de "Consentir" 
            button = self.driver.find_element(By.XPATH, '//button[@aria-label="Consentir"]')
            # Haz clic en el botón de "Consentir" 
            button.click()
            try:
                # Encuentra el botón de "Ask me later" 
                button = self.driver.find_element(By.XPATH, '//*[@id="__next"]/div[3]/div[2]/button')
                # Haz clic en el botón de "Consentir" 
                button.click()
            except:
                pass
            # Reducir el nivel de zoom 
            #zoom_out_script = "document.body.style.zoom='80%';"
            #self.driver.execute_script(zoom_out_script)

            # Encuentra todos los elementos 
            divSuplentes1 = self.driver.find_elements(By.XPATH, '//div[@display="flex" and contains(@class, "sc-fqkvVR sc-dcJsrY bASBgQ kHiXJe")]')
            divSuplentes2 = self.driver.find_elements(By.XPATH, '//div[@display="flex" and contains(@class, "sc-fqkvVR sc-dcJsrY bphgaj kHiXJe")]')
            
            # Crear un nuevo array y combinar los elementos de divSuplentes1 y divSuplentes2
            divSuplentes = []
            divSuplentes.extend(divSuplentes1)
            divSuplentes.extend(divSuplentes2)
            
            tamaño_divSuplentes = len(divSuplentes)
            self.progress_bar.setRange(0, tamaño_divSuplentes)
            self.output_textedit.append(f"{jugador+1}/{tamaño_divSuplentes}")
            
            # Definir los textos a buscar en los subelementos
            textos_a_buscar = ["Out", "Doubtful"]

            elemento=divSuplentes[jugador]
                
            for texto_a_buscar in textos_a_buscar:
                # Verificar si el elemento contiene un subelemento con el texto actual
                subelemento = elemento.find_elements(By.XPATH, f'.//div[contains(text(), "{texto_a_buscar}")]')
                
                if subelemento:
                    self.output_textedit.append(f"Se detectó jugador con status '{texto_a_buscar}' .")
                    entradas = []
                    estadisticas = {}
                    clave = "Ausencia"
                    valor = "Injury"
                    estadisticas[clave] = valor

                    entrada = {
                        clave: valor
                    }
                    # Agrega la entrada JSON a la lista
                    entradas.append(entrada)
                    
                    nombre=divSuplentes[jugador].text
                    # Palabra que deseas eliminar
                    palabra_a_eliminar = "Out"
                    nombrefi=nombre.replace(palabra_a_eliminar, "")
                    palabra_a_eliminar = "Doubtful"
                    nombreJson = nombrefi.replace(palabra_a_eliminar, "")
                    self.output_textedit.append(nombreJson)
                    
                    puntuaciónJson=None
            
                    # Crear el diccionario para el jugador
                    JsonJugador = {
                        nombreJson: {
                            "puntuacion": puntuaciónJson,
                            "estadisticas": entradas
                        }
                    }

                    self.performance_to_json(JsonJugador)
            
            # Definir los textos a buscar en los subelementos
            textos_a_buscar = ["Suspended"]

            elemento=divSuplentes[jugador]
            
            for texto_a_buscar in textos_a_buscar:
                # Verificar si el elemento contiene un subelemento con el texto actual
                subelemento = elemento.find_elements(By.XPATH, f'.//div[contains(text(), "{texto_a_buscar}")]')

                if subelemento:
                    self.output_textedit.append(f"Se detectó jugador con status'{texto_a_buscar}' .")
                    
                    entradas = []
                    estadisticas = {}
                    clave = "Ausencia"
                    valor = "Suspension"
                    estadisticas[clave] = valor

                    entrada = {
                        clave: valor
                    }
                    # Agrega la entrada JSON a la lista
                    entradas.append(entrada)
                    
                    nombre=divSuplentes[jugador].text
                    # Palabra que deseas eliminar
                    palabra_a_eliminar = "Suspended"
                    # Eliminar la palabra del string
                    nombreJson = nombre.replace(palabra_a_eliminar, "")
                    self.output_textedit.append(nombreJson)
                    
                    puntuaciónJson=None
            
                    # Crear el diccionario para el jugador
                    JsonJugador = {
                        nombreJson: {
                            "puntuacion": puntuaciónJson,
                            "estadisticas": entradas
                        }
                    }

                    self.performance_to_json(JsonJugador)
            
            divSuplentes[jugador].click()
            time.sleep(45)
            self.obtener_informacion_jugador()
            self.progress += 1
            self.invocar_actualizacion(self.progress)   

            jugador += 1 
            
            if jugador >= len(divSuplentes):
                self.driver.quit()
                break
                
            self.driver.quit()

class PlayerScraperWindowMF(QDialog, QWidget):
    def __init__(self, window_title):
        super().__init__()
        self.setWindowTitle(window_title)

        self.teams_data = {
        "Real Madrid": "https://cdn.gomister.com/file/cdn-common/teams/15.png?version=20231117",
        "Real Sociedad": "https://cdn.gomister.com/file/cdn-common/teams/16.png?version=20231117",
        "Atlético de Madrid": "https://cdn.gomister.com/file/cdn-common/teams/2.png?version=20231117",
        "Girona": "https://cdn.gomister.com/file/cdn-common/teams/222.png?version=20231117",
        "Osasuna": "https://cdn.gomister.com/file/cdn-common/teams/50.png?version=20231117",
        "Athletic Club": "https://cdn.gomister.com/file/cdn-common/teams/1.png?version=20231117",
        "Valencia": "https://cdn.gomister.com/file/cdn-common/teams/19.png?version=20231117",
        "Granada": "https://cdn.gomister.com/file/cdn-common/teams/10.png?version=20231117",
        "Getafe": "https://cdn.gomister.com/file/cdn-common/teams/9.png?version=20231117",
        "Villarreal": "https://cdn.gomister.com/file/cdn-common/teams/20.png?version=20231117",
        "Las Palmas": "https://cdn.gomister.com/file/cdn-common/teams/11.png?version=20231117",
        "Mallorca": "https://cdn.gomister.com/file/cdn-common/teams/408.png?version=20231117",
        "Rayo Vallecano": "https://cdn.gomister.com/file/cdn-common/teams/14.png?version=20231117",
        "Barcelona": "https://cdn.gomister.com/file/cdn-common/teams/3.png?version=20231117",
        "Celta de Vigo": "https://cdn.gomister.com/file/cdn-common/teams/5.png?version=20231117",
        "Cádiz": "https://cdn.gomister.com/file/cdn-common/teams/499.png?version=20231117",
        "Alavés": "https://cdn.gomister.com/file/cdn-common/teams/48.png?version=20231117",
        "Almería": "https://cdn.gomister.com/file/cdn-common/teams/21.png?version=20231117",
        "Sevilla": "https://cdn.gomister.com/file/cdn-common/teams/17.png?version=20231117",
        "Betis": "https://cdn.gomister.com/file/cdn-common/teams/4.png?version=20231117",
        }

        self.progress = 0

        # Crear un diseño de cuadrícula
        layout = QGridLayout(self)
        # Establecer el tamaño máximo de la segunda columna
        layout.setColumnStretch(1, 1)
        layout.setColumnStretch(2, 1)

        # Variables para almacenar la carpeta y la ruta seleccionadas
        self.selected_folder = ""
        self.selected_path = ""
        
        self.driver = None

        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("Mister Fantasy Scraper")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        layout.addWidget(label_text, 0, 0,1, 2)

        # LABEL SUBTÍTULO
        label_subtext = QLabel("Obtener el listado de todos los jugaodres disponibles en la web de Mister Fantasy MD y toda su informaciónm y estadísticas asociada.")
        layout.addWidget(label_subtext, 1, 0, 1, 2)


        ### SELECCIONAR JORNADA INPUT ####################################################
        # INPUT NÚMERO JORNADA 
        label_number = QLabel("Jornada a scrapear:")
        layout.addWidget(label_number, 2, 0)
        # Estilos 
        self.number_input = QSpinBox(self)
        self.number_input.setMinimum(1)  # Establecer el valor mínimo (jornada 1)
        self.number_input.setMaximum(38)  # Establecer el valor máximo (Jornada 38)
        self.number_input.setSingleStep(1)  # Establecer el paso
        self.number_input.setMaximumSize(38, 20)
        self.number_input.setMinimumSize(38, 20)
        # Aliniación
        layout.addWidget(self.number_input, 2, 1)
        

        #------- GAP vacio -----------------------------------------
        empty_widget = QWidget()
        empty_widget.setFixedHeight(10)  # Tamaño del gap (10 px)
        layout.addWidget(empty_widget, 3, 0)
        #-----------------------------------------------------------


        ###  SELECCIONAR RUTA DONDE GUARDAR EL EXCEL OUTPUT DEL SCRAPER  #################
        # LABEL TEXTO 
        label_text = QLabel("Ruta output scraper:")
        layout.addWidget(label_text, 4, 0)

        # INPUT TEXTO (QLineEdit en lugar de QSpinBox)
        self.text_input = QLineEdit(self)
        # Alineación
        layout.addWidget(self.text_input, 4, 1)
        # Estilos 
        self.text_input.setMinimumWidth(750)
        

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación  
        layout.addWidget(select_folder_button, 5, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos 
        select_folder_button.setMinimumWidth(140)

        #------- GAP vacio -----------------------------------------
        empty_widget = QWidget()
        empty_widget.setFixedHeight(10)  # Tamaño del gap (10 px)
        layout.addWidget(empty_widget, 6, 0)
        #-----------------------------------------------------------


        ###  BOTÓN PARA INICIAR SCRAPER  ################################################
        # Crear un botón llamado "Scrapear"
        scrape_button = QPushButton("Scrapear")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar barra progreso
        scrape_button.clicked.connect(self.iniciar_scrapear_thread)
        scrape_button.clicked.connect(self.start_progress)

        # Alineación 
        layout.addWidget(scrape_button, 7, 0)
        # Estilos
        self.number_input.setMaximumSize(38, 20)


        ###  BARRA DE PROGRESO  ################################################
        # Crear Barra de progreso
        self.progress_bar = QProgressBar(self)
        layout.addWidget(self.progress_bar)


        ###  VENTANA OUTPUT SCRAPER  ####################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        layout.addWidget(self.output_textedit, 8, 0, 9, 0)  # row, column, rowSpan, columnSpan


        ###  ESTABLECER DISEÑO DE LA VENTANA  ###########################################
        self.setMinimumSize(500, 500) # Configurar el tamaño mínimo de la ventana
        # Configurar el diseño para la ventana
        self.setLayout(layout)

        # Configurar el título de la ventana
        self.setWindowTitle("Mister Fantasy Mundo Deportivo Scraper")

        # Evento cerrar ventana 
        self.destroyed.connect(self.cleanup)

    def cleanup(self):
        # Realizar cualquier limpieza necesaria aquí
        QApplication.quit()

    def start_progress(self):
        # Establecer el rango de la barra de progreso según tus necesidades
        self.progress_bar.setRange(0, 511)

        ruta_output = self.text_input.text()
        if ruta_output!="":
            self.progress_bar.setValue(0)
    
    def click_mas(self):
        # Pinchar en el botón del menu "Más"
        masMenu = self.driver.find_element(By.XPATH, '//*[@id="content"]/header/div[2]/ul/li[5]/a')

        try:
            masMenu.click()
        except (ElementNotInteractableException, NoSuchElementException):
            # Maneja la excepción y espera antes de intentar nuevamente
            self.output_textedit.insertPlainText("Anuncio detectado, reiniciando driver...")
            self.driver.refresh()
            time.sleep(3) 
            masMenu.click()

    def actualizar_version(self,version):
      for equipo, url in self.teams_data.items():
        # Dividir la URL en base al signo de interrogación
        partes = url.split('?')
        
        # Verificar si hay una parte después del signo de interrogación y actualizar la versión
        if len(partes) > 1:
            partes[1] = f"version={version}"
            
            # Volver a unir las partes para formar la URL actualizada
            nueva_url = '?'.join(partes)
            
            # Actualizar la URL en el diccionario
            self.teams_data[equipo] = nueva_url

        #print(version)
        #print("nuevaaaa url-->  ",nueva_url)

    def obtener_valor_por_etiqueta(self,label_deseado):
        # Función para obtener el valor basado en la etiqueta
        elemento = self.driver.find_element(By.XPATH, f"//div[@class='item']//div[@class='label' and text()='{label_deseado}']/following-sibling::div[@class='value']")
        valor = elemento.text
        return valor

    def extraer_info_jugador(self,jornada_absolute,jornada_a_scrapear):
        
        nombre = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[2]")
        apellido = self.driver.find_element(By.XPATH, " /html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[3]")
        valorS= self.driver.find_element(By.XPATH,'/html/body/div[6]/div[3]/div[2]/div[2]/div/div/div[1]/div[2]')
        valor=valorS.text

        media_puntos_local = self.obtener_valor_por_etiqueta("Media en casa")
        media_puntos_visitante = self.obtener_valor_por_etiqueta("Media fuera")
        try:
            edad = self.obtener_valor_por_etiqueta("Edad")
            altura = self.obtener_valor_por_etiqueta("Altura")
            peso = self.obtener_valor_por_etiqueta("Peso")
        except:
            edad = None
            altura = None
            peso = None
                
        if peso == "kg":
            peso = None

            
        #### OBTENER EQUIPO JUGADOR ####

        # Obtener src del equipo
        team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[1]/a/img")
        image_url = team_logo_element.get_attribute("src")

        # Comparar la URL de la imagen con las URLs en teams_data
        equipo = None
        proximo_rival=None
        local= False
        for equipo_nombre, equipo_url in self.teams_data.items():
            if image_url == equipo_url:
                equipo = equipo_nombre
                
                #### OBTENER RESULTADO ÚLTIMO PARTIDO ####
                try:
                    divpartido = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div[1]/div[3]/div")
                except:
                    divpartido = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div/div[2]/div")
                
                # Encuentra el div del partido
                item_elements = divpartido.find_elements(By.CLASS_NAME, 'item')
            
                # Encuentra las imágenes dentro del div partido
                img_elements = item_elements[0].find_elements(By.CLASS_NAME, 'team-logo')

                # Guarda las src de las imágenes en variables
                if len(img_elements) >= 2:
                    src_img1 = img_elements[0].get_attribute('src')
                    src_img2 = img_elements[1].get_attribute('src')
                    if src_img1 == image_url:
                        local = True
                        for equipo_nombre, equipo_url in self.teams_data.items():
                            if src_img2 == equipo_url:
                                proximo_rival=equipo_nombre
                    else:
                        local=False
                        for equipo_nombre, equipo_url in self.teams_data.items():
                            if src_img1 == equipo_url:
                                proximo_rival=equipo_nombre
                else:
                    self.output_textedit.insertPlainText("No se encontro el próximo partido")
                

        #### OBTENER POSICIÓN DEL JUGADOR ####
        elemento = self.driver.find_element(By.XPATH, '//i[contains(@class, "pos-")]')
        # Obtener el valor del atributo class
        clases = elemento.get_attribute("class").split()

        # Determinar la posición
        posicion = None
        for clase in clases:
            if clase.startswith("pos-") and "pos-big" in clases:
                if clase == "pos-1":
                    posicion = "PT"
                elif clase == "pos-2":
                    posicion = "DF"
                elif clase == "pos-3":
                    posicion = "MC"
                elif clase == "pos-4":
                    posicion = "DL"
                break

            
        #### OBTENER PUNTOS DEL JUGADOR ####
        # Encontrar jornada 
        elementos_principales = self.driver.find_elements(By.CLASS_NAME, 'btn-player-gw')

        # Iterar sobre cada elemento encontrado
        subelemento_gw=None
        jornada_name=None
        for elemento_principal in elementos_principales:
            # Encontrar subelemento con la clase 'gw' dentro de cada elemento principal
            subelemento_gw = elemento_principal.find_element(By.CLASS_NAME, 'gw')

            # Verificar si el texto coincide con el de la jornada
            if subelemento_gw.text == jornada_a_scrapear:
                jornada_name = subelemento_gw.text
                break             
        
        if jornada_name ==jornada_absolute:
            # Encontrar jornada en la web con otro elemennto como referencia
            localizador = self.driver.find_element(By.XPATH, "//h4[text()='Valor']")
            self.driver.execute_script("arguments[0].scrollIntoView(true);", localizador)   
            
            time.sleep(1)
            
            try:
                subelemento_gw.click()
            except:
                elemento_principal.click()
            
            time.sleep(2)
            
            try:
                # PUNTOS MISTER FANTASY
                main_provider = self.driver.find_element(By.CLASS_NAME, 'main-provider')
                points_element = main_provider.find_element(By.CLASS_NAME, 'points')
                final_points = points_element.get_attribute('data-points')

                # PUNTOS AS, MARCA Y MUNDO DEPORTIVO 
                providers_div = self.driver.find_element(By.CLASS_NAME, "providers")
                li_elements = providers_div.find_elements(By.TAG_NAME, "li")

                points_array = []

                for li_element in li_elements:

                    points_div = li_element.find_element(By.CLASS_NAME, "points")
                    points_value = points_div.text
                    points_array.append(points_value)

                as_points=points_array[0]
                marca_points=points_array[1]
                mundo_deportivo_points=points_array[2]
                
                #### OBTENER PARTIDO ANTERIOR ####
                # Encontrar el div principal con la clase "player-match"
                player_match_div = self.driver.find_element(By.CLASS_NAME, "player-match")

                # Encontrar los subelementos dentro del div principal
                team_1 = player_match_div.find_element(By.CLASS_NAME, "left").find_element(By.CLASS_NAME, "team").text
                goals_team_1 = [int(goal.text) for goal in player_match_div.find_elements(By.CLASS_NAME, "goals")][0]  
                goals_team_2 = [int(goal.text) for goal in player_match_div.find_elements(By.CLASS_NAME, "goals")][1]  
                team_2 = player_match_div.find_element(By.CLASS_NAME, "right").find_element(By.CLASS_NAME, "team").text

                if team_1 == equipo:
                    ultimo_rival=team_2

                    if goals_team_1 > goals_team_2:
                        result = "Win"
                    elif goals_team_1 < goals_team_2:  
                        result = "Loss"
                    else:
                        result = "Draw"
                else:
                    ultimo_rival=team_1

                    if goals_team_1 > goals_team_2:
                        result = "Loss"
                    elif goals_team_1 < goals_team_2:  
                        result = "Win"
                    else:
                        result = "Draw"

                self.driver.back()
                
            except:
                final_points=None
                as_points=None
                marca_points=None
                mundo_deportivo_points=None
                ultimo_rival=None
                result=None
                
        else:
            final_points="NA"
            as_points="NA"
            marca_points="NA"
            mundo_deportivo_points="NA"
            ultimo_rival="NA"
            result="NA"
        

        #### IMPRIMIR TODOS LOS DATOS ####
        self.output_textedit.append("_____________________________________________")
        time.sleep(0.2)
        self.output_textedit.append(f"-{self.progress+1}. {nombre.text}, {apellido.text}")
        time.sleep(0.2)
        self.output_textedit.append(f"Valor: {valor}")
        time.sleep(0.2)
        self.output_textedit.append(f"Posición: {posicion}")
        time.sleep(0.2)
        self.output_textedit.append(f"Equipo: {equipo}")
        time.sleep(0.2)
            
        self.output_textedit.append("- - - - - - - - - - - - - - - - - - - - - - - - - -")
        time.sleep(0.2)

        self.output_textedit.append(f"Puntuación Fantasy: {final_points}")
        time.sleep(0.2)
        self.output_textedit.append(f"Puntuación Fantasy: {as_points}")
        time.sleep(0.2)
        self.output_textedit.append(f"Puntuación Marca: {marca_points}")
        time.sleep(0.2)
        self.output_textedit.append(f"Puntuación Mundo Deportivo: {mundo_deportivo_points}")
        time.sleep(0.2)
        
        self.output_textedit.append("- - - - - - - - - - - - - - - - - - - - - - - - - -")
        time.sleep(0.2)
            
        self.output_textedit.append(f"Último rival: {ultimo_rival}")
        time.sleep(0.2)
        self.output_textedit.append(f"Resultado del partido: {result}")
        time.sleep(0.2)

        self.output_textedit.append(f"Próximo rival: {proximo_rival}")
        time.sleep(0.2)
        self.output_textedit.append(f"Próximo partido es local: {local}")
        time.sleep(0.2)
        self.output_textedit.append(f"Media en casa: {media_puntos_local}")
        time.sleep(0.2)
        self.output_textedit.append(f"Media fuera: {media_puntos_visitante}")
        time.sleep(0.2)
        self.output_textedit.append(f"Edad: {edad}")
        time.sleep(0.2)
        self.output_textedit.append(f"Altura: {altura}")
        time.sleep(0.2)
        self.output_textedit.append(f"Peso: {peso}")
        time.sleep(0.2)

        self.progress += 1
        self.invocar_actualizacion(self.progress)

        #Definir ruta donde guardar el output del scraper
        ruta_output = self.text_input.text()
        save_as=f"{ruta_output}/"+jornada_absolute+".xlsx"
        self.output_textedit.append(save_as)
        jugador = nombre.text + " " + apellido.text

        try:
            wb = openpyxl.load_workbook(save_as)
        except FileNotFoundError:
            # Crear un nuevo libro de trabajo y una hoja
            wb = openpyxl.Workbook()
            sheet = wb.active
            encabezado = ["Jugador", "Valor", "Posición", "Equipo", "Puntuación Fantasy", "Puntuación AS", "Puntuación Marca", "Puntuación Mundo Deportivo", "Último rival", "Resultado del partido", "Próximo rival", "Próximo partido es local", "Media en casa", "Media fuera", "Edad", "Altura", "Peso"]
            sheet.append(encabezado)
            # Guardar el archivo Excel
            wb.save(save_as)

        # Seleccionar la hoja activa
        sheet = wb.active

        # Lista de variables a almacenar
        nueva_fila = [jugador, valor, posicion, equipo, final_points, as_points, marca_points, mundo_deportivo_points, ultimo_rival, result, proximo_rival, local, media_puntos_local, media_puntos_visitante, edad, altura, peso]

        # Escribir la nueva fila en la hoja de cálculo
        sheet.append(nueva_fila)

        # Guardar el archivo Excel
        wb.save(save_as)
        
    def iniciar_scrapear_thread(self):  
        # Crear un hilo y ejecutar la función en segundo plano
        if usuario!="":
            thread = threading.Thread(target=self.scrapear_funcion)
            thread.start()
        else:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('No has iniciado sesion en la aplicación. Logueate con tus credenciales de Mister Fantasy MD en la ventana Perfil para acceder al mercado de jugaodes.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)


    def invocar_actualizacion(self, nuevo_valor):
        QMetaObject.invokeMethod(self.progress_bar, "setValue", Qt.ConnectionType.QueuedConnection, Q_ARG(int, nuevo_valor))

    def scrapear_funcion(self):
        self.output_textedit.append("Starting scraper...")

        # GESTIÓN DEL INPUT DEL USUARIO
        # Obtener el valor de la jornada desde el QSpinBox
        numero_jornada = str(self.number_input.value())

        # Concatena 'J' delante del número
        jornada_a_scrapear = 'J' + numero_jornada

        # Mostrar el valor en el QTextEdit
        self.output_textedit.append(f"Jornada seleccionada: {jornada_a_scrapear}")

        ruta_output = self.text_input.text()
        self.output_textedit.append(f"Ruta para la salida del scraper selecionada: {ruta_output}")
        self.output_textedit.append(f"________________________________________________________________________________________")

        if ruta_output=="":
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText("\n¡La ruta de la carpeta donde guardar los datos scrapeados no está inicializada!, Configúrala antes de empezar a scrapear")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            self.output_textedit.append(f"________________________________________________________________________________________")
            return
        
        rutaDel=f"{ruta_output}/"+jornada_a_scrapear+".xlsx"
        if os.path.exists(rutaDel):
             os.remove(rutaDel)

        self.driver = webdriver.Chrome()
        realizar_login(self.driver)
        time.sleep(5)

        # Espera a que se cargue la página
        self.driver.implicitly_wait(10)

        #Hacer click en el btn Jugadores con la función click_mas() para manejar errores generados por anuncios intrusivos
        self.click_mas()

        # Pinchar en el botón "Jugaodres" para acceder al listado de jugadores 
        jugadoresbtn = self.driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div[1]/button[2]')

        try:
            jugadoresbtn.click()
        except (ElementNotInteractableException, NoSuchElementException):
            # Maneja la excepción y espera antes de intentar nuevamente
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText("\nAnuncio detectado, reiniciando driver...")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            self.driver.refresh()
            time.sleep(3)
            self.click_mas()
            time.sleep(3)
            try:
                jugadoresbtn = self.driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div[1]/button[2]')
                jugadoresbtn.click()
            except: 
                self.output_textedit.append("Reinicia el script :(")
                sys.exit()

        pag=2
        index=0
        absolute=1
        jornada_absolute=""
        while True:

            # Encontrar todos los elementos li
            elementos_lis = self.driver.find_elements(By.XPATH, "/html/body/div[6]/div[3]/div[3]/ul/li")

            # Longitud de la lista de elementos encontrados
            length=len(elementos_lis)

            while index < length:
                # Encontrar todos los elementos li
                elementos_li = self.driver.find_elements(By.XPATH, "/html/body/div[6]/div[3]/div[3]/ul/li")
                elementos_li[index].click()

                time.sleep(1)
                
                try:
                    team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[2]/div[1]/div/div[1]/div[1]/a/img")
                except:
                    try:
                        team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div/div[3]/div/div[1]/div[2]/img[1]")
                    except:
                        team_logo_element = self.driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div[3]/div/div[3]/div/div[1]/div[2]/img[2]")
                
                image_url = team_logo_element.get_attribute("src")
                # Dividir la URL utilizando el signo de igual como delimitador
                parts = image_url.split('=')
                # El valor de version está en la segunda parte después del =
                version = parts[1]
                self.actualizar_version(version)
                
                if absolute == 1:
                    # Encontrar jornada 
                    elementos_principales = self.driver.find_elements(By.CLASS_NAME, 'btn-player-gw')

                    # Iterar sobre cada elemento encontrado
                    subelemento_gw=None
                    for elemento_principal in elementos_principales:
                        # Encontrar subelemento con la clase 'gw' dentro de cada elemento principal
                        subelemento_gw = elemento_principal.find_element(By.CLASS_NAME, 'gw')
                        
                        # Verificar si el texto coincide con el de la jornada
                        if subelemento_gw.text == jornada_a_scrapear:
                            jornada_absolute = subelemento_gw.text
                            break   
                absolute = 2
                
                self.extraer_info_jugador(jornada_absolute,jornada_a_scrapear)
                
                #Retroceder página
                self.driver.back()
                time.sleep(1)
                elementos_li = self.driver.find_elements(By.XPATH, "/html/body/div[6]/div[3]/div[3]/ul/li")
                if index == 0:
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", elementos_li[index])
                else:
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", elementos_li[index-1])
                time.sleep(1)
                index += 1

            #Pulsar Ver más
            try:
                ver_mas = self.driver.find_element(By.XPATH, '/html/body/div[6]/div[3]/div[3]/div[1]/button')
                ver_mas.click()
                time.sleep(4)
            except:
                break

            #Jugador cambio de pagina
            elementos_li = self.driver.find_elements(By.XPATH, "/html/body/div[6]/div[3]/div[3]/ul/li")
            elementos_li[index].click()
            time.sleep(2)
            self.extraer_info_jugador(jornada_absolute,jornada_a_scrapear)
            self.driver.back()
            
            self.output_textedit.append("____________________________________")
            self.output_textedit.append("------------------------------------")
            self.output_textedit.append(f"Siguiente página... ({pag})")
            self.output_textedit.append("------------------------------------")
            
            index=1
            pag+=1

        self.driver.quit()    
        self.output_textedit.append("Todos los jugadores scrapeados")


class trainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.progress = 0

        # Crear un diseño de cuadrícula dentro del QVBoxLayout
        grid_layout = QGridLayout(self)

        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("ENTRENAR MODELEO")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        grid_layout.addWidget(label_text, 0, 0)

        # LABEL SUBTÍTULO 1
        label_subtext1 = QLabel("Pruba con los diferentes algoritmos disponibles a entrenar varios modelo y conparar entre ellos su desenpeño para selecionar el que mejores predicciones realice. ")
        grid_layout.addWidget(label_subtext1, 1, 0, 1, 2)

        ### SELECCIONAR RUTA DATASET DE ENTRADA ##################################################
        # LABEL DE TEXTO
        label_text = QLabel("Selecionar ruta de la carpeta de los datasets de entrada de cada jornada: ")
        grid_layout.addWidget(label_text, 2, 0, 1, 2)

        # INPUT DE TEXTO
        self.text_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input, 3, 0, 1, 2)

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 4, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        ### SELECTOR ALGORITMO A USAR ##################################################
        # LABEL DE TEXTO
        label_text = QLabel("Seleciona un algoritmo de entrenamiento: ")
        grid_layout.addWidget(label_text, 5, 0)

        self.combo_box = QComboBox()
        self.combo_box.addItem("Gradient Boosted Tree model")
        self.combo_box.addItem("K-NN model")
        self.combo_box.addItem("Linear Regression model")
        # Establecer el ancho máximo para la QComboBox
        self.combo_box.setMaximumWidth(185)
        grid_layout.addWidget(self.combo_box, 5, 1)

        ### SELECTOR ATRIBUTO A PREDECIR ###########################################################
        label_choice = QLabel("Seleccionar atributo del jugador predecir:")
        grid_layout.addWidget(label_choice, 6, 0)

        self.combo_box1 = QComboBox()
        self.combo_box1.addItem("Entrenar para predecir valor de mercado que alcanzará un jugaodr en la próxima jornada")
        self.combo_box1.addItem("Entrenar para predecir puntos que obtendrá un jugaodr en la próxima jornada")
        
        # Establecer el ancho máximo para la QComboBox
        self.combo_box1.setMaximumWidth(500)
        grid_layout.addWidget(self.combo_box1, 6, 1)

        ###  SELECCIONAR RUTA DONDE GUARDAR EL MODELO  ###################################
        # LABEL TEXTO 
        label_text = QLabel("Ruta donde guardar el modelo generado:")
        grid_layout.addWidget(label_text, 7, 0)

        # INPUT TEXTO (QLineEdit en lugar de QSpinBox)
        self.text_input2 = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input2, 8, 0, 1, 2)
        # Estilos 
        self.text_input2.setMinimumWidth(350)

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder2(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 9, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        self.selected_option = None
        ### BOTÓN PARA EMPEZAR ENTRENAMIENTO ###########################################################
        # Crear un botón
        self.scrape_button = QPushButton("Iniciar entrenamiento del modelo")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.scrape_button.clicked.connect(self.iniciar_thread)

        # Alineación y estilos
        grid_layout.addWidget(self.scrape_button, 10, 0,alignment=Qt.AlignmentFlag.AlignLeft)
        self.scrape_button.setMinimumWidth(225)

        ###  BARRA DE PROGRESO  ################################################
        # Crear Barra de progreso
        self.progress_bar = QProgressBar(self)
        grid_layout.addWidget(self.progress_bar,10,1)

        ###  VENTANA OUTPUT TRAIN  ####################################################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        grid_layout.addWidget(self.output_textedit, 11, 0, 10, 0)  # row, column, rowSpan, columnSpan

        ###  GRÁFICA  ####################################################################################
        # Crear una instancia de la figura de Matplotlib y el widget de lienzo
        #self.fig = Figure()
        #self.canvas = FigureCanvas(self.fig)
        #grid_layout.addWidget(self.canvas,8,1,alignment=Qt.AlignmentFlag.AlignRight)

        # Crear un subplot vacío
        #self.ax = self.fig.add_subplot(111)

        # Ajustar el tamaño de la gráfica
        #self.canvas.setFixedSize(250, 200)

        # Ocultar la gráfica inicialmente
        #self.canvas.setVisible(False)

    def start_progress(self):
        # Establecer el rango de la barra de progreso según tus necesidades
        self.progress_bar.setRange(0, 10)

        ruta_output = self.text_input.text()
        if ruta_output!="":
            self.progress_bar.setValue(0)

    def invocar_actualizacion(self, nuevo_valor):
        QMetaObject.invokeMethod(self.progress_bar, "setValue", Qt.ConnectionType.QueuedConnection, Q_ARG(int, nuevo_valor))

    def iniciar_thread(self):  
        # Crear un hilo y ejecutar la función en segundo plano
        thread = threading.Thread(target=self.train_function)
        thread.start()

    def train_function(self):
        def visualizar_distribucion(df, columna, etiqueta_x, titulo, borde_color):
            print(etiqueta_x)

            # Visualizar la distribución de las puntuaciones
            data = df[columna].dropna()

            sns.set_style("whitegrid")  # Configurar el estilo de la gráfica
            sns.histplot(data, bins=30, kde=False)  # Cambiando kde a False

            # Configurar título y etiquetas
            plt.title(titulo)
            plt.xlabel(etiqueta_x)
            plt.ylabel('Frecuencia')

            # Ajustar las propiedades del gráfico de barras para que no estén conectadas
            bars = plt.gca().patches
            for bar in bars:
                bar.set_width(0.8)  # Ancho de la barra
                bar.set_edgecolor(borde_color)  # Color del borde de la barra (ajusta según tus preferencias)

            # Mostrar la gráfica
            plt.show()
            time.sleep(1)

        def plot_boxplot(df, columna, titulo, etiqueta_x):
            # Visualizar el boxplot de la columna de interés
            data = df[columna].dropna()

            sns.set_style("whitegrid")  # Configurar el estilo de la gráfica
            sns.boxplot(x=data)

            # Configurar título y etiquetas
            plt.title(titulo)
            plt.xlabel(etiqueta_x)
            plt.ylabel('Valor')

            # Ajustar las propiedades del boxplot
            boxes = plt.gca().artists
            for box in boxes:
                box.set_linewidth(2)  # Ancho del borde

            # Mostrar la gráfica
            plt.show()
            time.sleep(1)
        
        def plot_lineplot(x_values, y_values, title, xlabel, ylabel):
            
            plt.plot(x_values, y_values, marker='o')
            plt.title(title)
            plt.xlabel(xlabel)
            plt.ylabel(ylabel)
            plt.show()
            time.sleep(1)
        
        def plot_correlation_matrix(df, title):
            # Calcular la matriz de correlación
            correlation_matrix = df.corr()

            # Configurar el estilo de la gráfica
            sns.set(style='whitegrid')

            # Crear una figura y un eje
            plt.figure(figsize=(10, 8))

            # Crear la matriz de correlación utilizando seaborn
            sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt=".2f", linewidths=.5)

            # Configurar título
            plt.title(title)

            # Mostrar la gráfica
            plt.show()

        #Lista de hilos que se van iniciando
        hilos=[]

        self.start_progress()
        #### PARTE 0 : LEER INPUTS + COMPROBAR QUE TODAS LOS INPUTS (rutas de archivos y carpetas) HAN SIDO INICIALIZADAS #####################################################################
        # Ruta a la carpeta que contiene los archivos json de Sofaescore
        carpeta_datasets = self.text_input.text()
        carpeta_save=self.text_input2.text()
        if not carpeta_datasets or not carpeta_save:
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            self.output_textedit.mergeCurrentCharFormat(formato_rojo)
            if not carpeta_datasets:
                self.output_textedit.insertPlainText("La ruta de la carpeta del dataset de entrenamiento no ha sido inicializada.\n")
            
            if not carpeta_save:
                self.output_textedit.insertPlainText("La ruta de la carpeta donde guardar el modelo generado no ha sido inicializada.\n")
            
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            self.output_textedit.mergeCurrentCharFormat(formato_negro)
            return

        # FASE 1: fusionar todos los dataset de entrada de cada jornada selecionados en uno solo #########################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Generando dataset de entrada...\n")
        carpeta_datasets = self.text_input.text()
        try:
            # Obtener la lista de archivos en la carpeta de entrada
            archivos_excel = [archivo for archivo in os.listdir(carpeta_datasets) if archivo.endswith('.xlsx')]

            # Comprobar si hay archivos Excel en la carpeta
            if not archivos_excel:
                output_textedit = self.output_textedit
                color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
                formato_rojo = QTextCharFormat()
                formato_rojo.setForeground(color_rojo)
                output_textedit.mergeCurrentCharFormat(formato_rojo)
                self.output_textedit.insertPlainText("No hay archivos Excel (.xlsx) en la carpeta de entrada.\n")
                formato_negro = QTextCharFormat()
                formato_negro.setForeground(QColor(0, 0, 0))
                output_textedit.mergeCurrentCharFormat(formato_negro)
                return
                
            else:
                # Crear una lista para almacenar los DataFrames individuales
                lista_dataframes = []

                # Iterar sobre cada archivo Excel y almacenar los DataFrames en la lista
                for archivo in archivos_excel:
                    ruta_archivo = os.path.join(carpeta_datasets, archivo)
                    df = pd.read_excel(ruta_archivo)
                    lista_dataframes.append(df)

                # Concatenar los DataFrames en uno solo
                df_combinado = pd.concat(lista_dataframes, ignore_index=True)
                
                # Guardar el DataFrame combinado en un nuevo archivo Excel
                #archivo_salida = carpeta_datasets + "/dataset_training.xlsx"
                #df_combinado.to_excel(archivo_salida, index=False)
                time.sleep(0.5)
                self.output_textedit.insertPlainText(f"Dataset de entrada fusionado exitosamente.\n")
        except:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('\nCarpeta del datasets para el entrenamiento de modelos erroneo. Asegurate de introducir la ruta correcta a la carpeta para entrenar el modelo.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return

        self.progress += 1
        self.invocar_actualizacion(self.progress)

        #Fase 2: Eliminar columnas que no aportan (tiene muvhos valores nulos o igules en la misma columna) ##############################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Analizando relevacia de cada columna del dataset...\n")
        # Definir umbrales para el porcentaje de valores iguales o nulos
        umbral_porcentaje_iguales = 88
        umbral_porcentaje_nulos = 88

        # Obtener información sobre cada columna
        columnas_a_eliminar = []

        for columna in df.columns:
            # Calcular el porcentaje de valores iguales
            porcentaje_iguales = (df[columna].value_counts().max() / len(df)) * 100
            
            # Calcular el porcentaje de valores nulos
            porcentaje_nulos = (df[columna].isnull().sum() / len(df)) * 100

            self.output_textedit.insertPlainText(f"{columna}: - Porcentaje de valores Iguales: {porcentaje_iguales} - Porcenta de valores Nulos: {porcentaje_nulos}\n")
            time.sleep(0.7)
            
            # Verificar si el porcentaje supera los umbrales y agregar la columna a la lista de eliminación
            if porcentaje_iguales >= umbral_porcentaje_iguales or porcentaje_nulos >= umbral_porcentaje_nulos:
                columnas_a_eliminar.append(columna)

        # Eliminar las columnas seleccionadas del DataFrame
        df = df.drop(columnas_a_eliminar, axis=1)

        # Mostrar las columnas seleccionadas para eliminar
        self.output_textedit.insertPlainText(f" \n")
        self.output_textedit.insertPlainText(f"Columnas seleccionadas para eliminar: {columnas_a_eliminar}\n")
        time.sleep(0.5)
        self.output_textedit.insertPlainText(f"Columnas elimadas correctamente\n")
        time.sleep(0.5)

        # Eliminamos manualmente la columna nombre xq esta duplicado dentro del dataset consecuencia de la fusión de los datasets de SF y MF
        columna_a_eliminar="Nombre"

        # Verifica si la columna existe antes de intentar eliminarla
        if columna_a_eliminar in df.columns:
            # Elimina la columna por su nombre
            df = df.drop(columns=columna_a_eliminar)

            time.sleep(0.5)          
            self.output_textedit.insertPlainText(f"Columna '{columna_a_eliminar}' eliminada con éxito.\n")
        else:
            self.output_textedit.insertPlainText(f"La columna '{columna_a_eliminar}' no existe en el DataFrame y no se pudo eliminar.\n")
        
        self.progress += 1
        self.invocar_actualizacion(self.progress)

        # FASE 3: Gestión de MISSING VALUES ##############################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Manejando Missing Values...\n")

        # Obtener información sobre el tipo de dato de cada columna
        tipos_de_dato = df.dtypes

        # Inicializar los valores faltantes basándote en el tipo de dato
        for columna, tipo in tipos_de_dato.items():
            if pd.api.types.is_numeric_dtype(tipo):
                # Si es numérico, inicializar con 0
                df[columna] = df[columna].fillna(0)
            elif pd.api.types.is_string_dtype(tipo):
                # Si es cadena, inicializar con numpy.nan
                df[columna] = df[columna].fillna(np.nan)

        self.output_textedit.insertPlainText(f"Gestión de Missing Values completada exitosamente.\n")
        
        self.progress += 1
        self.invocar_actualizacion(self.progress)
        time.sleep(0.5)

        # FASE 4: Analisis de distribución de la variable label  ################################################################################################
        # Actualizar los datos y volver a dibujar la gráfica
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Generando gráfica de la distribución de la variable a predecir...\n")

        # Obtener el opción predecir valor o puntos
        index = self.combo_box1.currentIndex()
        self.selected_option = index + 1

        # Realizar acciones en función de la opción seleccionada
        if self.selected_option == 1:
            hilo = threading.Thread(target=visualizar_distribucion(df, 'Valor', 'Valor', 'Distribución puntuaciones jugadores', 'black'))
            hilo.start()
            hilos.append(hilo)

        elif self.selected_option == 2:
            hilo = threading.Thread(target=visualizar_distribucion(df, 'Puntuación Fantasy', 'Puntuación', 'Distribución puntuaciones jugadores', 'white'))
            hilo.start()
            hilos.append(hilo)
            
        time.sleep(0.5)
        self.output_textedit.insertPlainText(f"Histograma generado exitosamente.\n")
        self.progress += 1
        self.invocar_actualizacion(self.progress)

        # FASE 5: Analisis de outliers ############################################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Analizando outliers...\n")
        if self.selected_option == 1:
            hilo = threading.Thread(target=plot_boxplot(df, 'Valor', 'Distribución puntuaciones jugadores','Eje x'))
            hilo.start()
            hilos.append(hilo)

        elif self.selected_option == 2:
            hilo = threading.Thread(target=plot_boxplot(df, 'Puntuación Fantasy', 'Distribución puntuaciones jugadores','Eje x'))
            hilo.start()
            hilos.append(hilo)
        
        time.sleep(0.5)
        self.output_textedit.insertPlainText(f"Box plot generado exitosamente.\n")
        self.progress += 1
        self.invocar_actualizacion(self.progress)

        # FASE 6: Gestión de variables categóricas ###########################################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Gestionando variables categóricas con one-hot encoding...\n")

        # Identificar variables categóricas
        categoricas = df.select_dtypes(include=['object']).columns.tolist()

        # Mostrar las variables categóricas identificadas
        self.output_textedit.insertPlainText(f"Variables categóricas transformadas: {categoricas}\n")

        # Aplicar codificación de frecuencia a las variables categóricas
        encoder = ce.CountEncoder()
        df[categoricas] = encoder.fit_transform(df[categoricas])

        time.sleep(0.5)
        self.output_textedit.insertPlainText(f"Variables categóricas manejadas exitosamente.\n")
        self.progress += 1
        self.invocar_actualizacion(self.progress)
         
        # FASE 7: Matriz de correlación ###########################################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Generando matriz de correlación...\n")

        hilo = threading.Thread(target=plot_correlation_matrix(df,'Matriz de Correlación'))
        hilo.start()
        hilos.append(hilo)

        time.sleep(0.5)
        self.invocar_actualizacion(self.progress)
        self.output_textedit.insertPlainText(f"Matriz de correlación generada exitosamente.\n")
        time.sleep(0.5)
        print(" \n")
        time.sleep(0.5)

        # Finalizar hilos generados para cada plot
        for hilo in hilos:
            hilo.join()
        print("hilos de plots eliminados")

        self.output_textedit.insertPlainText(f"Obteniendo correlación media de cada variable con el resto de varoables...\n")
        # Calcular la matriz de correlación
        correlation_matrix = df.corr()

        # Calcular la media de correlación para cada variable
        mean_correlation = correlation_matrix.abs().mean()

        for indice, valor in mean_correlation.items():
            self.output_textedit.insertPlainText(f"{indice}: media de correlación: {valor}\n")
            time.sleep(0.5)

        # Definir umbral para la media de correlación
        threshold = 0.1

        # Identificar columnas con media de correlación menor al umbral
        low_correlation_columns = mean_correlation[mean_correlation < threshold].index

        self.output_textedit.insertPlainText(f"\nEliminado columnas con baja correlación:\n")
        time.sleep(0.5)
        for i, columna in enumerate(low_correlation_columns):
            self.output_textedit.insertPlainText(f"{columna}")
                
            # Verificar si es el último elemento
            if i < len(low_correlation_columns) - 1:
                self.output_textedit.insertPlainText(", ")
            else:
                self.output_textedit.insertPlainText(".")
                
        time.sleep(0.5)
        self.output_textedit.insertPlainText("\n")

        # Eliminar columnas con baja correlación
        df = df.drop(columns=low_correlation_columns)

        self.progress += 1
        self.invocar_actualizacion(self.progress)

        columnas_array = df.columns.to_numpy()

        # Guardar el nuevo DataFrame en un nuevo archivo Excel
        #df.to_excel('dataset_trasEDA.xlsx', index=False)

        # FASE 8: Entrenar modeleo con el daatset procesado con el algoritmo selecionado ###########################################################################################################################
        selected_model = self.combo_box.currentText()
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')

        if selected_model == "Gradient Boosted Tree model":
            algoritmo_utilizado="Gradient Boosted Tree model"
            if self.selected_option == 1:
                self.output_textedit.insertPlainText(f"Entrenando con Gradient Boosted Trees con el atributo Valor como label.\n")
                
                # Dividir los datos de entrenamiento
                X = df.drop(["Valor"], axis=1)
                y = df["Valor"]  # Variable objetivo ahora es "Valor"                
               
            elif self.selected_option == 2:
                self.output_textedit.insertPlainText(f"Entrenando con Gradient Boosted Trees con el atributo Puntuación Fnatsy como label.\n")
                
                #Crea un nuevo DataFrame X eliminando la columna llamada "Puntuación Fantasy" del DataFrame original df_entrenamiento. El parámetro axis=1 indica que la operación se realiza a lo largo de las columnas.
                X =df.drop(["Puntuación Fantasy"], axis=1)

                #Crea una Serie y que contiene la columna "Puntuación Fantasy" del DataFrame original df_entrenamiento. En otras palabras, se extraen las etiquetas y los valores de la columna "Puntuación Fantasy" y se almacenan en la Serie y.
                y = df["Puntuación Fantasy"] #Variable objetivo
                
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

            # Tiempo de inicio del train
            elapsed_time=0
            start_time = time.time()

            # Inicializar el modelo Gradient Boosted Tree
            modelo = GradientBoostingRegressor(random_state=42)

            # Definimos el número de folds a realizar
            folds = 10

            # Aplicar validación cruzada con múltiples métricas
            scoring = ['neg_mean_squared_error', 'r2', 'neg_mean_absolute_error', 'explained_variance']
            cv_results = cross_validate(modelo, X_train, y_train, cv=folds, scoring=scoring)

            # Mostrar estadísticas para cada fold
            for i in range(folds):
                mse = -cv_results['test_neg_mean_squared_error'][i]
                r2 = cv_results['test_r2'][i]
                mae = -cv_results['test_neg_mean_absolute_error'][i]
                evs = cv_results['test_explained_variance'][i]

                self.output_textedit.insertPlainText(f"\nFold {i + 1}:\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"Mean Squared Error: {mse}\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"R2 Score: {r2}\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"Mean Absolute Error: {mae}\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"Explained Variance Score: {evs}\n")
                time.sleep(0.4)
            
            # Encontrar el índice del fold con el mejor MSE (o cualquier otra métrica de tu elección)
            best_fold_index = np.argmax(cv_results['test_neg_mean_squared_error'])

            # Obtener el mejor modelo entrenado
            best_model = copy.deepcopy(modelo)  # Para evitar afectar al modelo original
            best_model.fit(X_train, y_train)  # Ajustar el mejor modelo con el conjunto de entrenamiento

            # Tiempo de final del train
            end_time = time.time()
            # Calcula la diferencia de tiempo
            elapsed_time = end_time - start_time
            # Convertir a minutos y segundos
            minutes, seconds = divmod(elapsed_time, 60)

            # Mostrar estadísticas agregadas
            average_mse = -cv_results['test_neg_mean_squared_error'].mean()
            average_r2 = cv_results['test_r2'].mean()
            average_mae = -cv_results['test_neg_mean_absolute_error'].mean()
            average_evs = cv_results['test_explained_variance'].mean()

            self.output_textedit.insertPlainText("\nEstadísticas finales obtenidas:\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Mean Squared Error (Promedio): {average_mse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"R2 Score (Promedio): {average_r2}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Mean Absolute Error (Promedio): {average_mae}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Explained Variance Score (Promedio): {average_evs}\n")
            time.sleep(0.4)
            
        elif selected_model == "K-NN model":
            algoritmo_utilizado="KNN model"
            # FASE 8.2. Preparar los datos para entrenar con ellos ######################################################################
            if self.selected_option == 1:
                self.output_textedit.insertPlainText(f"Entrenando con K-NN con el atributo Valor como label.\n")

                X = df.drop('Valor', axis=1)  # Features
                Y = df['Valor']  # Variable de salida

            elif self.selected_option == 2:
                self.output_textedit.insertPlainText(f"Entrenando con K-NN con el atributo Puntuación Fnatsy como label.\n")

                X = df.drop('Puntuación Fantasy', axis=1)  # Features
                Y = df['Puntuación Fantasy']  # Variable de salida
            
            # División en 80/20 para entrenamiento y validación
            X_train, X_val, y_train, y_val = train_test_split(X, Y, test_size=0.2, random_state=42)

            # Normalización de datos
            scaler = MinMaxScaler()
            X_train_scaled = scaler.fit_transform(X_train)
            X_val_scaled = scaler.transform(X_val)
            
            # Definir los valores de k que quieres probar
            k_values = [3, 4, 5, 6, 7, 8, 9, 10]
            n_cv_iterations = 10  # Número de iteraciones en el cross-validation

            # Inicializar variables para almacenar el menor MSE y su información asociada
            min_mse = float('inf')
            best_k = None
            best_iteration = None

            # Inicializar listas para almacenar el MSE medio y el error medio para cada k
            avg_mse_values = []
            avg_error_values = []

            # FASE 8.2.2 Realizar cross validation ######################################################################
            # Tiempo de inicio del train
            elapsed_time=0
            start_time = time.time()

            # Bucle para probar diferentes valores de k
            for k in k_values:
                k_mse = []
                k_error = []
                self.output_textedit.insertPlainText(f'_________________________________________________________\n')
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f'k = {k}:\n')
                time.sleep(0.4)
                
                # Bucle para realizar el cross-validation en cada iteración
                for iteration in range(1, n_cv_iterations + 1):
                    # Inicializar el modelo k-NN
                    knn_model = KNeighborsRegressor(n_neighbors=k)

                    # Configurar el objeto KFold para controlar las iteraciones del cross-validation
                    kf = KFold(n_splits=10, shuffle=True, random_state=iteration)

                    # Realizar validación cruzada y obtener las puntuaciones
                    cv_scores = cross_val_score(knn_model, X_train_scaled, y_train, cv=kf, scoring='neg_mean_squared_error')

                    # Convertir las puntuaciones negativas a positivas para MSE
                    cv_scores = -cv_scores

                    # Calcular el promedio del MSE y el error medio
                    mean_mse = cv_scores.mean()
                    mean_error = np.sqrt(mean_mse)  # Raíz cuadrada del MSE para obtener el error medio
                    
                    # Almacenar el menor MSE y su información asociada
                    if mean_mse < min_mse:
                        min_mse = mean_mse
                        best_k = k
                        best_iteration = iteration

                    # Almacenar el MSE y el error para la iteración actual
                    k_mse.append(mean_mse)
                    k_error.append(mean_error)

                    # Imprimir información para cada valor de k e iteración
                    self.output_textedit.insertPlainText(f'     --------------------------------------\n')
                    time.sleep(0.4)
                    self.output_textedit.insertPlainText(f'     Iteración {iteration} para k = {k}\n')
                    time.sleep(0.4)
                    self.output_textedit.insertPlainText(f'        -Mean CV MSE: {mean_mse}\n')
                    time.sleep(0.4)
                    self.output_textedit.insertPlainText(f'        -Standard Deviation CV MSE: {cv_scores.std()}\n')
                    time.sleep(0.4)
                    self.output_textedit.insertPlainText(f'        -Error medio: {mean_error}\n')
                    time.sleep(0.3)
                
                # Tiempo de final del train
                end_time = time.time()
                # Calcula la diferencia de tiempo
                elapsed_time = end_time - start_time
                # Convertir a minutos y segundos
                minutes, seconds = divmod(elapsed_time, 60)

                # Calcular el MSE medio y el error medio para todas las iteraciones y almacenarlos
                avg_mse_values.append(np.mean(k_mse))
                avg_error_values.append(np.mean(k_error))

                # Imprimir el MSE medio y el error medio final para cada valor de k
                self.output_textedit.insertPlainText(f'____________________________________________________\n')
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f'MSE medio final para k = {k}: {avg_mse_values[-1]}\n')
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f'Error medio final para k = {k}: {avg_error_values[-1]}\n')
                time.sleep(0.4)

            # Imprimir el menor MSE y su información asociada al final
            self.output_textedit.insertPlainText(f'____________________________________________________\n')
            self.output_textedit.insertPlainText(f'\nMenor MSE obtenido: {min_mse} con k = {best_k} en la iteración {best_iteration}\n')

            hilo = threading.Thread(target=plot_lineplot(k_values, avg_mse_values, 'MSE Medio para Cada k', 'Número de vecinos (k)', 'MSE Medio'))
            hilo.start()

        elif selected_model == "Linear Regression model":
            algoritmo_utilizado="Linear Regression"
            if self.selected_option == 1:
                self.output_textedit.insertPlainText(f"Entrenando con regresión lineal con el atributo Valor como label.\n")
                X = df.drop('Valor', axis=1)  # Features
                Y = df['Valor']  # Variable de salida

            elif self.selected_option == 2:
                self.output_textedit.insertPlainText(f"Entrenando con regresión lineal con el atributo Puntuación Fantasy como label.\n")
                X = df.drop('Puntuación Fantasy', axis=1)  # Features
                Y = df['Puntuación Fantasy']  # Variable de salida
        
            # Dividimos los datos en un 80% para entrenamiento y un 20% para prueba
            X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.2, random_state=42)

            # Tiempo de inicio del train
            elapsed_time=0
            start_time = time.time()

            # Inicializar el modelo de regresión lineal
            modelo = LinearRegression()

            # Definimos el número de folds a realizar
            folds=5

            # Aplicar validación cruzada con múltiples métricas
            scoring = ['neg_mean_squared_error', 'r2', 'neg_mean_absolute_error', 'explained_variance']
            cv_results = cross_validate(modelo, X_train, y_train, cv=folds, scoring=scoring)

            # Mostrar estadísticas para cada fold
            for i in range(folds):  
                mse = -cv_results['test_neg_mean_squared_error'][i]  
                r2 = cv_results['test_r2'][i]
                mae = -cv_results['test_neg_mean_absolute_error'][i]  
                evs = cv_results['test_explained_variance'][i]

                self.output_textedit.insertPlainText(f"\nFold {i + 1}:\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"Mean Squared Error: {mse}\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"R2 Score: {r2}\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"Mean Absolute Error: {mae}\n")
                time.sleep(0.4)
                self.output_textedit.insertPlainText(f"Explained Variance Score: {evs}\n")
                time.sleep(0.3)

            # Encontrar el índice del fold con el mejor MSE (o cualquier otra métrica de tu elección)
            best_fold_index = np.argmax(cv_results['test_neg_mean_squared_error'])

            # Obtener el mejor modelo entrenado
            best_model = copy.deepcopy(modelo)  # Para evitar afectar al modelo original
            best_model.fit(X_train, y_train)  # Ajustar el mejor modelo con el conjunto de entrenamiento
            
            # Tiempo de final del train
            end_time = time.time()
            # Calcula la diferencia de tiempo
            elapsed_time = end_time - start_time
            # Convertir a minutos y segundos
            minutes, seconds = divmod(elapsed_time, 60)

            # Mostrar estadísticas agregadas
            average_mse = -cv_results['test_neg_mean_squared_error'].mean()
            average_r2 = cv_results['test_r2'].mean()
            average_mae = -cv_results['test_neg_mean_absolute_error'].mean() 
            average_evs = cv_results['test_explained_variance'].mean()

            self.output_textedit.insertPlainText("\nEstadísticas finales obtenidas:\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Mean Squared Error (Promedio): {average_mse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"R2 Score (Promedio): {average_r2}")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Mean Absolute Error (Promedio): {average_mae}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Variance Score (Promedio): {average_evs}\n")
            time.sleep(0.4)

        self.progress += 1
        self.invocar_actualizacion(self.progress)

        # FASE 8.2.3 Validar modelo generado en el entrenamiento ################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Testeando modelo generado en el entrenamiento...\n")

        if selected_model == "Gradient Boosted Tree model":
            # Inicializar el modelo Gradient Boosted Tree
            modelo = GradientBoostingRegressor(random_state=42)

            # Ajustar (fit) el modelo con el conjunto de entrenamiento
            modelo.fit(X_train, y_train)

            # Realizar predicciones en el conjunto de prueba
            y_pred_test = modelo.predict(X_test)

            # Calcular métricas de evaluación en el conjunto de prueba
            test_mse = mean_squared_error(y_test, y_pred_test)
            test_rmse = np.sqrt(test_mse)
            test_mae = mean_absolute_error(y_test, y_pred_test)
            test_r2 = r2_score(y_test, y_pred_test)

            # Imprimir las métricas en el conjunto de prueba
            self.output_textedit.insertPlainText("\nMétricas en el conjunto de prueba:\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"      -Mean Squared Error (MSE) en conjunto de test: {test_mse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"      -Root Mean Squared Error (RMSE) en conjunto de test: {test_rmse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"      -Mean Absolute Error (MAE) en conjunto de test: {test_mae}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"      -R^2 Score en conjunto de test: {test_r2}\n")
            time.sleep(0.4)

        elif selected_model == "Linear Regression model":
           # Inicializar el modelo de regresión lineal
            modelo = LinearRegression()

            # Ajustar el modelo con el conjunto de entrenamiento
            modelo.fit(X_test, y_test)

            # Realizar predicciones en el conjunto de test
            y_test_pred = modelo.predict(X_test)

            # Calcular varias métricas en el conjunto de test
            test_mse = mean_squared_error(y_test, y_test_pred)
            test_rmse = np.sqrt(test_mse)  # Raíz cuadrada del MSE para obtener el RMSE
            test_mae = mean_absolute_error(y_test, y_test_pred)
            test_r2 = r2_score(y_test, y_test_pred)

        # Imprimir las métricas
            self.output_textedit.insertPlainText("\nMétricas en el conjunto de prueba:\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"     -Mean Squared Error (MSE) en conjunto de test: {test_mse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"     -Root Mean Squared Error (RMSE) en conjunto de test: {test_rmse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"     -Mean Absolute Error (MAE) en conjunto de test: {test_mae}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"     -R^2 Score en conjunto de test: {test_r2}\n")
            time.sleep(0.4)
            
        elif selected_model == "K-NN model":
            # Ajustar el mejor modelo con el mejor k en el conjunto de entrenamiento
            best_model = KNeighborsRegressor(n_neighbors=best_k)
            best_model.fit(X_train_scaled, y_train)

            # Realizar predicciones en el conjunto de test
            y_val_pred = best_model.predict(X_val_scaled)

            # Calcular varias métricas en el conjunto de test
            val_mse = mean_squared_error(y_val, y_val_pred)
            val_rmse = np.sqrt(val_mse)  # Raíz cuadrada del MSE para obtener el RMSE
            val_mae = mean_absolute_error(y_val, y_val_pred)
            val_r2 = r2_score(y_val, y_val_pred)

            # Imprimir varias métricas en el conjunto de test
            self.output_textedit.insertPlainText("\nMétricas en el conjunto de prueba:\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f'     -Mean Squared Error (MSE) obtenido en el test del modelo con k = {best_k}:   {val_mse}\n')
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f'     -Root Mean Squared Error (RMSE) obtenido en el test del modelo con k = {best_k}:   {val_rmse}\n')
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f'     -Mean Absolute Error (MAE) obtenido en el test del modelo con k = {best_k}:   {val_mae}\n')
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f'     -R^2 Score obtenido en el test del modelo con k = {best_k}:   {val_r2}\n')
            time.sleep(0.4)
            
        self.progress += 1
        self.invocar_actualizacion(self.progress)

        # FASE 8.2.5 Resumen del entrenamiento ################################################################################################
        # Obtener la fecha actual
        fecha_actual = datetime.now()
        # Formatear la fecha como una cadena (opcional)
        fecha_actual_str = fecha_actual.strftime("%Y-%m-%d--%H-%M-S")

        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Resumen del entrenamiento:\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Fecha de entrenamiento:  {fecha_actual_str}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Número de ejemplares utilizados:  {len(archivos_excel)}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Algoritmo utilizado:  {algoritmo_utilizado}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Tiempo de entrenamiento:  {int(minutes)} minutos y {seconds:.2f} segundos\n")
        time.sleep(0.3)

        if selected_model == "Gradient Boosted Tree model" or selected_model == "Linear Regression model":

            self.output_textedit.insertPlainText(f"Mejor fold obtenido': {best_fold_index}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Mean Squared Error en el entrenamiento: {average_mse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"R2 Score en el entrenamiento: {average_r2}")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Mean Absolute Error en el entrenamiento: {average_mae}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Variance Score en el entrenamiento: {average_evs}\n")
            time.sleep(0.4)

            self.output_textedit.insertPlainText(f"Mean Squared Error (MSE) en conjunto de test: {test_mse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Root Mean Squared Error (RMSE) en conjunto de test: {test_rmse}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"Mean Absolute Error (MAE) en conjunto de test: {test_mae}\n")
            time.sleep(0.4)
            self.output_textedit.insertPlainText(f"R2 Score en conjunto de test: {test_r2}\n")
            time.sleep(0.4)

            # Guardar la información del entrenamiento, incluyendo las columnas eliminadas
            entrenamiento_info = {
                'Fecha de entrenamiento': fecha_actual_str,
                'Número de ejemplares utilizados': len(archivos_excel),
                'Algoritmo utilizado': algoritmo_utilizado,
                'Tiempo de entrenamiento (minutos)': minutes,
                'Tiempo de entrenamiento (segundos)': seconds,
                'Mejor fold obtenido': best_fold_index,
                'MSE obtenido en el entrenamiento': average_mse,
                'Variance Score obtenido en el entrenamiento': average_evs,
                'MAE obtenido en el entrenamiento': average_mae,
                'R^2 obtenido en el entrenamiento': average_r2,
                'MSE obtenido en el test del modelo': test_mse, 
                'RMSE obtenido en el test del modelo': test_rmse, 
                'MAE obtenido en el test del modelo': test_mae, 
                'R^2 obtenido en el test del modelo': test_r2, 
                'Columnas con las que ha sido entrenado': columnas_array.tolist()  
            }

        elif selected_model == "K-NN model":

            self.output_textedit.insertPlainText(f"Mejor K obtenida:  {best_k}\n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"MSE obtenido en el entrenamiento:  {min_mse}\n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"MSE obtenido en el test del modelo:  {val_mse}\n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"RMSE obtenido en el test del modelo:  {val_rmse}\n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"MAE obtenido en el test del modelo:  {val_mae}\n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"R^2 obtenido en el test del modelo:  {val_r2}\n")
            time.sleep(0.3)

            # Guardar la información del entrenamiento, incluyendo las columnas eliminadas
            entrenamiento_info = {
                'Fecha de entrenamiento': fecha_actual_str,
                'Número de ejemplares utilizados': len(archivos_excel),
                'Algoritmo utilizado': algoritmo_utilizado,
                'Tiempo de entrenamiento (minutos)': minutes,
                'Tiempo de entrenamiento (segundos)': seconds,
                'Mejor K obtenida': best_k,
                'MSE obtenido en el entrenamiento': min_mse,
                'MSE obtenido en el test del modelo': val_mse, 
                'RMSE obtenido en el test del modelo': val_rmse, 
                'MAE obtenido en el test del modelo': val_mae, 
                'R^2 obtenido en el test del modelo': val_r2, 
                'Columnas con las que ha sido entrenado': columnas_array.tolist()  
            }

        # FASE 8.2.4 Guardar modelo e información asociada de la generación del modelo en el entrenanienta ################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Guardando modelo generado e información asociada del entrenamiento y test del modelo...\n")
        time.sleep(1)

        if self.selected_option == 1:
            model_name= algoritmo_utilizado+"_Valor_Mercado_to_predict_"+fecha_actual_str+".pkl" 
        elif self.selected_option == 2:
            model_name= algoritmo_utilizado+"_Puntuación_Fantasy_to_predict_"+fecha_actual_str+".pkl"

        carpeta_save=self.text_input2.text()

        ruta_save=carpeta_save+"/"+model_name
        
        # Crear un diccionario que contiene el modelo y la información del entrenamiento
        combined_data = {
            'model': best_model,
            'training_info': entrenamiento_info
        }

        # Guardar el diccionario en un archivo
        joblib.dump(combined_data, ruta_save)

        self.output_textedit.insertPlainText(f"Modelo guardando correctamente en  {ruta_save}\n")
        self.progress += 1
        self.invocar_actualizacion(self.progress)
        time.sleep(2)
       
class predictWindow(QWidget):
    def __init__(self):
        super().__init__()

        # Crear un diseño de cuadrícula dentro del QVBoxLayout
        grid_layout = QGridLayout(self)

        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("PREDECIR DATOS")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        grid_layout.addWidget(label_text, 0, 0)

        # LABEL SUBTÍTULO 1
        label_subtext1 = QLabel("Predice el valor de mercado o los puntos que obtendrá el jugador en la sigueinte jornada de la liga mediante el modelo generado en el entrenamiento.")
        grid_layout.addWidget(label_subtext1, 1, 0, 1, 2)

        ### SELECCIONAR RUTA DATASET DE ENTRADA ##################################################
        # LABEL DE TEXTO
        label_text = QLabel("Ruta de los futbolitas a predecir: ")
        grid_layout.addWidget(label_text, 2, 0)

        # INPUT DE TEXTO
        self.text_file_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_file_input, 3, 0, 1, 2)

        # BOTÓN PARA SELECCIONAR ARCHIVO
        select_file_button = QPushButton("Seleccionar Archivo")
        select_file_button.clicked.connect(lambda: select_file(self))

        # Alineación
        grid_layout.addWidget(select_file_button, 4, 1, alignment=Qt.AlignmentFlag.AlignRight)

        # Estilos
        select_file_button.setMinimumWidth(140)

        ### SELECCIONAR RUTA MODELO A USAR #################################################################
        # LABEL DE TEXTO
        label_text = QLabel("Ruta del modelo a utilzar para predecir: ")
        grid_layout.addWidget(label_text, 5, 0)

        # INPUT DE TEXTO
        self.text_file2_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_file2_input, 6, 0, 1, 2)

        # BOTÓN PARA SELECCIONAR ARCHIVO
        select_file_button = QPushButton("Seleccionar modelo")
        select_file_button.clicked.connect(lambda: select_file2(self))
        # Alineación
        grid_layout.addWidget(select_file_button, 7, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_file_button.setMinimumWidth(140)

        ###  SELECCIONAR RUTA DONDE GUARDAR EL EXCEL OUTPUT DEL SCRAPER  ###################################
        # LABEL TEXTO 
        label_text = QLabel("Ruta donde guardar predicciones realizadas: ")
        grid_layout.addWidget(label_text, 8, 0)

        # INPUT TEXTO (QLineEdit en lugar de QSpinBox)
        self.text_input = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input, 9, 0, 1, 2)
        # Estilos 
        #self.text_input.setMinimumWidth(350)

        # BOTÓN PARA SELECCIONAR CARPETA
        select_folder_button = QPushButton("Seleccionar Carpeta")
        select_folder_button.clicked.connect(lambda: select_folder(self))
        # Alineación
        grid_layout.addWidget(select_folder_button, 10, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        select_folder_button.setMinimumWidth(140)

        ### BOTÓN PARA EMPEZAR ENTRENAMIENTO ###########################################################
        # Crear un botón
        self.scrape_button = QPushButton("Realizar prediciones")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.scrape_button.clicked.connect(self.iniciar_scrapear_thread)

        # Alineación y estilos
        grid_layout.addWidget(self.scrape_button, 11, 0)
        self.scrape_button.setMaximumWidth(150)

        ###  VENTANA OUTPUT predecir  ####################################################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        grid_layout.addWidget(self.output_textedit, 12, 0, 10, 0)  # row, column, rowSpan, columnSpan
    
    def iniciar_scrapear_thread(self):  
        # Crear un hilo y ejecutar la función en segundo plano
        thread = threading.Thread(target=self.train_function)
        thread.start()

    def train_function(self):
        #### PARTE 0 : LEER INPUTS + COMPROBAR QUE TODAS LOS INPUTS (rutas de archivos y carpetas) HAN SIDO INICIALIZADAS #####################################################################
        # Ruta a la carpeta que contiene los archivos json de Sofaescore
        file_futbolistas = self.text_file_input.text()
        file_modelo = self.text_file2_input.text()
        carpeta_save=self.text_input.text()

        if not file_futbolistas or not file_modelo or not carpeta_save:
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            self.output_textedit.mergeCurrentCharFormat(formato_rojo)
            if not file_futbolistas:
                self.output_textedit.insertPlainText("La ruta de la carpeta del dataset de entrenamiento no ha sido inicializada.\n")

            if not file_modelo:
                self.output_textedit.insertPlainText("La ruta de la carpeta del dataset de entrenamiento no ha sido inicializada.\n")
            
            if not carpeta_save:
                self.output_textedit.insertPlainText("La ruta de la carpeta donde guardar el modelo generado no ha sido inicializada.\n")
            
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            self.output_textedit.mergeCurrentCharFormat(formato_negro)
            return
        
        #### PARTE 1 : CARGAR MODELO E INFORMACIÓN ASOCIADA del entrenamiento #########################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Cargando modelo e información estadística asociada del entrenamiento del mismo...\n\n")
       
        try:
            # Cargar el diccionario desde el archivo
            loaded_combined_data = joblib.load(file_modelo)
        except:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('Modelo cargado erroneo. Asegurate de introducir la ruta correcta del modleo con el que predecir los valores deseados.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return

        # Acceder al modelo y la información del entrenamiento desde el diccionario
        loaded_model = loaded_combined_data['model']
        loaded_entrenamiento_info = loaded_combined_data['training_info']

        # Imprimir todos los valores de la información del entrenamiento
        self.output_textedit.insertPlainText(f"Fecha de entrenamiento: {loaded_entrenamiento_info['Fecha de entrenamiento']} \n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText(f"Número de ejemplares utilizados: {loaded_entrenamiento_info['Número de ejemplares utilizados']} \n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText(f"Algoritmo utilizado: {loaded_entrenamiento_info['Algoritmo utilizado']} \n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText(f"Tiempo de entrenamiento: {loaded_entrenamiento_info['Tiempo de entrenamiento (minutos)']} minutos. \n")
        time.sleep(0.2)
        self.output_textedit.insertPlainText(f"Tiempo de entrenamiento: {loaded_entrenamiento_info['Tiempo de entrenamiento (segundos)']} seconds. \n")
        time.sleep(0.2)

        if loaded_entrenamiento_info['Algoritmo utilizado'] == "Linear Regression" or loaded_entrenamiento_info['Algoritmo utilizado'] == "Gradient Boosted Tree model":
            #self.output_textedit.insertPlainText(f"Mejor fold obtenido:  {loaded_entrenamiento_info['Mejor fold obtenido']} \n")
            time.sleep(0.2)                                                                                       
            self.output_textedit.insertPlainText(f"Mean Squared Error en el entrenamiento:  {loaded_entrenamiento_info['MSE obtenido en el entrenamiento']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"R2 Score en el entrenamiento:  {loaded_entrenamiento_info['Variance Score obtenido en el entrenamiento']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"Mean Absolute Error en el entrenamiento:  {loaded_entrenamiento_info['MAE obtenido en el entrenamiento']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"Variance Score en el entrenamiento:  {loaded_entrenamiento_info['R^2 obtenido en el entrenamiento']} \n")      
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"Mean Squared Error (MSE) en conjunto de test:  {loaded_entrenamiento_info['MSE obtenido en el test del modelo']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"Root Mean Squared Error (RMSE) en conjunto de test:  {loaded_entrenamiento_info['RMSE obtenido en el test del modelo']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"Mean Absolute Error (MAE) en conjunto de test:  {loaded_entrenamiento_info['MAE obtenido en el test del modelo']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"R2 Score en conjunto de test:  {loaded_entrenamiento_info['R^2 obtenido en el test del modelo']} \n")      
            time.sleep(0.3)
          
        elif loaded_entrenamiento_info['Algoritmo utilizado'] == "K-NN model":
            self.output_textedit.insertPlainText(f"Mejor K obtenida: {loaded_entrenamiento_info['Mejor K obtenida']} \n")
            time.sleep(0.2)
            self.output_textedit.insertPlainText(f"MSE obtenido en el entrenamiento: {loaded_entrenamiento_info['MSE obtenido en el entrenamiento']} \n")
            time.sleep(0.2)

            self.output_textedit.insertPlainText(f"MSE obtenido en el test del modelo:  {loaded_entrenamiento_info['MSE obtenido en el test del modelo']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"RMSE obtenido en el test del modelo:  {loaded_entrenamiento_info['RMSE obtenido en el test del modelo']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"MAE obtenido en el test del modelo:  {loaded_entrenamiento_info['MAE obtenido en el test del modelo']} \n")
            time.sleep(0.3)
            self.output_textedit.insertPlainText(f"R^2 obtenido en el test del modelo:  {loaded_entrenamiento_info['R^2 obtenido en el test del modelo']} \n")      
            time.sleep(0.3)

        columnas_a_mantener = loaded_entrenamiento_info['Columnas con las que ha sido entrenado']
        self.output_textedit.insertPlainText(f"Columnas con las que ha sido entrenado: {columnas_a_mantener} \n")
        time.sleep(0.2)
        
        #### PARTE 2 : CARGAR DATASET DE JUGADORES A PREDECIR #########################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Cargando fichero de jugadores a predecir...\n")
        try:
            # Cargar el DataFrame de jugadores
            df = pd.read_excel(file_futbolistas)
        except:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText('\Fichero de jugaodres a predecir de mi plantilla o mercado erroneo. Asegurate de introducir la ruta correcta del archico para poder predecir los valores de los mismos.\n')
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            return

        time.sleep(0.5)
        self.output_textedit.insertPlainText(f"Fichero cargado exitosamente...\n\n")
        time.sleep(0.5)

        #### PARTE 3 : ELIMINAR COLUMNAS eliminadas en el entrenamiento  #########################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Eliminando columnas que no fueron usadas para entrenar el modelo...\n")

        # Eliminar las columnas que no están en 'columnas_a_mantener'
        columnas_a_eliminar = [col for col in df.columns if col not in columnas_a_mantener]
        df = df.drop(columns=columnas_a_eliminar)

        self.output_textedit.insertPlainText(f"Columnas que no fueron usadas para entrenar el modelo eliminadas correctamente.\n\n")

        #### PARTE 4 : CODIFICAR VARIABLES CATEGÓRICAS ############################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Codificando variables categóricas.\n\n")

        # Identificar variables categóricas
        categoricas = df.select_dtypes(include=['object']).columns.tolist()

        # Mostrar las variables categóricas identificadas
        self.output_textedit.insertPlainText(f"Variables categóricas transformadas: {categoricas}\n")

        # Aplicar codificación de frecuencia a las variables categóricas
        encoder = ce.CountEncoder()
        df[categoricas] = encoder.fit_transform(df[categoricas])

        time.sleep(0.5)
        self.output_textedit.insertPlainText(f"Variables categóricas manejadas exitosamente.\n")

        #### PARTE 5 : MANEJAR MISSING VALUES ######################################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText(f"Manejando Missing Values...\n")

        # Obtener información sobre el tipo de dato de cada columna
        tipos_de_dato = df.dtypes

        # Inicializar los valores faltantes basándote en el tipo de dato
        for columna, tipo in tipos_de_dato.items():
            if pd.api.types.is_numeric_dtype(tipo):
                # Si es numérico, inicializar con 0
                df[columna] = df[columna].fillna(0)
            elif pd.api.types.is_string_dtype(tipo):
                # Si es cadena, inicializar con numpy.nan
                df[columna] = df[columna].fillna(np.nan)

        self.output_textedit.insertPlainText(f"Gestión de Missing Values completada exitosamente.\n")
        
        #### PARTE 5 : EJECUTAR MODELO ######################################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        if '_Puntuación_Fantasy_to_predict' in file_modelo:
            self.output_textedit.insertPlainText("Prediciendo valores de mercado de los jugadores seleccionados para la próxima jornada de liga.\n")
        
            # Eliminar la columna a predecir del DataFrame df
            columna_a_eliminar = 'Puntuación Fantasy'
            df = df.drop(columna_a_eliminar, axis=1)
            print("Check elimada col puntos MF")
            
        elif '_Valor_Mercado_to_predict' in file_modelo:
            self.output_textedit.insertPlainText("Prediciendo puntuaciones de Mister Fantasy MD de los jugadores seleccionados para la próxima jornada de liga.\n")
            
            # Eliminar la columna a predecir del DataFrame df
            columna_a_eliminar = 'Valor'
            df = df.drop(columna_a_eliminar, axis=1)
            print("Check elimada col valor")

        else:
            self.output_textedit.insertPlainText("Modelo cargado incompatible.\n")
            return

        # Registra el tiempo de inicio
        inicio_tiempo = time.time()

        # Realizar predicciones
        predicciones = loaded_model.predict(df)

        # Crear un DataFrame con las predicciones
        df_predicciones = pd.DataFrame({'Predicciones': predicciones})

        # Registra el tiempo de finalización
        fin_tiempo = time.time()
        # Calcula la diferencia de tiempo para obtener el tiempo total
        tiempo_total = fin_tiempo - inicio_tiempo
        # Convertir a minutos y segundos
        minutes, seconds = divmod(tiempo_total, 60)

        #### PARTE 6 : OBTENER ESTADISTICAS DE LAS PREDIciONES REALIZADAS ######################################################################################################################

        df = pd.read_excel(file_futbolistas)
       
        # Llena los valores nulos con ceros en las columnas 'Valor' y 'Puntuación Fantasy'
        df['Valor'].fillna(0, inplace=True)
        df['Puntuación Fantasy'].fillna(0, inplace=True)

        # Inicializa y_true con ceros
        y_true = np.zeros(len(df))

        if '_Valor_Mercado_to_predict' in file_modelo:
            y_true = df['Valor'].values
        elif '_Puntuación_Fantasy_to_predict' in file_modelo:
            y_true = df['Puntuación Fantasy'].values
        
        # Calcula el error cuadrático medio (MSE)
        mse = mean_squared_error(y_true, predicciones)
        
        # Calculo mean ansolute error(MAE)
        mae = mean_absolute_error(y_true, predicciones)

        # Calculo de R-cuadrado
        r2 = r2_score(y_true, predicciones)

        #### PARTE 7 : MOSTRAR MEJOR JUGADOR POR POSICIÓN ######################################################################################################################
        # Cargar el DataFrame de jugadores
        dfp = pd.read_excel(file_futbolistas)

        #Imprimir nombre + posicion + predicción
        for index, (nombre, pos, row) in enumerate(zip(dfp['Jugador'], dfp['Posición'], df_predicciones.iterrows())):
            self.output_textedit.insertPlainText(f"Jugador: {nombre}, Posición: {pos}, Predicción: {row[1]['Predicciones']}\n")
            time.sleep(0.2)

        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        
        if '_Valor_Mercado_to_predict' in file_modelo:
            self.output_textedit.insertPlainText("Mejores jugadores para cada posicón por puntos obtenidos:\n")
        elif '_Puntuación_Fantasy_to_predict' in file_modelo:
            self.output_textedit.insertPlainText("Mejores jugadores para cada posicón por puntos obtenidos:\n")

        # Crear un diccionario para almacenar los jugadores agrupados y ordenados por posición y predicción
        grupos_por_posicion = {}

        # Iterar sobre los datos
        for index, (nombre, pos, row) in enumerate(zip(dfp['Jugador'], dfp['Posición'], df_predicciones.iterrows())):
            # Crear un subgrupo para la posición si aún no existe
            if pos not in grupos_por_posicion:
                grupos_por_posicion[pos] = []

            # Agregar el jugador al subgrupo correspondiente
            grupos_por_posicion[pos].append({
                'jugador': nombre,
                'posicion': pos,
                'prediccion': row[1]['Predicciones']
            })

        # Ordenar cada subgrupo por posición y predicción de mayor a menor
        for posicion, jugadores in grupos_por_posicion.items():
            grupos_por_posicion[posicion] = sorted(jugadores, key=lambda x: x['prediccion'], reverse=True)

        # Imprimir cada subgrupo ordenado
        for posicion, jugadores in grupos_por_posicion.items():
            self.output_textedit.insertPlainText(f"\nPosición {posicion}:\n")
            index=1
            for jugador_info in jugadores:

                if '_Valor_Mercado_to_predict' in file_modelo:
                    self.output_textedit.insertPlainText(f"{index}- {jugador_info['jugador']} :  Predicción de valor de mercado: {jugador_info['prediccion']}\n")
                elif '_Puntuación_Fantasy_to_predict' in file_modelo:
                    self.output_textedit.insertPlainText(f"{index}- {jugador_info['jugador']} :  Predicción de puntuación: {jugador_info['prediccion']}\n")
                
                time.sleep(0.2)
                index+=1
            self.output_textedit.insertPlainText(f" \n")

        #### PARTE 8 : GEERAR ESTADISTICAS DE LA EJCUCIÓN DEL MODLEO para realizar la sprediciones ######################################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText("Estadísticas de la ejecución del modelo:\n")

        # Obtener la fecha actual
        fecha_actual = datetime.now()
        # Formatear la fecha como una cadena (opcional)
        fecha_actual_str = fecha_actual.strftime("%Y-%m-%d--%H-%M-S")

        algoritmo_utilizado='undefined'
        if "KNN" in file_modelo:
            algoritmo_utilizado = 'KNN'
        elif "Linear Regression" in file_modelo:
            algoritmo_utilizado = 'Linear_Regresion'
        elif "Gradient Boosted Tree" in file_modelo:
            algoritmo_utilizado = 'Gradient_Boosted_Tree'

        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Fecha de ejecución:  {fecha_actual_str}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Total de jugaodres predecidos:  {len(df_predicciones)}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Algoritmo utilizado:  {algoritmo_utilizado}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"Tiempo de ejecución:  {int(minutes)} minutos y {seconds:.2f} segundos\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"MSE (Error cuadrático medio) obtenido en las prediciones:  {mse}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"MAE (Error absoluo medio) obtenido en lss prediciones:  {mae}\n")
        time.sleep(0.3)
        self.output_textedit.insertPlainText(f"R2 (Coeficiente de determinación) obtenido en las prediciones:  {r2}\n")
        time.sleep(0.3)
        
        #### PARTE 8 : GUARDAR DATOS PREDECIDOS DE CADA JUGADOR ######################################################################################################################
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText("Guardando datos predecidos...\n")

        for posicion, jugadores in grupos_por_posicion.items():
            grupos_por_posicion[posicion] = sorted(jugadores, key=lambda x: x['prediccion'], reverse=True)

        # Concatenar todos los subgrupos en un solo DataFrame
        df_total = pd.concat([pd.DataFrame(jugadores) for jugadores in grupos_por_posicion.values()])

        # Guardar todos los jugadores en un solo archivo Excel
        #Obtener la fecha actual
        fecha_actual = datetime.now()
        fecha_actual_str = fecha_actual.strftime("%Y-%m-%d--%H-%M-S")
        
        if '_Valor_Mercado_to_predict' in file_modelo:
            excel_file_path = carpeta_save+"/predicciones_Valores_Mercado_jugadores_con_"+algoritmo_utilizado+"_" + fecha_actual_str + ".xlsx"
        elif '_Puntuación_Fantasy_to_predict' in file_modelo:
            excel_file_path = carpeta_save+"/predicciones_Puntos_jugadores_con_"+algoritmo_utilizado+"_" + fecha_actual_str + ".xlsx"

        df_total.to_excel(excel_file_path, index=False)
              
        self.output_textedit.insertPlainText(f"Datos predecidos guardados correctamente en {excel_file_path}.\n")


class login(QWidget):   
    def __init__(self): 
        super().__init__()

        # Crear un diseño de cuadrícula dentro del QVBoxLayout
        grid_layout = QGridLayout(self)

        # TITULO VENTANA  ###########################################################################################
        # LABEL TÍTULO
        label_text = QLabel("MI PERFIL")
        # Aplicar estilos para destacar el texto
        label_text.setStyleSheet("font-weight: bold; color: black; font-size: 20px;")
        grid_layout.addWidget(label_text, 0, 0)

        # LABEL SUBTÍTULO 1
        label_subtext1 = QLabel("Danos acceso a tu cuenta de Mister Fantasy MD logueandote en el siguiente formulario para permitir a la aplicación obtener informacion de los jugadores de la liga. ")
        grid_layout.addWidget(label_subtext1, 1, 0, 1, 2)

        # LABEL SUBTÍTULO 2
        label_subtext2 = QLabel("* Tus credenciales nunca serán guardadas y se eliminaran autoamticamente al cerrar la aplicación. *")
        # Aplicar estilos para destacar el texto
        label_subtext2.setStyleSheet("color: red;")
        grid_layout.addWidget(label_subtext2, 2, 0, 1, 2)

        # INPUT CREDENCIALES  #########################################################################################
        # LABEL DE TEXTO
        label_text2 = QLabel("Usuario: ")
        grid_layout.addWidget(label_text2, 3, 0, alignment=Qt.AlignmentFlag.AlignTop)

        # INPUT DE TEXTO
        self.text_input1 = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input1, 3, 1)

        ### SELECCIONAR PSW ##################################################
        # LABEL DE TEXTO
        label_text = QLabel("Contraseña: ")
        grid_layout.addWidget(label_text, 4, 0)

        # INPUT DE TEXTO
        self.text_input2 = QLineEdit(self)
        # Alineación
        grid_layout.addWidget(self.text_input2, 4, 1)

        ### BOTÓN PARA EJECUTAR FUNCIÓN PARA FUSIONAR EXCELLS ###########################################################
        # Crear un botón
        self.save_button = QPushButton("Loguearte en Mister Fantasy MD")

        # Conectar la señal clicked del botón a la función iniciar_scrapear_thread e iniciar la barra de progreso
        self.save_button.clicked.connect(self.iniciar_scrapear_thread)

        # Alineación
        grid_layout.addWidget(self.save_button, 5, 1, alignment=Qt.AlignmentFlag.AlignRight)
        # Estilos
        self.save_button.setMinimumWidth(100)
        #self.save_button.setMaximumWidth(150)

        ###  VENTANA OUTPUT SCRAPER  ####################################################################################
        # Crear un QTextEdit para la salida
        self.output_textedit = QTextEdit(self)
        grid_layout.addWidget(self.output_textedit, 6, 0, 11, 0)  # row, column, rowSpan, columnSpan

    def iniciar_scrapear_thread(self):  
        # Crear un hilo y ejecutar la función en segundo plano
        thread = threading.Thread(target=self.guardar_credenciales)
        thread.start()

    def guardar_credenciales(self):
        self.output_textedit.insertPlainText('________________________________________________________________________________________\n')
        self.output_textedit.insertPlainText('Comprobando credenciales introducidas...\n')

        usuario_input = self.text_input1.text()
        contrasena_input = self.text_input2.text()

        if usuario_input!="" and contrasena_input!="":
            
            self.driver = webdriver.Chrome()

            # Navega a la página web que deseas hacer scraping
            self.driver.get("https://mister.mundodeportivo.com/new-onboarding/#market")

            # Espera a que se cargue la página
            self.driver.implicitly_wait(15)

            # Encuentra el botón de "Consentir" 
            button = self.driver.find_element(By.XPATH, '//*[@id="didomi-notice-agree-button"]')
            # Haz clic en el botón de "Consentir" 
            button.click()

            # Encuentra el botón de "Siguinete" 
            button = self.driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[2]/button')
            # Haz clic en el botón de "Siguiente" 
            button.click()
            button.click()
            button.click()
            button.click()

            # Encuentra el botón de "sing con gmail" 
            button = self.driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/button[3]')
            button.click()

            # Localiza el elemento del input gmail
            inputgmail = self.driver.find_element(By.XPATH, '//*[@id="email"]')

            # Borra cualquier contenido existente en la caja de texto (opcional)
            inputgmail.clear()

            # Ingresa texto en la caja de texto
            inputgmail.send_keys(usuario_input)

            # Localiza el elemento del input gmail
            inputpsw = self.driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/form/div[2]/input')

            # Borra cualquier contenido existente en la caja de texto (opcional)
            inputpsw.clear()

            # Ingresa texto en la caja de texto
            inputpsw.send_keys(contrasena_input)

            # Encuentra el botón de "sing con gmail" 
            button = self.driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/form/div[3]/button')
            button.click()
            try:
                # Encuentra el botón para comprobar si se ha conseguido loguear en la web de Mister Fantasy
                button = self.driver.find_element(By.XPATH, '//*[@id="content"]/header/div[2]/ul/li[2]/a')
                
                # Haz clic en el botón
                button.click()

                self.driver.quit()

                self.output_textedit.insertPlainText('Credenciales correctas.\n')

                # Acceder a las variables globales desde la clase
                global usuario, contrasena
                usuario = usuario_input
                contrasena = contrasena_input

                return
            
            except NoSuchElementException:
                self.driver.quit()
                output_textedit = self.output_textedit
                color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
                formato_rojo = QTextCharFormat()
                formato_rojo.setForeground(color_rojo)
                output_textedit.mergeCurrentCharFormat(formato_rojo)
                output_textedit.insertPlainText('Usuario o contraseña incorrectos.\n')
                formato_negro = QTextCharFormat()
                formato_negro.setForeground(QColor(0, 0, 0))
                output_textedit.mergeCurrentCharFormat(formato_negro)
                return
        else:
            output_textedit = self.output_textedit
            color_rojo = QColor(255, 0, 0)  # Valores RGB para rojo
            formato_rojo = QTextCharFormat()
            formato_rojo.setForeground(color_rojo)
            output_textedit.mergeCurrentCharFormat(formato_rojo)
            output_textedit.insertPlainText("Credenciales no inicializadas.\n")
            formato_negro = QTextCharFormat()
            formato_negro.setForeground(QColor(0, 0, 0))
            output_textedit.mergeCurrentCharFormat(formato_negro)
            

def main():
    app = QApplication(sys.argv)
    ventana_principal = VentanaPrincipal()
    ventana_principal.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()